from argparse import ArgumentParser, RawTextHelpFormatter
import os
import openpyxl
import datetime
from bdpl_ingest import *
import pickle
import shutil
from lxml import etree
import glob

def set_up_mco_files(mco_prep_dir, unit_name, shipmentDate, current_batch):
    
    mco_spreadsheet = os.path.join(mco_prep_dir, '{}_{}_MCO_deposit_batch_{}.xlsx'.format(unit_name, shipmentDate, str(current_batch).zfill(2)))
    
    #If MCO spreadsheet doesn't exist, create it and add header rows
    if not os.path.exists(mco_spreadsheet):
        mco_wb = openpyxl.Workbook()
        mco_ws = mco_wb.active
        
        deposit_date = datetime.datetime.today().strftime('%Y-%m-%d')
        
        reference_info = ['BDPL deposit to MCO: {} shipment {}, batch {}'.format(unit_name, shipmentDate, str(current_batch).zfill(2)), 'micshall@iu.edu']
        mco_ws.append(reference_info)
        
        mco_header = ['Other Identifier', 'Other Identifier Type', 'Other Identifier', 'Other Identifier Type', 'Other Identifier', 'Other Identifier Type', 'Title', 'Creator', 'Date Issued', 'Abstract', 'Physical Description', 'Publish', 'File', 'Label']
        mco_ws.append(mco_header)
        
        mco_wb.save(mco_spreadsheet)
        
    #if MCO spreadsheet does exist, open and assign variables.
    else:
        mco_wb = openpyxl.load_workbook(mco_spreadsheet)
        mco_ws = mco_wb['Sheet']
        
    batch_copy_list = os.path.join(mco_prep_dir, 'batch_copy_list_{}.txt'.format(str(current_batch).zfill(2)))
        
    return mco_spreadsheet, mco_wb, mco_ws, batch_copy_list

def prep_content(unit_name, shipmentDate, folders):
    
    #set variables
    ship_dir = folders['ship_dir']
    unit_home = folders['unit_home']  
    mco_prep_dir = os.path.join(ship_dir, 'mco_prep') 
    
    if not os.path.exists(mco_prep_dir):
        os.makedirs(mco_prep_dir)
        
    batch_size = 50
    
    #get shipment spreadsheet and open with openpyxl
    spreadsheet = find_spreadsheet(folders, unit_name, shipmentDate)
    wb = openpyxl.load_workbook(spreadsheet)
    app_ws = wb['Appraisal']
    
    '''Give user opportunity to add other file extensions'''
    format_lists = os.path.join(mco_prep_dir, 'mco_formats.txt')
    if os.path.exists(format_lists):
        with open(format_lists, 'rb') as f:
            audio_formats, video_formats = pickle.load(f)
    else:
        audio_formats = ['.wav']
        video_formats = ['.mpg', '.mkv']
        
    print('\nThis script will prepare batches of up to {} BDPL items for deposit to Media Collections Online.  Large shipments will require running this script multiple times.\n\nNOTE: all content is assumed to be from optical media.  Adjust the "PhysicalDescription" variable if other content sources are used.'.format(batch_size))
    
    print('\nThis script will look for files with the following extensions:\n\tAudio: {}\n\tVideo: {}'.format(', '.join(audio_formats), ', '.join(video_formats)))
    
    print('\nAdd additional file extensions?\n\tA: add AUDIO formats\n\tV: add VIDEO formats\n\tR: RESET to default formats (.wav and .mpg /.mkv)\n\tC: skip adding formats and CONTINUE')
    
    while True:
        add_formats = input('\nEnter selection (A, V, R, or C): ')
        
        if add_formats.upper() == 'C':
            break
            
        elif add_formats.upper() == 'A':
            new_audio = input('Enter exact audio file format extension (with "."): ')
            audio_formats.append(new_audio.lower())
            
        elif add_formats.upper() == 'R':
            audio_formats = ['.wav']
            video_formats = ['.mpg', '.mkv']
        
        elif add_formats.upper() == 'V':
            new_video = input('Enter exact video file format extension (with "."): ')
            video_formats.append(new_video.lower())
        
        else:
            continue
  
    #save lists in case process is interrupted
    with open(format_lists, 'wb') as f:
        pickle.dump((audio_formats, video_formats), f)
        
    #check if we've already run any batches.  Set up a dictionary to store info; we'll also pickle this dict to file in case we have to stop prepping and resume later.
    current_batch = 1
    current_item_count = 0
    batch_info = {}
    
    batch_info_file = os.path.join(mco_prep_dir, 'batch_info.txt')
    if os.path.exists(batch_info_file):
        with open(batch_info_file, 'rb') as f:
            batch_info = pickle.load(f)
        
            if len(batch_info) > 0:
                current_batch = len(batch_info)
                
                current_item_count = len(batch_info[current_batch])
                
                if current_item_count == batch_size:
                    current_item_count = 0
                    current_batch += 1

    #set up MCO deposit spreadsheet
    mco_spreadsheet, mco_wb, mco_ws, batch_copy_list = set_up_mco_files(mco_prep_dir, unit_name, shipmentDate, current_batch)
    
    '''loop through each barcode in appraisal spreadsheet'''
    for barcode in app_ws['A'][1:]:
        if not barcode is None:
        
            master_list = [item for sublist in list(batch_info.values()) for item in sublist]
            
            #set key variables based on barcode
            item_barcode = barcode.value
            folders = bdpl_folders(unit_name, shipmentDate, item_barcode)
            destination = folders['destination']
            files_dir = folders['files_dir']
            image_dir = folders['image_dir']
            
            if not os.path.exists(destination):
                #print('\tBarcode folder does not exist')
                continue

            #check to see if item was already completed
            if item_barcode in master_list:
                print('\tItem already completed')
                continue
            
            print('\nCurrent item: {}'.format(item_barcode))
            
            #get metadata for barcode and pickle it
            print('\n\tGathering metadata for barcode...')
            if not load_metadata(folders, item_barcode, spreadsheet):
                continue
            
            #get pickled metadata so we can use it
            metadata_dict = pickleLoad('metadata_dict', folders, item_barcode)
            
            #if content goes to MCO, make sure we are not exceeding our batch size limit
            if 'mco' in metadata_dict['initial_appraisal'].lower():
                
                #if we've hit our batch size limit, start a new batch and reset file count. Also, start a new spreadsheet.
                if current_item_count == batch_size:
                    current_batch += 1
                    current_item_count = 0
                    
                    #set up new spreadsheet if we've started a new batch
                    mco_spreadsheet, mco_wb, mco_ws, batch_copy_list = set_up_mco_files(mco_prep_dir, unit_name, shipmentDate, current_batch)
                    
                current_item_count += 1
                
            else:
                print('\tDo not deposit to MCO')
                continue
            
            '''Some metadata fields need to be compiled and/or selected from alternatives'''
            #check for title:
            if metadata_dict.get('item_title') and metadata_dict.get('item_title') not in ['', '-', 'N/A']:
                item_title = metadata_dict['item_title']
            else:
                item_title = metadata_dict['label_transcription']
                
            #check for a description
            item_description = metadata_dict.get('item_description', '')
                
            #check for dates
            if metadata_dict.get('assigned_dates') and metadata_dict.get('assigned_dates') not in ['', '-', 'N/A']:
                date_issued = metadata_dict['assigned_dates']
            else:
                if metadata_dict['begin_date'] == metadata_dict['end_date']:
                    try:
                        date_issued = metadata_dict['begin_date'].replace('undated', '')
                    except AttributeError:
                        date_issued = metadata_dict['begin_date']
                else:
                    date_issued = "{}/{}".format(metadata_dict['begin_date'], metadata_dict['end_date'])
                
            #check if content came from a DVD or CD (by job type).  NOTE: we are currently assuming all content comes from optical media
            phys_descr = 'Optical disc'
            
            # if metadata_dict.get('jobType') and metadata_dict.get('jobType') in ['DVD', 'CDDA']:
                # phys_descr = 'Optical disc'
            # else:
                # phys_descr = ''
            
            '''add metadata to MCO dictionary'''
            mco_metadata = {'BDPLID' : item_barcode, 
                            'ID_Type_3' : 'BDPL ID', 
                            'CollectionID' : metadata_dict['current_coll_id'], 
                            'ID_Type_1' : 'Collection ID',  
                            'AccessionID' : metadata_dict['current_accession'], 
                            'ID_Type_2' : 'Accession ID', 
                            'Title' : item_title, 
                            'Creator' : metadata_dict['collection_creator'], 
                            'DateIssued' : date_issued, 
                            'Abstract' : item_description, 
                            'PhysicalDescription' : phys_descr, 
                            'Publish' : 'No'}
            
            #try to clear out any bad data
            for k, v in mco_metadata.items():
                if v in ['-', 'N/A', ' ']:
                    mco_metadata[k] = ''
                    
            '''Now look for our files'''
            
            audio_file_list = [f for f in os.listdir(files_dir) if os.path.splitext(f)[-1].lower() in audio_formats]
            cue_file_list = [f for f in os.listdir(files_dir) if os.path.splitext(f)[-1].lower() == '.cue']
            video_file_list = [f for f in os.listdir(files_dir) if os.path.splitext(f)[-1].lower() in video_formats]
            
            for ls in (audio_file_list, video_file_list):
                print(ls)
                
                #if list is empty, continue.
                if len(ls) == 0:
                    continue
                
                #establish type of content. NOTE: may need to change labels, based upon content source...
                if ls == audio_file_list:
                    label = 'CD'
                else:
                    label = 'DVD'
                
                for i in range(0, len(ls)):
                    
                    mco_file = ls[i]
                    
                    print('\n\tWorking on {}'.format(mco_file))
                    
                    #get current count of 'file' fields in MCO spreadshet
                    column_count = {i : cell.value for i, cell in enumerate(mco_ws[2], 1) if cell.value == 'File'}
                    
                    #assign values to mco metadata dictionary
                    mco_metadata['File_{}'.format(i)] = mco_file
                    mco_metadata['Label_{}'.format(i)] = '{} part {}'.format(label, i+1)
                    
                    #if we exceed current # of 'File'/'Label' fields in the spreadsheet, we need to add a new one
                    if i+1 > len(column_count):
                        #new 'file' column will be two over from the last one
                        current_max = max(column_count.keys()) + 2
                        mco_ws.insert_cols(current_max)
                        mco_ws.cell(row=2, column=current_max, value='File')
                        
                        #add new 'label' column
                        current_max += 1
                        mco_ws.insert_cols(current_max)
                        mco_ws.cell(row=2, column=current_max, value='Label')
                    
                    #write file to batch copy list
                    target = os.path.join(files_dir, mco_file)
                    print('\t\tWriting file to copy list...')
                    with open(batch_copy_list, 'ab') as f:
                        f.write('{}\n'.format(target).encode())
                    
                    #for audio files with CUE: create structure.xml file
                    if label == 'CD':
                        found_cue = [c for c in cue_file_list if os.path.splitext(c)[0] == os.path.splitext(mco_file)[0]]
                        
                        #if we've found a wav_cue_file file, convert to structure.xml
                        if found_cue:
                            
                            #old files: make sure we note it's an optical disk...
                            if mco_metadata['PhysicalDescription'] == '':
                                mco_metadata['PhysicalDescription'] = 'Optical disc'
                            
                            cue_file = os.path.join(files_dir, found_cue[0])
                            
                            print('\t\tCreating structure.xml file for audio...', cue_file)
                            
                            #get info from wav_cue_file file.  NOTE: old procedures resulted in an encoding issue and do not reference the WAV file; let's try to fix those!
                            while True:
                                try:
                                    with open(cue_file, 'r') as f:
                                        cue_contents = f.read().splitlines()
                                        
                                    if 'BINARY' in cue_contents[0]:
                                        fix_cue(files_dir, cue_file)
                                    elif 'WAVE' in cue_contents[0]:
                                        break
                                
                                #if we get UnicodeDecodeError, we should fix cue files in both files_dir and image_dir
                                except UnicodeDecodeError:
                                    #fix the wav cue file
                                    fix_cue(files_dir, cue_file)
                                    
                                    #grab bin cue(s) and fix them, too.
                                    bin_cues = glob.glob(os.path.join(image_dir, '*.cue'))
                                    for bc in bin_cues:
                                        fix_cue(image_dir, bc)
                            
                            #pull out tracks and time indices
                            tracks = [c.strip().replace(' AUDIO', '').capitalize() for c in cue_contents if "TRACK" in c]
                            times = [':'.join(c.split()[2].split(':', 2)[:2]) for c in cue_contents if "INDEX 01" in c]
                            
                            #set up dictionary to store track information
                            track_info = {}
                            
                            #loop through our list of tracks; each track should have a corresponding INDEX 01 timestamp
                            for i in range(0, len(tracks)):
                                
                                #get the begin time.  
                                if times[i] == '00:00':
                                    begin_time = '0'
                                else:
                                    begin_time = times[i]
                                
                                #get end time; final track does not need one
                                if times[i] == times[-1]:
                                    track_info[tracks[i]] = {'begin':begin_time}
                                else:
                                    end_time = calc_end_time(times[i+1])
                                    track_info[tracks[i]] = {'begin':begin_time, 'end':end_time}
                                
                            #start creating our structure xml doc with lxml
                            structure_xml = os.path.join(mco_prep_dir, '{}.structure.xml'.format(mco_file))
                            
                            item = etree.Element('item')
                            item.attrib['label'] = mco_metadata.get('Title', 'Audio recording').replace('"', "'").replace('&', 'and').strip()
                            
                            #loop through our track info to pull in info for the individual 'spans'
                            for track in track_info.keys():
                                span = etree.SubElement(item, 'span')
                                span.attrib['label'] = track
                                span.attrib['begin'] = track_info[track]['begin']
                                #only include 'end' attribute if we have an end time
                                if track_info[track].get('end'):
                                    span.attrib['end'] = track_info[track]['end']
                            
                            #write etree to file in mco assets folder
                            structure_tree = etree.ElementTree(item)
                            structure_tree.write(structure_xml, pretty_print=True)
                            
                            #write file to copy list
                            with open(batch_copy_list, 'ab') as f:
                                f.write('{}\n'.format(structure_xml).encode())
                                
            #Document progress; store info to batch_info dict and pickle that as well (just in case).  Also save to spreadsheet.
            if batch_info.get(current_batch):
                batch_info[current_batch].append(item_barcode)
            else:
                batch_info[current_batch] = [item_barcode]
                
            with open(batch_info_file, 'wb') as f:
                pickle.dump(batch_info, f)

            '''write info for this barcode to mco spreadsheet'''
            mco_ws.append(list(mco_metadata.values()))
            mco_wb.save(mco_spreadsheet)    
    
    print('\n\n-----------------------------------------------------------------------------\n\nMCO preparation complete. Run "move" operation after review of MCO spreadsheet(s).')
        
def calc_end_time(timestamp):
    
    min, sec = timestamp.split(':')
    
    if sec == '00':
        sec = 59
        min = int(min) - 1
    else:
        sec = int(sec) - 1
    
    return "{}:{}".format(str(min).zfill(2), str(sec).zfill(2))

def fix_cue(folder, cue_file):
    
    print('\tFixing {}...'.format(cue_file))
    
    #get audio file
    if 'disk-image' in folder:
        audio = '{}.bin'.format(os.path.basename(os.path.splitext(cue_file)[0]))
        type = 'BINARY'
    else: 
        audio = '{}.wav'.format(os.path.basename(os.path.splitext(cue_file)[0]))
        type = 'WAVE'
    
    #get info from cue file
    with open(cue_file, 'rb') as infile:
        cue_info = infile.readlines()[1:]
    
    #write correct 1st line
    with open(cue_file, 'w') as outfile:
        if type == 'BINARY':
            outfile.write('FILE "{}" BINARY\n'.format(audio))
        else:
            outfile.write('FILE "{}" WAVE\n'.format(audio))
    
    #append rest of cue info
    with open(cue_file, 'ab') as outfile:
        for line in cue_info:
            outfile.write(line) 

def move_content(unit_name, shipmentDate, folders, mco_dropbox):
    
    print('\nThis script will move individual batches of content to {} for deposit to Media Collections Online(MCO).\n\n\tNOTE: Make sure that the previous batch has been uploaded to MCO and then delete files before moving next batch.')
    
    #set variables
    ship_dir = folders['ship_dir']
    unit_home = folders['unit_home']  
    mco_prep_dir = os.path.join(ship_dir, 'mco_prep') 
    batch_size = 50
    
    #set up destination for files in dropbox
    mco_assets = os.path.join(mco_dropbox, 'assets')
    if not os.path.exists(mco_assets):
        os.makedirs(mco_assets)
    
    #get info about batches
    batch_info_file = os.path.join(mco_prep_dir, 'batch_info.txt')
    if os.path.exists(batch_info_file):
        with open(batch_info_file, 'rb') as f:
            batch_info = pickle.load(f)
        
    #get list of all batches and check for log of completed batches
    batches = list(batch_info.keys())
    
    moved_batches = os.path.join(mco_prep_dir, 'moved_batches.txt')
        
    while True:
        
        #remove any completed batches from our list
        if os.path.exists(moved_batches):
            with open(moved_batches, 'r') as f:
                for m in f.read().splitlines():
                    if int(m) in batches:
                        batches.remove(m)
                    
        if len(batches) > 0:
            print('\nThe following batch(es) need to be moved to {}:\n\t{}'.format(mco_dropbox, '\n\t'.join(str(b) for b in batches)))
        
            option = input('\nEnter the batch number that will be moved to the MCO dropbox or Q to quit: ')
            
            try:
                if int(option) in batches:
                    current_batch = option
                    break
                else:
                    print('\nResponse not recognized; re-enter option.\n\n')
                    
            except ValueError:
                if option.lower() == 'q':
                    return
                else:
                    print('\nResponse not recognized; re-enter option.\n\n')
        
        else:
            print('\nAll batches have been moved to {}. Check with digital preservation librarian if you believe this is an error.'.format(mco_dropbox))
            return
            
           
    #get spreadsheet and copy list for this batch
    mco_spreadsheet, mco_wb, mco_ws, batch_copy_list = set_up_mco_files(mco_prep_dir, unit_name, shipmentDate, current_batch)
    
    #set up failed list...
    failed_copies = os.path.join(mco_prep_dir, 'failed_copies_{}.txt'.format(str(current_batch).zfill(2)))
    failed = False
    
    #check to see if there are already files in the destination...
    if checkFiles(mco_assets):
        print('\n\nNOTE: Destination currently has files.  Proceed with copy operation?')
        
        while True:
            option = input ('(Y or N): ')
        
            if option.lower() == 'y':
                break
            elif option.lower() == 'n':
                return
            else:
                continue
    
    #now loop through list and copy files to assets folder
    with open(batch_copy_list, 'r') as f:
        for file in f.read().splitlines():
            if os.path.exists(file):
                shutil.copy(file, mco_assets)
            else:
                with open(failed_copies, 'a') as f:
                    f.write('{}\n'.format(file))
                failed = True
                
    #if no failures, copy over spreadsheet
    if not failed:
        shutil.copy(mco_spreadsheet, mco_dropbox)
        
        print('\n\nBatch successfully copied to MCO dropbox.  Before moving any additional batches, make sure materials are ingested by MCO and remove files from dropbox.')
        
        #write to completed list.
        with open(moved_batches, 'a') as f:
            f.write('{}\n'.format(str(current_batch)))
        
    else:
        print('\n\nNOTE: One or more files failed to be copied to MCO dropbox; see {} for more details.'.format(failed_copies))
        
    return
    
def main():
    #clear screen
    newscreen()
    
    #set up arg parser so user is required to add shipment directory
    parser = ArgumentParser(
        description='This script has two options:\n\t(a) Prepare specified content for deposit to Media Collections Online (MCO)\n\t(b) Move content from Scandium to the MCO dropbox.',
        formatter_class=RawTextHelpFormatter
    )

        #option 1: include unit and shipment date to package content; option 2: include unit, shipment date, and mco_path to move content.
    parser.add_argument(
        '-prep', 
        help='Prep metadata and files to deposit to MCO',
        action='store_true'
    )
    
    parser.add_argument(
        '-move', 
        help='Move files to MCO dropbox',
        action='store_true'
    )
    
    parser.add_argument(
        'unit_name',
        help='Unit abbreviation'
    )

    parser.add_argument(
        'shipmentDate',
        help='Shipment date'
    )
    
    parser.add_argument(
        '-mco', '--mco_dropbox',
        help='Path to MCO dropbox'
    )
    
    args = vars(parser.parse_args())
    
    #make sure only one action is selected
    if args['move'] == args['prep']:
        parser.error('\nScript can only be used to prep content OR to move files.  Make sure you have selected appropriate option.')
    
    #make sure that unit_name and shipmentDate are included
    if not all(k in args for k in ('unit_name', 'shipmentDate')):
        parser.error('\n\nScript requires a valid unit abbreviation and associated shipment date, and MCO dropbox location.')
    
    #if unit_name and shipmentDate are present, make sure they are valid
    else:
            
        #set other variables
        unit_name = args['unit_name']
        shipmentDate = args['shipmentDate']
        folders = bdpl_folders(unit_name, shipmentDate)
        ship_dir = folders['ship_dir']
        unit_home = folders['unit_home']        
    
        if not os.path.exists(unit_home):
            parser.error('\n\nScript requires a valid unit abbreviation; no directory exists for unit {}.'.format(folders['unit_home']))
        if not os.path.exists(ship_dir):
            parser.error('\nScript requires a valid shipment date for unit {}; shipment "{}" does not exist.'.format(unit_name, shipmentDate))
    
    #'Prep' option
    if args['prep']:
        prep_content(unit_name, shipmentDate, folders)
    
    #'Move' option:
    if args['move']:
        
        mco_dropbox = args['mco_dropbox']
        if not os.path.exists(mco_dropbox):
            print('\n')
            parser.error('\n\nScript requires a valid path to an MCO dropbox location; {} does not exist.'.format(mco_dropbox))
    
        move_content(unit_name, shipmentDate, folders, mco_dropbox)
  
if __name__ == "__main__":
    main()