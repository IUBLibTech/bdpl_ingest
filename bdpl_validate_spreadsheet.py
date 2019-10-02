import shutil
import os
import openpyxl
from collections import Counter

def main():
    while True:
        spreadsheet = input('\nEnter Python-appropriate path to the spreadsheet: ')
    
        if os.path.exists(spreadsheet):
            break
    
    wb = openpyxl.load_workbook(spreadsheet)
    ws = wb['Inventory']
    
    print('\nOptions:\n\t 1 - Validate BDPL spreadsheet submitted by unit\n\t 2 - Validate barcodes on RipStation "userdata.txt" file')
    
    while True:
        option = input('\nEnter your desired option (1 /2): ')
        if option.strip() == '1':
            validate_spreadsheet(wb, ws, spreadsheet)
            break
        elif option.strip() == '2':
            validate_userdata(wb, ws, spreadsheet)
            break
        else:
            continue

def validate_spreadsheet(wb, ws, spreadsheet):
    print('\nNOTE: validating spreadsheet against template version 20190724; update script if new template is in use.')
    
    '''make sure correct column headers are in use'''
    
    #current column headers--UPDATE IF NEEDED!
    template_headers = ['Identifier', 'Accession ID (if known)', 'Collection Title (if assigned)', 'Collection ID (if assigned)', 'Creator (if known)', 'Physical location of media', 'Source type (select from drop down or type)', 'Label transcription', 'Initial appraisal notes (including potential sensitive data)', 'Instructions for BDPL staff ', 'Restriction Statement', 'Restriction end date (YYYY-MM-DD)', 'Move directly to SDA without appraisal? ']
    
    new_headers = []
    for cell in ws[1]:
        if not cell.value is None:
            new_headers.append(cell.value)
    
    column_list = 'ABCDEFGHIJKLMNOP'
    
    #compare current headers vs. template
    errors = 0
    for i in range(0, len(template_headers)):
        try:
            if new_headers[i] != template_headers[i]:
                print('\n\nWARNING: Column %s has incorrect header!\n\tHeader is "%s"; SHOULD be "%s"' % (column_list[i], new_headers[i], template_headers[i]))
                errors += 1
        except IndexError as e:
            print('\n\nWARNING: %s not included.' % template_headers[i])
            errors += 1
    
    if errors == 0:
        print('\n\nNo errors with spreadsheet headers!')
    
    new_bcs = {}    
    missing_barcodes = []
    
    for barcode in ws['A'][1:]:
        if not barcode.value is None:
            new_bcs[barcode.value] = barcode.row
        else:
            missing_barcodes.append(barcode.row)
    
    #note if any barcode values are missing
    if len(missing_barcodes) > 0:
        print('\n\n\nWARNING: the following rows are missing barcodes; verify if any action is needed:')
        for row in missing_barcodes:
            print('\tRow: %s' % row)
    else:
        print('\n\nNo Missing barcodes!')
    
    new_bcs_list = list(new_bcs.keys())
    
    #check for any duplicate barcodes in new spreadsheet
    duplicate_barcodes = [item for item, count in Counter(new_bcs_list).items() if count > 1]
    
    #note if any barcodes in the spreadsheet are duplicates
    if len(duplicate_barcodes) > 0:
        print('\n\nWARNING: spreadsheet includes duplicate barcode values:')
        for dup in duplicate_barcodes:
            print('\t%s\tRow: %s' % (dup, new_bcs[dup]))
    else:
        print('\n\nNo duplicte barcodes in spreadsheet.')
    
    #make a copy of the master workbook
    master_spreadsheet = 'Y:/spreadsheets/bdpl_master_spreadsheet.xlsx'
    master_copy = os.path.join('C:/temp', 'bdpl_master_copy.xlsx')
    
    if not os.path.exists('C:/temp'):
        os.mkdir('C:/temp')
    
    try:
        shutil.copy(master_spreadsheet, master_copy)
    except FileNotFoundError:
        print('\n\nUnable to access master spreadsheet at ', master_spreadsheet)
    
    #add all current barcodes into a list
    try:
        master_wb = openpyxl.load_workbook(master_copy)
        item_ws = master_wb['Item']

        master_list = []
        
        for barcode in item_ws['A'][1:]:
            if not barcode.value is None:
                master_list.append(barcode.value)
        
        already_used = [x for x in new_bcs_list if x in master_list]
        
        #note if any barcodes are already in the SDA are duplicated
        if len(already_used) > 0:
            print('\n\nWARNING: spreadsheet includes barcodes that have already been deposited to the SDA:')
            for dup in already_used:
                print('\t%s\tRow: %s' % (dup, new_bcs[dup]))
    except FileNotFoundError:
        print('\n\nUnable to access copy of master spreadsheet at ', master_copy)

def validate_userdata(wb, ws, spreadsheet):
    while True:
        userdata = input('\nEnter Python-appropriate path to "userdata.txt": ')
    
        if os.path.exists(userdata):
            break
    
    #get full list of barcodes in userdata.txt
    with open(userdata, 'r') as ud:
        ud_list = ud.read().splitlines()
    
    #get list of barcodes in inventory
    barcodes = []
    iterrows = ws.iter_rows()
    
    next(iterrows)
    
    for row in iterrows:
        if not row[0].value is None:
            barcodes.append(str(row[0].value))
            
    bad_barcodes = [bc for bc in ud_list if not bc in barcodes]
    
    if len(bad_barcodes) == 0:
        print('\nAll barcodes in userdata.txt are located in', spreadsheet)
    else:
        print('\nWARNING: the following barcodes are not listed in %s:' % spreadsheet)
        print('\n\t%s' % '\n\t'.join(bad_barcodes))

if __name__ == '__main__':

    os.system('cls')
    
    #print BDPL screen
    fname = "C:/BDPL/scripts/bdpl.txt"
    if os.path.exists(fname):
        with open(fname, 'r') as fin:
            print(fin.read())

    main()