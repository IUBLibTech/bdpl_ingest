import subprocess
import os
import shutil
import bagit
import sys
import openpyxl
import hashlib
import datetime
import pickle
import glob
import csv
from lxml import etree
import uuid

from bdpl_ingest import *
from bdpl_ripstation_ingest import *



def write_list(list_name, message):
    with open(list_name, 'a') as f:
        f.write(message)
    
def check_list(list_name, item_barcode):
    if not os.path.exists(list_name):
        return False
    with open(list_name, 'r') as f:
        for item in f:
            if item_barcode in item.strip():
                return True
            else:
                continue
        return False

def get_size(start_path):
    total_size = 0
    if os.path.isfile(start_path):
        total_size = os.path.getsize(start_path)
    else:
        for dirpath, dirnames, filenames in os.walk(start_path):
            for f in filenames:
                fp = os.path.join(dirpath, f)
                # skip if it is symbolic link
                if not os.path.islink(fp):
                    total_size += os.path.getsize(fp)
    return total_size


def raw_text(text):
    """Returns a raw string representation of text"""
    
    escape_dict={'\a':r'\a',
           '\b':r'\b',
           '\c':r'\c',
           '\f':r'\f',
           '\n':r'\n',
           '\r':r'\r',
           '\t':r'\t',
           '\v':r'\v',
           '\'':r'\'',
           '\"':r'\"',
           '\0':r'\0',
           '\1':r'\1',
           '\2':r'\2',
           '\3':r'\3',
           '\4':r'\4',
           '\5':r'\5',
           '\6':r'\6',
           '\7':r'\7',
           '\8':r'\8',
           '\9':r'\9'}
    
    new_string=''
    for char in text:
        try: 
            new_string+=escape_dict[char]
        except KeyError: 
            new_string+=char
    return new_string

def separate_content(sep_dest, file, log, bag_report_dir, item_barcode):
    #We will store results for this file in a temp file for the whole barcode.
    separated_file_stats = os.path.join(bag_report_dir, '%s-separation-stats.txt' % item_barcode)
    temp_list = []
    
    #if this temp file exists, retrieve it and check if this file has already been moved.
    if os.path.exists(separated_file_stats):
        with open(separated_file_stats, 'rb') as f:
            temp_list = pickle.load(f)    
        
        #if we already have a list of separated files, check to see if file has already been moved
        for f in temp_list:
            if f['file'] == file:
                #if file previously failed, remove it from the list.
                if f['result'] != 'Moved':
                    temp_list.remove(f)
                    break
                #if the file was moved, return to main() and go on to next one.
                else:
                    return
    
    print('\t\t%s' % file)
    
    #create folder structure, if needed
    if not os.path.exists(os.path.dirname(sep_dest)):
        os.makedirs(os.path.dirname(sep_dest))
    
    #get separation stats: size 
    size = get_size(file)
    
    mod_date = datetime.datetime.fromtimestamp(os.path.getmtime(file)).isoformat()
    
    #get format puid using siegrfried
    try:
        cmd = 'sf -csv "%s"' % file
        #just pull out the puid from siegrfried output
        puid = subprocess.check_output(cmd, text=True).split('\n')[1].split(',')[5]
    except subprocess.CalledProcessError as e:
        pass
    
    #check if file is a disk image or extracted file
    if 'disk-image' in file:
        type = 'disk-image'
    else:
        type = 'extracted-file'
    
    try:
        #remove read only attribute, if necessary
        # cmd = 'ATTRIB -r "%s"' % file
        # subprocess.call(cmd, shell=True)
        
        #now move file
        shutil.move(file, sep_dest)
        with open(log, 'a') as f:
            f.write('%s\t%s\t%s\n' % (file, size, mod_date))
        result = 'Moved'
            
    except (shutil.Error, OSError, IOError, PermissionError) as e:
        #print('\n\tError separating %s: %s' % (file, e))
        result = e

    #add file to our temp list, noting result of operation and file size.  Write temp list to file.
    temp_list.append({'barcode' : item_barcode, 'file' : file, 'size' : size, 'result' : result, 'type' : type, 'puid' : puid})
    
    with open(separated_file_stats, 'wb') as f:
        pickle.dump(temp_list, f)

def return_spreadsheet_value(ws, current_row, current_column):

    return ws.cell(row=current_row, column=current_column).value

def main():
    
    '''SET VARIABLES'''
    #Identify where files will be moved ###### UPDATE TO ARCHIVER location #######
    archiver_drop_off = 'W:/Archiver_spool/general%2fmediaimages'
    
    if not os.path.exists(archiver_drop_off):
        print('WARNING: incorrect server shared mapped to W: drive.  Reconnect if necessary.')
        sys.exit(1)
    
    #open master workbook and get ready to write
    master_spreadsheet = 'W:/spreadsheets/bdpl_master_spreadsheet.xlsx'
    master_wb = openpyxl.load_workbook(master_spreadsheet)
    item_ws = master_wb['Item']
    cumulative_ws = master_wb['Cumulative']
    
    print('\nNOTE: Archiver drop off location is at %s; master spreadsheet is at %s' % (archiver_drop_off, master_spreadsheet))
    
    #get unit name and shipment folder
    while True:
        unit_name = input('\nEnter unit abbreviation: ')
        
        shipmentDate = input('\nEnter shipment date: ')
        
        ship_dir = os.path.join('Z:\\', unit_name, 'ingest', shipmentDate)
        
        if os.path.exists(ship_dir):
            break
        else:
            print('\nWARNING: cannot locate folder %s.  Check unit name and shipment date.' % ship_dir)
            continue
            
    packaging_info = os.path.join(ship_dir, 'packaging-info.txt')
    
    #if we've already worked on this shipment, check for info file; if found, assign variables.
    if os.path.exists(packaging_info):
        info = {}
        with open(packaging_info, 'rb') as f:
            info = pickle.load(f)
        
        unit_name = info['unit'] 
        spreadsheet = info['spreadsheet']
        separations_manifest = info['separations_manifest']
    
    else:
        spreadsheet = os.path.join(ship_dir, '%s_%s.xlsx' % (unit_name, shipmentDate))
        if not os.path.exists(spreadsheet):
            print('\nWARNING: Could not locate %s in shipment folder.  Be sure spreadsheet is present and correctly named (%s_%s.xlsx) and then run script again.' % (spreadsheet, unit_name, shipmentDate))
            sys.exit(1)

        sep_check = input('\nDoes shipment have any content to be separated/deaccessioned? (y/n) ')
                
        if sep_check.lower().strip() == 'y':           
                        
            separations_manifest = os.path.join(ship_dir, 'separations.txt')
                        
            if not os.path.exists(separations_manifest):
                print('\nWARNING: cannot identify separations manifest.  Please check directory to make sure separations.txt is present.')
                sys.exit(1)

        else:
            separations_manifest = ''
                    
    
        info = {'unit' : unit_name, 'spreadsheet' : spreadsheet, 'separations_manifest' : separations_manifest}

        with open(packaging_info, 'wb') as f:
            pickle.dump(info, f)
        
    #set shipment directory as current working directory
    os.chdir(ship_dir) 
     
    #open shipment workbook
    wb = openpyxl.load_workbook(spreadsheet)
    ws_app = wb['Appraisal']
    
    #folders/files for tracking status
    bag_report_dir = os.path.join(ship_dir, 'bag_reports')
    deaccession_dir = os.path.join(ship_dir, 'deaccessioned')
    unaccounted_dir = os.path.join(ship_dir, 'unaccounted')

    started_list = os.path.join(bag_report_dir, 'started.txt') #to document all the barcodes that we started to package/process; all should be accounted for in the fail/complete lists
    deaccession_list = os.path.join(bag_report_dir, 'deaccession.txt')    #for barcodes/items that will not be sent to SDA
    other_list = os.path.join(bag_report_dir, 'other-decision.txt') #for barcode folders have an alternate appraisal decision     
    failed_list = os.path.join(bag_report_dir, 'failed-packaging.txt') #for any failures; will note which stage the failure occurred 

    prep_list = os.path.join(bag_report_dir, 'info_prepped.txt') #for barcodes that were successfully bagged
    bagged_list = os.path.join(bag_report_dir, 'bagged.txt') #for barcodes that were successfully bagged
    tarred_list = os.path.join(bag_report_dir, 'tarred.txt') #for barcodes that were successfully tarred
    completed_list = os.path.join(bag_report_dir, 'completed.txt') #for barcode folders that were successfully cleaned and SIP creation completed

    moved_list = os.path.join(bag_report_dir, 'moved.txt') #for barcodes that reached the end of the process; should include any that were deaccessioned
    metadata_list = os.path.join(bag_report_dir, 'metadata.txt') #for barcodes that have metadata written to spreadsheet
    separated_list = os.path.join(bag_report_dir, 'separated-content.txt') #for barcodes that have undergone separations
    unaccounted_list = os.path.join(bag_report_dir, 'unaccounted.txt') #for barcodes that are in directory, but not in spreadsheet; need to check for data entry errors
    format_report = os.path.join(bag_report_dir, 'cumulative-formats.txt') # for tracking information on file formats
    puid_report = os.path.join(bag_report_dir, 'puid-report.txt') # list of all puids in shipment
    duration_doc = os.path.join(bag_report_dir, 'duration.txt')
    missing_doc = os.path.join(bag_report_dir, 'missing.txt')
    stats_doc = os.path.join(bag_report_dir, 'shipment_stats.txt')    
    '''SET UP: GATHER INITIAL STATS AND CHECK FOR INCONSISTENCIES WITH BARCODE FOLDERS IN SHIPMENT'''
    #make our report directory
    if not os.path.exists(bag_report_dir):
        os.mkdir(bag_report_dir)
    
    #get list of directories in our shipment folder; make sure these are folders and do not include any folders created during bagging process
    dir_list = [d for d in os.listdir(ship_dir) if os.path.isdir(os.path.join(ship_dir, d)) and not d in ['review', 'bag_reports', 'unaccounted', 'deaccessioned']]

    #get list of barcodes from spreadsheet
    spreadsheet_list = []
    for col in ws_app['A'][1:]:
        if not col.value is None:
            spreadsheet_list.append(str(col.value))
    
    #see if any barcode folders are missing
    missing_from_dir = list(set(spreadsheet_list) - set(dir_list))
    
    #don't include completed items in missing count    
    if os.path.exists(completed_list):
        for item in missing_from_dir:
            if check_list(completed_list, item):
                missing_from_dir.remove(item)  
    
    #check if there are any folders in the shipment NOT in spreadsheet
    missing_from_spreadsheet = list(set(dir_list) - set(spreadsheet_list))
    
    #If we have unaccounted barcodes; save list to file and move the dirs themselves to an 'unaccounted' folder
    if len(missing_from_spreadsheet) > 0:
        
        if not os.path.exists(unaccounted_dir):
            os.mkdir(unaccounted_dir)   
            
        with open(unaccounted_list, 'a') as f:
            for item in missing_from_spreadsheet:
                f.write('%s\n' % item)
                try:
                    shutil.move(item, unaccounted_dir)
                except (PermissionError, OSError) as e:
                    write_list(failed_list, '%s\tmove_unaccounted\t%s' % (item_barcode, e))

    #Get date info on items for shipment stats (acquired by using max/min of dir_list, using modified date as key)
    latest_date = datetime.datetime.fromtimestamp(os.stat(max(dir_list, key=os.path.getmtime)).st_ctime).strftime('%Y%m%d')
    earliest_date = datetime.datetime.fromtimestamp(os.stat(min(dir_list, key=os.path.getmtime)).st_ctime).strftime('%Y%m%d')        
    
    duration_stats = {}
    if os.path.exists(duration_doc):
        with open(duration_doc, 'wb') as file:
            duration_stats = pickle.load(file)
            
        if earliest_date < duration_stats['earliest']:
            duration_stats['earliest'] = earliest_date
            
        if latest_date > duration_stats['latest']:
            duration_stats['latest'] = latest_date
    
    else:
        duration_stats['earliest'] = earliest_date
        duration_stats['latest'] = latest_date

    tdelta = datetime.datetime.strptime(duration_stats['latest'], '%Y%m%d') - datetime.datetime.strptime(duration_stats['earliest'], '%Y%m%d')
    
    #use 1 day as minimum timedelta
    if tdelta < datetime.timedelta(days=1):
        duration_stats['duration'] = 1
    else:
        duration_stats['duration'] = int(str(tdelta).split()[0])
    
    #write duration and 'missing' stats to file
    with open(duration_doc, 'wb') as file:
        pickle.dump(duration_stats, file)
    with open(missing_doc, 'wb') as file:
        pickle.dump(missing_from_dir), file)

    '''INITIATE PACKAGING'''
    #get total number of rows, + 1
    maxrow = ws_app.max_rows() + 1
    
    #get all of our current spreadsheet columns
    ws_columns = get_spreadsheet_columns(ws_app)
    
    #now loop through rows, skipping the headers...
    for item in dir_list:
        
        item_barcode = item.strip()
        
        status, current_row = return_spreadsheet_row(ws_app, item_barcode)
        
        if not status:
            print('\nUnable to locate barcode in spreadsheet! Moving on to next item...')
            write_list(unaccounted_list, item_barcode)
            try:
                shutil.move(item_barcode, unaccounted_dir)
            except (PermissionError, OSError) as e:
                write_list(failed_list, '%s\tmove_unaccounted\t%s' % (item_barcode, e))
            continue

        #skip to next barcode if current one has already finished workflow
        if check_list(completed_list, item_barcode):
            print('\n%s completed.' % item_barcode)
            continue
        
        #document that we've started working on this barcode
        if not check_list(started_list, item_barcode):
            write_list(started_list, item_barcode)
 
        print('\nWorking on item: %s' % item_barcode)    
        
        initial_appraisal = return_spreadsheet_value(ws_app, current_row, ws_columns['initial_appraisal'])
        
        #if content will not be moved to SDA, just skip folder for now and write to skipped and moved lists
        if initial_appraisal == "Delete content":
            if check_list(deaccession_list, item_barcode):
                print('\n\t%s has been moved to the "deaccession" folder.' % item_barcode)
            else:
                if not os.path.exists(deaccession_dir):
                    os.mkdir(deaccession_dir)
                try:
                    shutil.move(item_barcode, deaccession_dir)
                    print('\n\tContent will not be transferred to SDA.  Continuing with next item.')
                    write_list(deaccession_list, item_barcode)
                
                except (PermissionError, OSError) as e:
                    write_list(failed_list, '%s\tdeaccession\t%s' % (item_barcode, e))
                    
            continue
        
        #if content has been determined to be of value, complete prep workflow.
        elif initial_appraisal == "Transfer to SDA":

            '''CHECK THAT FOLDER EXISTS'''            
            folders = bdpl_folders(unit_name, shipmentDate, item_barcode)
            
            destination = folders['destination']
            files_dir = folders['files_dir']
            log_dir = folders['log_dir']
            imagefile = folders['imagefile']
            temp_dir = folders['temp_dir']
            reports_dir = folders['reports_dir']
            files_dir = folders['files_dir']
            image_dir = folders['image_dir']
            metadata = folders['metadata']
            
            #complete initial preparations; skip if we're returning to item
            if not check_list(prep_list, item_barcode):
                
                file_count = return_spreadsheet_value(ws_app, current_row, ws_columns['item_file_count'])
                
                #doublecheck if no file information has been reported on spreadsheet
                if file_count is None or file_count == 0:
                    #check for content in our image_dir and files_dir; if both are empty, fail barcode.
                    if not checkFiles(image_dir) and not checkFiles(files_dir):
                        write_list(failed_list, '%s\tcheck_folder\tNO CONTENT IN BARCODE FOLDER: CHANGE APPRAISAL DECISION?' % item_barcode)
                            continue
            
                #get file format info to include with master spreadsheet.  If format and puid lists exist, load and then loop through format report...
                format_list = []
                if os.path.exists(format_report):
                    with open(format_report, 'rb') as f:
                        format_list = pickle.load(f)
                        
                puid_list = []                    
                if os.path.exists(puid_report):
                    with open(puid_report, 'rb') as f:
                        puid_list = pickle.load(f)
                            
                #get file format/puid information for cumulative stats
                format_csv = os.path.join(report_dir, 'formatVersions.csv')
                if os.path.exists(format_csv):
                    temp_list = []
                    with open(format_csv, 'r') as fi:
                        fi = csv.reader(fi)
                        #skip header row
                        next(fi)
                        #loop through format csv; create a dictionary for each row 
                        for line in fi:
                            temp_dict = {}
                            temp_dict['puid'] = line[1]
                            temp_dict['format'] = line[0]
                            temp_dict['version'] = line[2]
                            temp_dict['count'] = int(line[3])
                            
                            #add temp dict to a temp list
                            temp_list.append(temp_dict)
                            
                            #add puids to a master list
                            puid_list.append(line[1])
                    
                        format_list.append({item_barcode : temp_list})
                    
                #and now write this back to file.
                with open(format_report, 'wb') as f:
                    pickle.dump(format_list, f)
                with open(puid_report, 'wb') as f:
                    pickle.dump(puid_list, f)
                
                #confirm prep has completed
                write_list(prep_list, item_barcode)
            
            '''REMOVE SEPARATED CONTENT AND TEMP FILES/FOLDERS'''
            if not check_list(separated_list, item_barcode):
                print('\n\tSeparating unnecessary files...\n')
                
                #remove bulk_extractor folder, if present, as well as reports used solely for appraisal/review
                for dir in ['bulk_extractor', 'temp']:
                    remove_dir = os.path.join(destination, dir)
                    if os.path.exists(remove_dir):
                        shutil.rmtree(remove_dir)
             
                for f in ["duplicates.csv", "errors.csv", "formats.csv", "formatVersions.csv", "mimetypes.csv", "unidentified.csv", "uniqueyears.csv", "years.csv", 'email_domain_histogram.txt', 'find_histogram.txt', 'telephone_histogram.txt', 'report.html']:
                    report = os.path.join(report_dir, f)
                    if os.path.exists(report):
                        os.remove(report)
                        
                assets = os.path.join(report_dir, 'assets')
                if os.path.exists(assets):
                    shutil.rmtree(assets)
                        
                #remove any files that need to be separated
                if os.path.isfile(separations_manifest):
                    #set up a log file
                    separations_log = os.path.join(log_dir, 'separations.txt')
                    
                    #get a list of relevant lines from the separations manifest, splitting at the barcode (to avoid any differences with absolute paths)
                    to_be_separated = []
                    with open(separations_manifest, 'r') as f:
                        sep_list = f.read().splitlines()
                    for file in sep_list:
                        if item_barcode in file:
                            name = raw_text(file.replace('"', '').rstrip())
                            to_be_separated.append(name.split('%s\\' % shipmentDate, 1)[1])
                    
                    #if we've found any files, loop through list
                    if len(to_be_separated) > 0:
                        
                        for sep_item in to_be_separated:
                            wildcard_list = []
                            
                            #if a wildcard is used, we will use glob to build a list of all files/folders matching pattern
                            if '*' in sep_item:
                                
                                #recursive option
                                if '\\**' in sep_item:
                                    wildcard_list = glob.glob(sep_item, recursive=True)
                            
                                #wildcard at one level
                                elif '\\*' in sep_item:
                                    wildcard_list = glob.glob(sep_item)
                                
                                #now loop through this list of files/folders identified by glob
                                for wc in wildcard_list:
                                    sep_dest = os.path.join('deaccessioned', wc)
                                    separate_content(sep_dest, wc, separations_log, bag_report_dir, item_barcode)
                            
                            elif os.path.isdir(sep_item):
                                #build recursive list of all files in the folder
                                for root, dirs, files in os.walk(sep_item):
                                    for f in files:
                                        wildcard_list.append(os.path.join(root, f))
                                #loop through the list
                                for wc in wildcard_list:
                                    sep_dest = os.path.join('deaccessioned', wc)
                                    separate_content(sep_dest, wc, separations_log, bag_report_dir, item_barcode)   

                                #now remove the folder
                                cmd = 'RD /S /Q "%s"' % sep_item
                                try:
                                    subprocess.call(cmd, shell=True)
                                except subprocess.CalledProcessError as e:
                                    pass
                                             
                            elif os.path.isfile(sep_item):
                                sep_dest = os.path.join('deaccessioned', sep_item)
                                separate_content(sep_dest, sep_item, separations_log, bag_report_dir, item_barcode)                        
                
                        #compile stats and check to see if any errors reported
                        separated_file_stats = os.path.join(bag_report_dir, '%s-separation-stats.txt' % item_barcode)
                        file_stat_list = []
                        if os.path.exists(separated_file_stats):
                            with open(separated_file_stats, 'rb') as f:
                                file_stat_list = pickle.load(f)
                            
                            #loop though list of stats; add to file and byte counts and note any failures
                            sep_files = 0
                            sep_size = 0
                            di_sep = 0
                            temp_puid = []
                            success = True
                            for f in file_stat_list:
                                if f['result'] != 'Moved':
                                    success = False
                                    pass
                                else:
                                    if f['type'] == 'extracted-file':
                                        sep_files += 1
                                        sep_size += f['size']
                                        #create a running list of puids for later
                                        temp_puid.append(f['puid'])
                                    else:
                                        di_sep += 1
                                        
                                        #remove disk-image folder if no longer needed
                                        try:
                                            os.rmdir(image_dir)
                                        except OSError:
                                            pass
                            
                            #if we have any failures, this barcode will fail: we need to make sure any files designated for separation have been removed.        
                            if not success:
                                print('\n\tNOTE: one or more errors with separations.')
                                write_list(failed_list, '%s\tSeparations\tSee %s for details.' % (item_barcode, separated_file_stats))
                                continue
                            else:
                                print('\n\tSeparations completed: %s files separated (%s bytes)' % (sep_files, sep_size))
                                write_list(separated_list, '%s\t%s\t%s\t%s' % (item_barcode, str(sep_files), str(sep_size), str(di_sep)))
                                
                                #get a count for each puid in our temp list; add to dictionary
                                sep_puid_count = {puid:temp_puid.count(puid) for puid in temp_puid}
                                
                                separated_puids = os.path.join(bag_report_dir, '%s_separated_puids.txt' % item_barcode)
                                with open(separated_puids, 'wb') as f:
                                    pickle.dump(sep_puid_count, f)
                                
                                #os.remove(separated_file_stats)
                                
                                #add information on preservation event to PREMIS metadata with lxml
                                premis = os.path.join(metadata, '%s-premis.xml' % item_barcode)
                                PREMIS_NAMESPACE = "http://www.loc.gov/premis/v3"
                                PREMIS = "{%s}" % PREMIS_NAMESPACE
                                NSMAP = {'premis' : PREMIS_NAMESPACE, "xsi": "http://www.w3.org/2001/XMLSchema-instance"}

                                parser = etree.XMLParser(remove_blank_text=True)
                                
                                tree = etree.parse(premis, parser=parser)
                                root = tree.getroot()
                                
                                event = etree.SubElement(root, PREMIS + 'event')
                                eventID = etree.SubElement(event, PREMIS + 'eventIdentifier')
                                eventIDtype = etree.SubElement(eventID, PREMIS + 'eventIdentifierType')
                                eventIDtype.text = 'UUID'
                                eventIDval = etree.SubElement(eventID, PREMIS + 'eventIdentifierValue')
                                eventIDval.text = str(uuid.uuid4())

                                eventType = etree.SubElement(event, PREMIS + 'eventType')
                                eventType.text = 'deaccession'

                                eventDateTime = etree.SubElement(event, PREMIS + 'eventDateTime')
                                eventDateTime.text = str(datetime.datetime.now())

                                eventDetailInfo = etree.SubElement(event, PREMIS + 'eventDetailInformation')
                                eventDetail = etree.SubElement(eventDetailInfo, PREMIS + 'eventDetail')
                                if di_sep == 0:
                                    eventDetail.text = 'Removed %s files (%s bytes).  See "./logs/separations.txt" for list of deaccessioned content.' % (sep_files, sep_size) 
                                else:
                                    eventDetail.text = 'Removed %s files (%s bytes) and %s disk image.  See "./logs/separations.txt" for list of deaccessioned content.' % (sep_files, sep_size, di_sep) 
                                eventDetailInfo = etree.SubElement(event, PREMIS + 'eventDetailInformation')
                                eventDetail = etree.SubElement(eventDetailInfo, PREMIS + 'eventDetail')
                                eventDetail.text = 'Formal removal of an object from the inventory of a repository.'

                                eventOutcomeInfo = etree.SubElement(event, PREMIS + 'eventOutcomeInformation')
                                eventOutcome = etree.SubElement(eventOutcomeInfo, PREMIS + 'eventOutcome')
                                eventOutcome.text = '0'
                                eventOutDetail = etree.SubElement(eventOutcomeInfo, PREMIS + 'eventOutcomeDetail')
                                eventOutDetailNote = etree.SubElement(eventOutDetail, PREMIS + 'eventOutcomeDetailNote')
                                eventOutDetailNote.text = 'Successful completion'

                                linkingAgentID = etree.SubElement(event, PREMIS + 'linkingAgentIdentifier')
                                linkingAgentIDtype = etree.SubElement(linkingAgentID, PREMIS + 'linkingAgentIdentifierType')
                                linkingAgentIDtype.text = 'local'
                                linkingAgentIDvalue = etree.SubElement(linkingAgentID, PREMIS + 'linkingAgentIdentifierValue')
                                linkingAgentIDvalue.text = 'IUL BDPL'
                                linkingAgentRole = etree.SubElement(linkingAgentID, PREMIS + 'linkingAgentRole')
                                linkingAgentRole.text = 'implementer'
                                linkingAgentID = etree.SubElement(event, PREMIS + 'linkingAgentIdentifier')
                                linkingAgentIDtype = etree.SubElement(linkingAgentID, PREMIS + 'linkingAgentIdentifierType')
                                linkingAgentIDtype.text = 'local'
                                linkingAgentIDvalue = etree.SubElement(linkingAgentID, PREMIS + 'linkingAgentIdentifierValue')
                                linkingAgentIDvalue.text = 'bdpl_pag-prep.py (https://github.com/IUBLibTech/bdpl_ingest/blob/master/bdpl_bag-prep.py)'
                                linkingAgentRole = etree.SubElement(linkingAgentID, PREMIS + 'linkingAgentRole')
                                linkingAgentRole.text = 'executing software'
                                linkingObjectID = etree.SubElement(event, PREMIS + 'linkingObjectIdentifier')
                                linkingObjectIDtype = etree.SubElement(linkingObjectID, PREMIS + 'linkingObjectIdentifierType')
                                linkingObjectIDtype.text = 'local'
                                linkingObjectIDvalue = etree.SubElement(linkingObjectID, PREMIS + 'linkingObjectIdentifierValue')
                                linkingObjectIDvalue.text = item_barcode
                                
                                #add new event to root etree.Element; write to etree.ElementTree
                                root.append(event)
                                premis_text = etree.tostring(tree, pretty_print=True, xml_declaration=True, encoding="UTF-8")
                                
                                #now write to file
                                with open(premis, 'wb') as f:
                                    f.write(premis_text)
                                
                    else:
                        print('\tSeparations completed.')
                        write_list(separated_list, '%s\t0\t0\t0' % item_barcode) #include 0s for # of files separated, associated bytes, disk image
                        
            '''BAG FOLDER'''
            #make sure we haven't already bagged folder
            if not check_list(bagged_list, item_barcode):
            
                print('\n\tCreating bag for barcode folder...')
                
                begin_date = return_spreadsheet_value(ws_app, current_row, ws_columns['begin_date'])
                end_date = return_spreadsheet_value(ws_app, current_row, ws_columns['end_date'])
                content_source_type = return_spreadsheet_value(ws_app, current_row, ws_columns['content_source_type'])
                label_transcription = return_spreadsheet_value(ws_app, current_row, ws_columns['label_transcription'])
                try:
                    item_title = return_spreadsheet_value(ws_app, current_row, ws_columns['item_title'])
                except KeyError:
                    item_title = ''
                appraisal_notes = return_spreadsheet_value(ws_app, current_row, ws_columns['appraisal_notes'])
                
                dates = '%s-%s' % (begin_date, end_date)

                
                #description for bag-info.txt currently includes source media, label transcription, and appraisal note.
                desc = 'Source: %s. | Label: %s. | Title: %s. | Appraisal notes: %s. | Date range: %s' %(content_source_type, label_transcription, item_title, appraisal_notes, dates)
                desc = desc.replace('\n', ' ')    
        
                try:
                    bagit.make_bag(destination, {"Source-Organization" : unit_name, "External-Description" : desc, "External-Identifier" : item_barcode}, checksums=["md5"])
                    write_list(bagged_list, item_barcode)
                    print('\tBagging complete.')
                except (RuntimeError, PermissionError, bagit.BagError, OSError) as e:
                    print("\tUnexpected error: ", e)
                    write_list(failed_list, '%s\tbagit\t%s' % (item_barcode, e)
                    continue
            
            '''CREATE TAR'''
            #make sure file hasn't already been tarred
            if not check_list(tarred_list, item_barcode):
                
                #Make sure we have enough space to create tar file (just to be sure; we should check first, as a rule)
                print('\n\tChecking available space...')
                
                #first check available space
                (total, used, free) = shutil.disk_usage(os.getcwd())
                
                #now get size of destination
                cmd = 'du -s %s' % destination
                dir_size = int(subprocess.check_output(cmd, shell=True, text=True).split()[0])
                
                #check if the new archive will have sufficient space on disk; include addition 10240 bytes for tar file. Ff so, continue.  If not, exit with a warning
                available_space = int(free) - (dir_size * 2 + 10240)
                
                if available_space <= 0:
                    print('\n\tWARNING! Insufficient space to create tar archive.\n\t\tAvailable space: %s\n\t\tSize needed for archive: %s' % (free, string(dir_size)))
                    write_list(failed_list, '%s\tInsufficient space to create tar archive; need minimum of %s bytes' % (item_barcode, string(dir_size)))
                    continue
                else:
                    print('\tCheck complete; sufficient space for tar file.')
                
                #tar folder
                print('\n\tCreating tar archive...')
                if 'POL' in item_barcode:
                    tar_file = '%s.tar.gz' % os.path.basename(destination)
                    cmd = 'tar -czf %s %s' % (tar_file, os.path.basename(destination))
                else:
                    tar_file = '%s.tar' % os.path.basename(destination)
                    cmd = 'tar -cf %s %s' % (tar_file, os.path.basename(destination))
                
                try:
                    subprocess.check_output(cmd, shell=True)
                    write_list(tarred_list, item_barcode)
                    print('\tTar archive created')
                except (RuntimeError, PermissionError, IOError, EnvironmentError, subprocess.CalledProcessError) as e:
                    print("\tUnexpected error: ", e)
                    write_list(failed_list, '%s\ttar\t%s' % (item_barcode, e))
                    continue
                  
            '''MOVE TAR TO ARCHIVER LOCATION'''
            if not check_list(moved_list, item_barcode):
                #Just in case we are restarting from an error   
                tar_file = '%s.tar' % os.path.basename(destination)
                
                print('\n\tMoving tar file to Archiver folder...')
                
                complete_sip = os.path.join(ship_dir, tar_file)
                
                #get some stats on SIP
                print('\tCalculating SIP size...')
                SIP_size = get_size(complete_sip)
                
                print('\tCalculating SIP md5 checksum...')
                SIP_md5 = md5(complete_sip)
                
                SIP_dict = {'size' : SIP_size, 'md5' : SIP_md5, 'filename' : tar_file}
                
                #store values in a file just in case we have a failure or are otherwise interrupted...
                with open(os.path.join(bag_report_dir, 'SIP_%s.txt' % item_barcode), 'wb') as file:
                    pickle.dump(SIP_dict, file)
                    
                try:
                    shutil.move(complete_sip, archiver_drop_off)
                    write_list(moved_list, item_barcode)
                    print('\tTar file moved.')
                except (RuntimeError, PermissionError, IOError, EnvironmentError) as e:
                    print("\tUnexpected error: ", e)
                    write_list(failed_list, '%s\tmove\t%s' % (item_barcode, e))
                    continue
                        
            '''WRITE STATS TO MASTER SPREADSHEET'''
            if not check_list(metadata_list, item_barcode):
                
                sep_size = 0
                sep_files = 0
                di_sep = 0
                
                #check to see if files were separated; if so, get stats and adjust previous totals.
                if os.path.exists(separated_list):
                    with open(separated_list, 'r') as fi:
                        fi = csv.reader(fi, delimiter='\t')
                        for line in fi:
                            try:
                                if item_barcode == line[0]:
                                    sep_files = int(line[1])
                                    sep_size = int(line[2])
                                    di_sep = int(line[3])
                                    break
                            except IndexError as e:
                                print(e)
                                continue
                                    
                #Recalculate size of extracted files
                extent_raw = return_spreadsheet_value(ws_app, current_row, ws_columns['extent_raw'])
                
                if extent_raw is None:
                    extracted_size = get_size(os.path.join(destination, 'data', 'files')) 
                else:
                    extracted_size = (int(extent_raw) - sep_size)
                    
                #write corrected size back to spreadsheet
                ws_app.cell(row=current_row, column=ws_columns['extent_raw'], value=extracted_size)
                
                #if there are any files separated, adjust extracted file count
                extracted_no = int(row[17].value) - sep_files
                if extracted_no != row[17].value:
                    ws_app.cell(row=row[0].row, column=18, value=extracted_no)
                    
                wb.save(spreadsheet)
                
                #Retrieve SIP info (just in case process closed unexpectedly)
                try:
                    SIP_dict
                except NameError:
                    SIP_stats = os.path.join(bag_report_dir, 'SIP_%s.txt' % item_barcode)
                    SIP_dict = {}
                    with open(SIP_stats, 'rb') as file:
                        SIP_dict = pickle.load(file)
                
                try:
                    if row[28].value is None:
                        access_option = '-'
                    else:
                        access_option = str(row[28].value)
                except IndexError:
                    access_option = '-'
                    
                '''MAYBE CAPTURE THIS INFORMATION AT BEGINNING OF PROCESS SO IT CAN BE USED THROUGHTOUT?'''
                
                coll_title = str(row[2].value)
                coll_id = str(row[3].value)
                creator = str(row[4].value)
                source_type = str(row[6].value)
                label_transcript = str(row[7].value)
                appraisal_notes = str(row[8].value)
                restrict_stmt = str(row[9].value)
                restrict_end_date = str(row[10].value)
                migration_date = str(row[12].value)
                sip_creation_date = str(datetime.datetime.now())
                earliest_date = str(row[21].value) 
                latest_date = str(row[22].value)
                
                #determine appropriate row: overwrite if already existing or add new row, otherwise
                newrow = return_spreadsheet_row(item_ws, item_barcode)
                
                #write information on the specfic barcode
                item_ws.cell(row=newrow, column=1).value = item_barcode
                item_ws.cell(row=newrow, column=2).value = unit_name
                item_ws.cell(row=newrow, column=3).value = shipmentDate
                item_ws.cell(row=newrow, column=4).value = coll_title
                item_ws.cell(row=newrow, column=5).value = coll_id
                item_ws.cell(row=newrow, column=6).value = creator
                item_ws.cell(row=newrow, column=7).value = source_type
                item_ws.cell(row=newrow, column=8).value = label_transcript
                item_ws.cell(row=newrow, column=9).value = appraisal_notes
                item_ws.cell(row=newrow, column=10).value = earliest_date
                item_ws.cell(row=newrow, column=11).value = latest_date
                item_ws.cell(row=newrow, column=12).value = restrict_stmt
                item_ws.cell(row=newrow, column=13).value = restrict_end_date
                item_ws.cell(row=newrow, column=14).value = migration_date
                item_ws.cell(row=newrow, column=15).value = sip_creation_date
                item_ws.cell(row=newrow, column=16).value = extracted_no
                item_ws.cell(row=newrow, column=17).value = extracted_size
                item_ws.cell(row=newrow, column=18).value = SIP_dict['size']
                item_ws.cell(row=newrow, column=19).value = SIP_dict['md5']
                item_ws.cell(row=newrow, column=20).value = SIP_dict['filename']
                item_ws.cell(row=newrow, column=21).value = access_option
                  
                master_wb.save(master_spreadsheet)
                
                #add info to cumulative stats for shipment
                shipment_stats = {}
               
                #if we already have shipment stats, retrieve from file and update.  Otherwise, add values to dictionary
                if os.path.exists(stats_doc):
                    with open(stats_doc, 'rb') as file:
                        shipment_stats = pickle.load(file)
                    shipment_stats['sip_count'] += 1
                    shipment_stats['extracted_no'] += extracted_no
                    shipment_stats['extracted_size'] += extracted_size
                    shipment_stats['SIP_size'] += SIP_dict['size']
                else:
                    shipment_stats = {'sip_count' : 1, 'extracted_no' : extracted_no, 'extracted_size' : extracted_size, 'SIP_size' : SIP_dict['size']}
                
                #Write shipment stat back to file
                with open(stats_doc, 'wb') as file:
                    pickle.dump(shipment_stats, file)
                
                write_list(metadata_list, item_barcode)

                os.remove(os.path.join(bag_report_dir, 'SIP_%s.txt' % item_barcode))
                
            '''CLEAN ORIGINAL BARCODE FOLDER'''
            #remove original folder
            if not check_list(completed_list, item_barcode):
                print('\n\tRemoving original folder...')
                cmd = 'RD /S /Q "%s"' % destination
                try:
                    subprocess.check_output(cmd, shell=True)
                    write_list(completed_list, item_barcode)
                    print('\tFolder removed')
                #note failure, but write metadata to master spreadsheet
                except (PermissionError, subprocess.CalledProcessError, OSError) as e:
                    print("\tUnexpected error: ", e)
                    write_list(failed_list, '%s\tclean_original\t%s' % (item_barcode, e))
                    continue
            
            #barcode is now done!
            print('\n\t%s COMPLETED\n---------------------------------------------------------------' % item_barcode)
            
            #if barcode had previously failed, remove it from list.
            if os.path.exists(failed_list):
                with open(failed_list, "r") as f:
                    lines = f.read().splitlines()
            
                with open(failed_list, "w") as f:
                    for line in lines:
                        if not item_barcode in line:
                            f.write('%s\n' % line)
    
        #if other appraisal decision is indicated, note barcode in a list and move folder.
        else:
            if not check_list(other_list, item_barcode):
                write_list(other_list, '%s\t%s' % (item_barcode, str(row[27].value)))
                
                if not os.path.exists('review'):
                    os.mkdir('review')
                
                shutil.move(item_barcode, 'review')
                
                print('\n\tAlternate appraisal decision: %s. \n\tConfer with collecting unit as needed.' % str(row[27].value))
            
            else:
                print('\n\t%s has been moved to the "Review" folder.' % item_barcode)
            
    '''CHECK PROCESS FOR MISSING/FAILED ITEMS'''    
    #get lists from status files: how many barcodes are in each list
    list_dict = {'started' : started_list, 'moved' : moved_list, 'deaccessioned' : deaccession_list, 'other' : other_list, 'failed' : failed_list}
    tally_dict = {}
    for key, value in list_dict.items():
        if os.path.isfile(value):
            tally = []
            with open(value, 'r') as v:
                for line in v:
                    try:
                        tally.append(line.split()[0])
                    except IndexError:
                        tally = []
        else:
            tally = []
        tally_dict[key] = tally
    
    #determine if any didn't make it through the process to one of the conclusions (skipped, failed, moved, or other)
    subtotal = tally_dict['moved'] + tally_dict['other'] + tally_dict['failed'] + tally_dict['deaccessioned']
    missing = list(set(tally_dict['started']) - set(subtotal))
    
    print('\n\nSTATS:\n# of barcodes processed:\t\t%s' % len(tally_dict['started']))
    print('# of barcodes moved to Archiver:\t%s' % len(tally_dict['moved']))
    print('# of barcodes to be deaccessioned:\t%s' % len(tally_dict['deaccessioned']))
    print('# of barcodes that failed:\t\t%s' % len(tally_dict['failed']))
    print('# of barcodes that require review:\t%s' % len(tally_dict['other']))
        
    #make sure all items are accounted for
    if len(missing) > 0:
        print("\n# of barcodes not accounted for:\t%s\n\t%s" % (len(missing), '\n\t'.join(missing)))
    
    #give alert if there are barcode folders not listed in spreadsheet
    if len(missing_from_spreadsheet) > 0:
        print('\n\nNOTE: shipment folder included %s barcode folder(s) not listed in spreadsheet.  These have been moved to "unaccounted" folder; review as needed.' % len(missing_from_spreadsheet))
        
    #if any items have failed, quit; only update cumulative information once we are all done.
    if len(tally_dict['failed']) > 0 or len(tally_dict['other']) > 0:
        print('\n***Exiting process. Address remaining barcodes and run again to write shipment information to master spreadsheet.***')
        if len(tally_dict['failed']) > 0:
            cmd = 'notepad %s' % failed_list
            subprocess.check_output(cmd)
        if len(tally_dict['other']) > 0:
            cmd = 'notepad %s' % other_list
            subprocess.check_output(cmd)
        
        sys.exit(1)
        
    
    '''UPDATE CUMULATIVE INFORMATION'''
    print('\n\n------------------------------------------------------------\n\nUPDATING MASTER SPREADSHEET')
    #get info from stats document
    if os.path.exists(stats_doc):
        shipment_stats = {}
        with open(stats_doc, 'rb') as file:
            shipment_stats = pickle.load(file)
            
        #get duration info 
        duration_stats = {}
        with open(duration_doc, 'rb') as file:
            duration_stats = pickle.load(file)
        
        #check to make sure shipment hasn't already been written to worksheet; if so, we will update that information.
        newrow = cumulative_ws.max_row+1

        iterrows = cumulative_ws.iter_rows()
        next(iterrows)
        
        for row in iterrows:    
            if not row[0].value is None:
                if unit_name in row[0].value and shipmentDate in row[1].value:
                    newrow = row[0].row
                    break
        
        print('\nWriting cumulative information to spreadsheet...')
        
        #write (or update) shipment info in cumulative sheet of master workbook
        cumulative_ws.cell(row=newrow, column=1, value = unit_name) 
        cumulative_ws.cell(row=newrow, column=2, value = shipmentDate)
        cumulative_ws.cell(row=newrow, column=3, value = shipment_stats['sip_count'])
        cumulative_ws.cell(row=newrow, column=4, value = shipment_stats['extracted_no'])
        cumulative_ws.cell(row=newrow, column=5, value = shipment_stats['extracted_size'])
        cumulative_ws.cell(row=newrow, column=6, value = shipment_stats['SIP_size'])
        cumulative_ws.cell(row=newrow, column=7, value = duration_stats['earliest'])
        cumulative_ws.cell(row=newrow, column=8, value = duration_stats['latest'])
        cumulative_ws.cell(row=newrow, column=9, value = duration_stats['duration'])
    else:
        print('\nNOTE: Shipment statistics failed to be recorded.')
        
    #now record this shipment's format information, to be saved in master workbook.
    #first, reecover our puid and format reports
    format_list = []
    if os.path.exists(format_report):
        with open(format_report, 'rb') as f:
            format_list = pickle.load(f)
    
    puid_list = []
    if os.path.exists(puid_report):
        with open(puid_report, 'rb') as f:
            puid_list = pickle.load(f)
    
    #remove any duplicate entries in the puid list, sort, and append 'barcode' as the first item
    puid_list = list(set(puid_list))
    puid_list.sort()
    puid_list.insert(0, 'barcode')
    
    #now create a puid dictionary so that we can refer to columns in our 'formats' worksheet.  Add one to each enumerator value so we align with openpyxl column index
    puid_dict = {}
    for pu, id in enumerate(puid_list):
        puid_dict[id] = pu+1

    format_sheet = 'puids_%s_%s' % (unit_name, shipmentDate)
    
    #if this puid sheet already exists, we'll just remove it and start anew...
    if format_sheet in master_wb.sheetnames:
        master_wb.remove(master_wb[format_sheet])
    
    fws = master_wb.create_sheet(format_sheet)
    fws.append(puid_list)

    print('\nWriting format information...')
    
    #loop through the dictionaries of our format report
    for index in range(len(format_list)):
        for key in format_list[index]:
        
            #retrieve the information on puids we separated; this is a dictionary with puid as key and # of files separated as value
            separated_puids = os.path.join(bag_report_dir, '%s_separated_puids.txt' % key)
            sep_puid_count = {}
            if os.path.exists(separated_puids):
                with open(separated_puids, 'rb') as f:
                    sep_puid_count = pickle.load(f)
            
            #set the row variable; write barcode info to first cell
            newrow = fws.max_row+1
            fws.cell(row=newrow, column=puid_dict['barcode'], value=key)
            
            #now loop through of our list of dictionaries with stats on specific formats
            for item in format_list[index][key]:
                
                #see if files of a given puid were separated; if so, reduce the count accordingly
                if len(sep_puid_count) > 0:
                    if item['puid'] in list(sep_puid_count):
                        item['count'] -= sep_puid_count[item['puid']]
                
                #if count is now 0, continue to next puid; otherwise write information to format sheet
                if item['count'] < 1:
                    continue
                else:
                    fws.cell(row=fws.max_row, column=puid_dict[item['puid']], value=item['count'])
        
            #os.remove(separated_puids)
            
    #print format totals for shipment
    max_row = fws.max_row + 1
    
    fws.cell(row=max_row, column=1, value='Totals:')
    
    #loop through sheet and sum each column
    for col in fws.iter_cols(2, fws.max_column, 2, fws.max_row):
        count = 0
        for c in col:
            if not c.value is None:
                count += c.value
                colno = c.column
        fws.cell(row=max_row, column=colno, value=count)
    
    
    
    print('\nPackaging for shipment %s shipment %s completed!!' % (unit_name, shipmentDate))
    
    #save workbooks
    wb.save(spreadsheet)
    master_wb.save(master_spreadsheet)
    
    #save a copy of the master spreadsheet to SDA
    shutil.copy(master_spreadsheet, archiver_drop_off)
    

if __name__ == '__main__':
    
    os.system('cls')
    
    #print BDPL screen
    fname = "C:/BDPL/scripts/bdpl.txt"
    if os.path.exists(fname):
        with open(fname, 'r') as fin:
            print(fin.read())
    
    main()