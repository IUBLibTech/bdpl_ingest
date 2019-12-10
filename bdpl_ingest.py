#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Python 3

"""

This project was inspired by and includes elements of Brunnhilde, a Siegfried-based digital archives reporting tool
github.com/timothyryanwalsh/brunnhilde
Copyright (c) 2017 Tim Walsh, distributed under The MIT License (MIT)
http://bitarchivist.net

"""

from collections import OrderedDict
from collections import Counter
import csv
import datetime
import errno
import math
import os
import re
import shutil
import sqlite3
import subprocess
import sys
import uuid
import xml
import lxml
from lxml import etree
import tempfile
import fnmatch
from tkinter import *
import tkinter.filedialog
from tkinter import ttk
import glob
import pickle
import time
import openpyxl
import glob
import hashlib
import psutil
import chardet
from urllib.parse import unquote

# from dfxml project
import Objects

def get_encoding(input):
    with open(input, 'rb') as f:
        result = chardet.detect(f.read())
    
    return result['encoding']
    
def check_premis(term, folders, item_barcode):
    #check to see if an event is already in our premis list--i.e., it's been successfully completed.  Currently only used for most resource-intensive operations: virus scheck, sensitive data scan, format id, and checksum calculation.
    
    #set up premis_list
    premis_list = pickleLoad('premis_list', folders, item_barcode)
    
    #see if term has been recorded at all
    found = [dic for dic in premis_list if dic['eventType'] == term]
    
    #if not recorded, it hasn't been run
    if not found: 
        return False
    else:
        #for virus scans, we will assume that completion may have either a 0 or non-zero value.  No need to run again.
        if term == 'virus check':
            return True
        elif term == 'metadata extraction':
            if [dc for dc in found if 'tree v1.7.0' in dc['linkingAgentIDvalue']]:
                return True
        #for other microservices, check if operation completed successfully; if so, return True, otherwise False
        else:
            if [dc for dc in found if dc['eventOutcomeDetail'] in ['0', 0]]:
                return True
            else:
                return False

def check_item_status(folders, item_barcode):

    temp_dir = folders['temp_dir']

    #If a 'done' file exists, we know the whole process was completed
    done_file = os.path.join(temp_dir, 'done.txt')
    if os.path.exists(done_file): 
        print('\n\nNOTE: this item barcode has completed the entire ingest workflow.  Consult with the digital preservation librarian if you believe additional procedures are needed.')
    #if no 'done' file, see where we are with the item...
    else:
        premis_list = pickleLoad('premis_list', folders, item_barcode)
        if len(premis_list) > 0:
            print('\n\nIngest of item has been initiated; the following procedures have been completed:\n\t', '\n\t'.join(list(set((i['%s' % 'eventType'] for i in premis_list)))))


def first_run(unit_name, shipmentDate, item_barcode, gui_vars):
    #this function only runs when a record is loaded for the first time.

    folders = bdpl_folders(unit_name, shipmentDate, item_barcode)
    
    if gui_vars['platform'] == 'bdpl_ingest':
    
        newscreen()

        #check if key data has been entered
        if not verify_data(unit_name, shipmentDate, item_barcode):
            return False, 'basic info wrong'

    spreadsheet = find_spreadsheet(folders, unit_name, shipmentDate)
    if not os.path.exists(spreadsheet): 
        return False, "spreadsheet doesn't exist"
    
    #check that metadata exists in spreadsheet, create a dict of associated values, and check to see if barcode was already ingested
    if not load_metadata(folders, item_barcode, spreadsheet):
        return False, "spreadsheet metadata doesn't exist"
        
    #check if item has already been completed or, if started, what procedures have been completed
    check_item_status(folders, item_barcode)
    
    #write metadata to gui if we are using bdpl_ingest (skip for bdpl_ripstation_ingest and bdpl_legacy)
    if gui_vars['platform'] == 'bdpl_ingest':
        metadata_to_gui(gui_vars, folders, item_barcode)

    #now create folders if they don't exist
    create_folders(folders)
    
    if gui_vars['platform'] == 'bdpl_ingest':  
        print('\n\nRecord loaded successfully; ready for next operation.')
    
    return True, 'all set'
    
def create_folders(folders):
    
    folders_created = os.path.join(folders['temp_dir'], 'folders_created.txt')
    
    if not os.path.exists(folders_created):
        for target in ['destination', 'image_dir', 'files_dir', 'metadata', 'temp_dir', 'reports_dir', 'log_dir', 'media_image_dir']:
            try:
                os.makedirs(folders["%s" % target])
            except OSError as exception:
                if exception.errno != errno.EEXIST:
                    raise
        
        open(folders_created, 'a').close()

def bdpl_folders(unit_name, shipmentDate, item_barcode=None):
    #this function creates folder variables
    folders = {}
    folders['unit_home'] = os.path.join('Z:\\', unit_name, 'ingest' )
    folders['ship_dir'] = os.path.join(folders['unit_home'], '%s' % shipmentDate)
    folders['media_image_dir'] = os.path.join('Z:\\', 'media-images', '%s' % unit_name)
    
    if not item_barcode is None:
        folders['destination'] = os.path.join(folders['ship_dir'], "%s" % item_barcode)
        folders['image_dir'] = os.path.join(folders['destination'], "disk-image")
        folders['files_dir'] = os.path.join(folders['destination'], "files")
        folders['metadata'] = os.path.join(folders['destination'], "metadata")
        folders['temp_dir'] = os.path.join(folders['destination'], 'temp')
        folders['reports_dir'] = os.path.join(folders['metadata'], 'reports')
        folders['log_dir'] = os.path.join(folders['metadata'], 'logs')
        folders['imagefile'] = os.path.join(folders['image_dir'], '%s.dd' % item_barcode)
        folders['dfxml_output'] = os.path.join(folders['metadata'], '%s-dfxml.xml' % item_barcode)
        folders['bulkext_dir'] = os.path.join(folders['destination'], 'bulk_extractor')
        folders['bulkext_log'] = os.path.join(folders['log_dir'], 'bulkext-log.txt')

    
    return folders
    
def pickleLoad(list_name, folders, item_barcode):
    metadata = folders['metadata']
    temp_dir = folders['temp_dir']
    temp_file = os.path.join(temp_dir, '%s.txt' % list_name)
    
    #this list will be used to store anything pulled in from premis xml; we'll check later to see if anything was added
    temp_premis = []
    
    if list_name in ['premis_list', 'temp_dfxml', 'duplicates', 'fs_list', 'partition_info_list']:
        temp_object = []
    else:
        temp_object = {}
    
    #special steps if we're dealing with a premis list...
    if list_name == "premis_list":
        
        premis_path = os.path.join(metadata, '%s-premis.xml' % item_barcode)
        premis_xml_included = os.path.join(temp_dir, 'premis_xml_included.txt')
        
        #for our list of premis events, we want to pull in information that may have already been written to premis xml
        if os.path.exists(premis_path):
            
            #check to see if operation has already been completed (we'll write an empty file once we've done so)
            if not os.path.exists(premis_xml_included):
                PREMIS_NAMESPACE = "http://www.loc.gov/premis/v3"
                NSMAP = {'premis' : PREMIS_NAMESPACE, "xsi": "http://www.w3.org/2001/XMLSchema-instance"}
                parser = etree.XMLParser(remove_blank_text=True)
                tree = etree.parse(premis_path, parser=parser)
                root = tree.getroot()
                events = tree.xpath("//premis:event", namespaces=NSMAP)
                
                for e in events:
                    temp_dict = {}
                    temp_dict['eventType'] = e.findtext('./premis:eventType', namespaces=NSMAP)
                    temp_dict['eventOutcomeDetail'] = e.findtext('./premis:eventOutcomeInformation/premis:eventOutcome', namespaces=NSMAP)
                    temp_dict['timestamp'] = e.findtext('./premis:eventDateTime', namespaces=NSMAP)
                    temp_dict['eventDetailInfo'] = e.findall('./premis:eventDetailInformation/premis:eventDetail', namespaces=NSMAP)[0].text
                    temp_dict['eventDetailInfo_additional'] = e.findall('./premis:eventDetailInformation/premis:eventDetail', namespaces=NSMAP)[1].text
                    temp_dict['linkingAgentIDvalue'] = e.findall('./premis:linkingAgentIdentifier/premis:linkingAgentIdentifierValue', namespaces=NSMAP)[1].text
                    temp_premis.append(temp_dict)
                    
                #now create our premis_xml_included.txt file so we don't go through this again.
                open(premis_xml_included, 'a').close()
        
    #make sure there's something in the file
    if os.path.exists(temp_file) and os.path.getsize(temp_file) > 0:
        with open(temp_file, 'rb') as file:
            temp_object = pickle.load(file)
    
    #if anything was added from our premix.xml file, 
    if len(temp_premis) > 0:
        for d in temp_premis:
            if not d in temp_object:
                temp_object.append(d)
        
        #now sort based on ['timestamp']
        temp_object.sort(key=lambda x:x['timestamp'])
            
    return temp_object

def pickleDump(list_name, list_contents, folders):
    temp_dir = folders['temp_dir']
    temp_file = os.path.join(temp_dir, '%s.txt' % list_name)
     
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)
        
    with open(temp_file, 'wb') as file:
        pickle.dump(list_contents, file)

def secureCopy(file_source, folders, item_barcode):
    if not os.path.exists(file_source):
        print('\n\nThis file source does not appear to exist: "%s"\n\nPlease verify the correct source has been identified.' % file_source)
        return
    
    files_dir = folders['files_dir']
    log_dir = folders['log_dir']
    
    #function takes the file source and destination as well as  a specific premis event to be used in documenting action
    print('\n\nFILE REPLICATION: TERACOPY\n\n\tSOURCE: %s \n\tDESTINATION: %s' % (file_source, files_dir))
    
    #set variables for premis
    timestamp = str(datetime.datetime.now())             
    migrate_ver = "TeraCopy v3.26"
    
    #set variables for copy operation; note that if we are using a file list, TERACOPY requires a '*' before the source. 
    if os.path.isfile(file_source):
        copycmd = 'TERACOPY COPY *"%s" %s /SkipAll /CLOSE' % (file_source, files_dir)
    else:
        copycmd = 'TERACOPY COPY "%s" %s /SkipAll /CLOSE' % (file_source, files_dir)
    
    try:
        exitcode = subprocess.call(copycmd, shell=True, text=True)
    except subprocess.CalledProcessError as e:
        print('\n\tFile replication failed:\n\n\t%s' % e)
        return
    
    #check to see if files are actually present (TeraCopy may complete without copying...)
    if not checkFiles(files_dir):
        exitcode = '1'
    
    #need to find Teracopy SQLITE db and export list of copied files to csv log file
    list_of_files = glob.glob(os.path.join(os.path.expandvars('C:\\Users\%USERNAME%\AppData\Roaming\TeraCopy\History'), '*'))
    tera_db = max(list_of_files, key=os.path.getctime)
    
    conn = sqlite3.connect(tera_db)
    conn.text_factory = str
    cur = conn.cursor()
    results = cur.execute("SELECT * from Files")
    
    tera_log = os.path.join(log_dir, 'teracopy_log.csv')
    with open(tera_log, 'w', encoding='utf8') as output:
        writer = csv.writer(output, lineterminator='\n')
        header = ['Source', 'Offset', 'State', 'Size', 'Attributes', 'IsFolder', 'Creation', 'Access', 'Write', 'SourceCRC', 'TargetCRC', 'TargetName', 'Message', 'Marked', 'Hidden']
        writer.writerow(header)
        writer.writerows(results)
    
    #get count of files that were actually moved
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    moved_files = cur.execute("SELECT Source FROM Files WHERE IsFolder = 0 and State = 2")
    count = sum(1 for file in moved_files)
    print('\n\t%s files successfully transferred to %s.' % (count, files_dir))
    
    
    #capture premis
    premis_list = pickleLoad('premis_list', folders, item_barcode)
    premis_list.append(premis_dict(timestamp, 'replication', exitcode, copycmd, 'Created a copy of an object that is, bit-wise, identical to the original.', migrate_ver))
    pickleDump('premis_list', premis_list, folders)       
        

    print('\n\tFile replication completed; proceed to content analysis.')

def ddrescue_image(folders, item_barcode, sourceDevice, other_device):
    
    temp_dir = folders['temp_dir']
    log_dir = folders['log_dir']
    image_dir = folders['image_dir']
    imagefile = folders['imagefile']
    
    check_device = subprocess.check_output('cat /proc/partitions', text=True)
    
    if sourceDevice == 'Zip':
        ps_cmd = "Get-Partition | % {New-Object PSObject -Property @{'DiskModel'=(Get-Disk $_.DiskNumber).Model; 'DriveLetter'=$_.DriveLetter}}"
        cmd = 'powershell.exe "%s"' % ps_cmd
        out = subprocess.check_output(cmd, shell=True, text=True)
        for line in out.splitlines():
            if 'ZIP 100' in line:
                  drive_ltr = line.split()[2]
        
        try:
            drive_ltr
        except UnboundLocalError:
            print('\n\nNOTE: Zip drive not recognized.  If you have not done so, insert disk into drive and allow device to complete initial loading.')
            return
        
        #get device name from /proc/partitions
        for line in check_device.splitlines():
            if len(line.split()) == 5 and drive_ltr in line.split()[4]:
                dd_target = '/dev/%s' % line.split()[3]
    
    #use case involving internal hard drive
    elif sourceDevice == 'Other':
        #make sure device name is correct
        if other_device in check_device:
            dd_target = '/dev/%s' % other_device
        else:
            print('\nNOTE: device name "%s" not found in /proc/partitions; verify and try again.' % other_device)
            return
        
    else:
        dd_target = sourceDevice
        
    print('\n\nDISK IMAGE CREATION: DDRESCUE\n\n\tSOURCE: %s \n\tDESTINATION: %s' % (dd_target, imagefile))
    
    #set up premis list
    premis_list = pickleLoad('premis_list', folders, item_barcode)
    
    #create variables for mapfile and ddrescue commands (first and second passes)
    mapfile = os.path.join(temp_dir, '%s.map' % item_barcode)
           
    ddrescue_events1 = os.path.join(log_dir, 'ddrescue_events1.txt')
    ddrescue_events2 = os.path.join(log_dir, 'ddrescue_events2.txt')
    
    ddrescue_rates1 = os.path.join(log_dir, 'ddrescue_rates1.txt')
    ddrescue_rates2 = os.path.join(log_dir, 'ddrescue_rates2.txt')
    
    ddrescue_reads1 = os.path.join(log_dir, 'ddrescue_reads1.txt')
    ddrescue_reads2 = os.path.join(log_dir, 'ddrescue_reads2.txt')
    
    migrate_ver = subprocess.check_output('ddrescue -V', shell=True, text=True).split('\n', 1)[0]  
    timestamp1 = str(datetime.datetime.now())
    
    copycmd1 = 'ddrescue -n --log-events=%s --log-rates=%s --log-reads=%s %s %s %s' % (ddrescue_events1, ddrescue_rates1, ddrescue_reads1, dd_target, imagefile, mapfile)
    
    #run commands via subprocess; per ddrescue instructions, we need to run it twice    
    print('\n--------------------------------------First pass with ddrescue------------------------------------\n')
    exitcode1 = subprocess.call(copycmd1, shell=True, text=True)
    
    premis_list.append(premis_dict(timestamp1, 'disk image creation', exitcode1, copycmd1, 'First pass; extracted a disk image from the physical information carrier.', migrate_ver))
    
    #new timestamp for second pass (recommended by ddrescue developers)
    timestamp2 = str(datetime.datetime.now())
    
    copycmd2 = 'ddrescue -d -r2 --log-events=%s --log-rates=%s --log-reads=%s %s %s %s' % (ddrescue_events2, ddrescue_rates2, ddrescue_reads2, dd_target, imagefile, mapfile)
    
    print('\n\n--------------------------------------Second pass with ddrescue------------------------------------\n')
    
    exitcode2 = subprocess.call(copycmd2, shell=True, text=True)
    
    if os.path.exists(imagefile) and os.stat(imagefile).st_size > 0:
            print('\n\n\tDisk image created; proceeding to next step...')
            exitcode2 = 0
            premis_list.append(premis_dict(timestamp2, 'disk image creation', exitcode2, copycmd2, 'Second pass; extracted a disk image from the physical information carrier.', migrate_ver))
    else:
        print('\n\nDISK IMAGE CREATION FAILED: Indicate any issues in note to collecting unit.')
    
    #save premis
    pickleDump('premis_list', premis_list, folders)

def fc5025_image(folders, item_barcode, disk525):

    imagefile = folders['imagefile']
    log_dir = folders['log_dir']
    
    print('\n\n\DISK IMAGE CREATION: DeviceSideData FC5025\n\n\tSOURCE: 5.25" floppy disk \n\tDESTINATION: %s\n\n' % imagefile)
            
    #create premis list
    premis_list = pickleLoad('premis_list', folders, item_barcode)
    
    disk_type_options = { 'Apple DOS 3.3 (16-sector)' : 'apple33', 'Apple DOS 3.2 (13-sector)' : 'apple32', 'Apple ProDOS' : 'applepro', 'Commodore 1541' : 'c1541', 'TI-99/4A 90k' : 'ti99', 'TI-99/4A 180k' : 'ti99ds180', 'TI-99/4A 360k' : 'ti99ds360', 'Atari 810' : 'atari810', 'MS-DOS 1200k' : 'msdos12', 'MS-DOS 360k' : 'msdos360', 'North Star MDS-A-D 175k' : 'mdsad', 'North Star MDS-A-D 350k' : 'mdsad350', 'Kaypro 2 CP/M 2.2' : 'kaypro2', 'Kaypro 4 CP/M 2.2' : 'kaypro4', 'CalComp Vistagraphics 4500' : 'vg4500', 'PMC MicroMate' : 'pmc', 'Tandy Color Computer Disk BASIC' : 'coco', 'Motorola VersaDOS' : 'versa' }

    timestamp = str(datetime.datetime.now())
    copycmd = 'fcimage -f %s %s | tee -a %s' % (disk_type_options[disk525], imagefile, os.path.join(log_dir, 'fcimage.log'))

    exitcode = subprocess.call(copycmd, shell=True, text=True)
    
    if exitcode == 0:
        premis_list.append(premis_dict(timestamp, 'disk image creation', exitcode, copycmd, 'Extracted a disk image from the physical information carrier.', 'FCIMAGE v1309'))  
    
    else:
        #FC5025 reports non-0 exitcode if there are any read errors; therefore, if a disk image larger than 0 bytes exists, we will call it a success
        if os.stat(imagefile).st_size > 0:
            premis_list.append(premis_dict(timestamp, 'disk image creation', 0, copycmd, 'Extracted a disk image from the physical information carrier.', 'FCIMAGE v1309'))
        else:
            print('\n\nDisk image not successfully created; verify you have selected the correct disk type and try again (if possible).  Otherwise, indicate issues in note to collecting unit.')
            return
    print('\n\n\tDisk image created; proceeding to next step...')
    
    #save premis
    pickleDump('premis_list', premis_list, folders)

def disk_image_replication(folders, item_barcode):
    
    imagefile = folders['imagefile']
    files_dir = folders['files_dir']
    
    #now attempt to replicate/extract content from disk image
    print('\n\nDISK IMAGE FILE REPLICATION: ')
    
    #set our software versions for unhfs and tsk_recover, just in case...
    cmd = 'unhfs 2>&1'
    unhfs_carve_ver = subprocess.check_output(cmd, shell=True, text=True).splitlines()[0]
    tsk_carve_ver = 'tsk_recover: %s ' % subprocess.check_output('tsk_recover -V', text=True).strip()
    
    #now get information on filesystems and (if present) partitions.  We will need to choose which tool to use based on file system; if UDF or ISO9660 present, use TeraCopy; otherwise use unhfs or tsk_recover
    secureCopy_list = ['udf', 'iso9660']
    unhfs_list = ['osx', 'hfs', 'apple', 'apple_hfs', 'mfs', 'hfs plus']
    tsk_list = ['ntfs', 'fat', 'fat12', 'fat16', 'fat32', 'exfat', 'ext2', 'ext3', 'ext4', 'ufs', 'ufs1', 'ufs2', 'ext', 'yaffs2', 'hfs+']
    
    #recover lists
    fs_list = pickleLoad('fs_list', folders, item_barcode)
    partition_info_list = pickleLoad('partition_info_list', folders, item_barcode)
    
    #see what file systems we have
    if len(fs_list) > 0:
    
        print('\n\tDisktype has identified the following file system: ', ', '.join(fs_list))
        
        #now check for any partitions; if none, go ahead and use teracopy, tsk_recover, or unhfs depending on the file system
        if len(partition_info_list) == 0:

            print('\n\tNo partition information...')
            
            if any(fs in ' '.join(fs_list) for fs in secureCopy_list):
                secureCopy(optical_drive_letter(), folders, item_barcode)

            elif any(fs in ' '.join(fs_list) for fs in unhfs_list):
                carvefiles('unhfs', folders, files_dir, unhfs_carve_ver, '', item_barcode)
            
            elif any(fs in ' '.join(fs_list) for fs in tsk_list): 
                carvefiles('tsk_recover', folders, files_dir, tsk_carve_ver, '', item_barcode)
            
            else:
                print('\n\tFile system not recognized by tools')
        #if there are one or more partitions, use tsk_recover or unhfs        
        elif len(partition_info_list) >= 1:
        
            for part_dict in partition_info_list:
            
                if len(partition_info_list) == 1:
                    outfolder = files_dir
                else:
                    outfolder = os.path.join(files_dir, 'partition_%s' % part_dict['slot'])
                
                if part_dict['desc'] in unhfs_list:
                    carvefiles('unhfs', folders, outfolder, unhfs_carve_ver, part_dict['slot'], item_barcode)
                                  
                elif part_dict['desc'] in tsk_list:
                    carvefiles('tsk_recover', folders, outfolder, tsk_carve_ver, part_dict['start'], item_barcode)
    
    else:
        print('\n\tNo files to be replicated.')

def carvefiles(tool, folders, outfolder, carve_ver, partition, item_barcode): 
    imagefile = folders['imagefile']
    files_dir = folders['files_dir']
    dfxml_output = folders['dfxml_output']
    
    if not os.path.exists(outfolder):
        os.makedirs(outfolder)
    
    if tool == 'unhfs':
        
        if partition == '':
            carve_cmd = 'unhfs -sfm-substitutions -resforks APPLEDOUBLE -o "%s" "%s" 2>nul' % (outfolder, imagefile)
        else:
            carve_cmd = 'unhfs -sfm-substitutions -partition %s -resforks APPLEDOUBLE -o "%s" "%s" 2>nul' % (partition, outfolder, imagefile)
    
    else:
        
        if partition == '':
            carve_cmd = 'tsk_recover -a %s %s' % (imagefile, outfolder)
        
        else:
            carve_cmd = 'tsk_recover -a -o %s %s %s' % (partition, imagefile, outfolder)
        
    print('\n\tTOOL: %s\n\n\tSOURCE: %s \n\n\tDESTINATION: %s\n' % (tool, imagefile, outfolder))
    
    timestamp = str(datetime.datetime.now())  
    exitcode = subprocess.call(carve_cmd, shell=True)
    
    premis_list = pickleLoad('premis_list', folders, item_barcode)
    premis_list.append(premis_dict(timestamp, 'replication', exitcode, carve_cmd, "Created a copy of an object that is, bit-wise, identical to the original.", carve_ver))
    pickleDump('premis_list', premis_list, folders)
    
    #if no files were extracted, remove partition folder.
    if not checkFiles(outfolder) and outfolder != files_dir:
        os.rmdir(outfolder)
    
    #if tsk_recover has been run, go through and fix the file MAC times
    if tool == 'tsk_recover' and exitcode == 0:
        
        #generate DFXML with fiwalk
        if not os.path.exists(dfxml_output):
            produce_dfxml(imagefile, folders, item_barcode)
        
        #use DFXML output to get correct MAC times and update files
        fix_dates(outfolder, folders, item_barcode)
    
    elif tool == 'unhfs' and os.path.exists(outfolder):
        file_count = sum([len(files) for r, d, files in os.walk(outfolder)])
        print('\t%s files successfully transferred to %s.' % (file_count, outfolder))
        
    print('\n\tFile replication completed; proceed to content analysis.')

def time_to_int(str_time):
    """ Convert datetime to unix integer value """
    dt = time.mktime(datetime.datetime.strptime(str_time, 
        "%Y-%m-%dT%H:%M:%S").timetuple())
    return dt
    
def fix_dates(outfolder, folders, item_barcode):
    #adapted from Timothy Walsh's Disk Image Processor: https://github.com/CCA-Public/diskimageprocessor
    
    dfxml_output = folders['dfxml_output']
    
    print('\n\nFILE MAC TIME CORRECTION (USING DFXML)')
    
    timestamp = str(datetime.datetime.now())
     
    try:
        for (event, obj) in Objects.iterparse(dfxml_output):
            # only work on FileObjects
            if not isinstance(obj, Objects.FileObject):
                continue

            # skip directories and links
            if obj.name_type:
                if obj.name_type not in ["r", "d"]:
                    continue

            # record filename
            dfxml_filename = obj.filename
            dfxml_filedate = int(time.time()) # default to current time

            # record last modified or last created date
            try:
                mtime = obj.mtime
                mtime = str(mtime)
            except:
                pass

            try:
                crtime = obj.crtime
                crtime = str(crtime)
            except:
                pass

            # fallback to created date if last modified doesn't exist
            if mtime and (mtime != 'None'):
                mtime = time_to_int(mtime[:19])
                dfxml_filedate = mtime
            elif crtime and (crtime != 'None'):
                crtime = time_to_int(crtime[:19])
                dfxml_filedate = crtime
            else:
                continue

            # rewrite last modified date of corresponding file in objects/files
            exported_filepath = os.path.join(outfolder, dfxml_filename)
            if os.path.isdir(exported_filepath):
                os.utime(exported_filepath, (dfxml_filedate, dfxml_filedate))
            elif os.path.isfile(exported_filepath):
                os.utime(exported_filepath, (dfxml_filedate, dfxml_filedate)) 
            else:
                continue

    except (ValueError, xml.etree.ElementTree.ParseError):
        print('\nUnable to read DFXML!')
        pass
    
    premis_list = pickleLoad('premis_list', folders, item_barcode)
    premis_list.append(premis_dict(timestamp, 'metadata modification', 0, 'https://github.com/CCA-Public/diskimageprocessor/blob/master/diskimageprocessor.py#L446-L489', 'Corrected file timestamps to match information extracted from disk image.', 'Adapted from Disk Image Processor Version: 1.0.0 (Tim Walsh)'))
    pickleDump('premis_list', premis_list, folders)

def lsdvd_check(folders, item_barcode, drive_letter):
    temp_dir = folders['temp_dir']
    reports_dir = folders['reports_dir']
    log_dir = folders['log_dir']
    
    #set up PREMIS list
    premis_list = pickleLoad('premis_list', folders, item_barcode)
    
    #document information about DVD titles; set variables and get lsdvd version
    lsdvd_temp = os.path.join(temp_dir, 'lsdvd.txt')
    cmd = 'lsdvd -V > %s 2>&1' % lsdvd_temp
    
    subprocess.check_output(cmd, shell=True, text=True)
    
    with open(lsdvd_temp, 'r', encoding='utf-8') as f:
        lsdvd_ver = f.read().split(' - ')[0]
    
    #now run lsdvd to get info about DVD, including # of titles
    lsdvdout = os.path.join(reports_dir, "%s_lsdvd.xml" % item_barcode)
    timestamp = str(datetime.datetime.now())
    lsdvdcmd = 'lsdvd -Ox -x %s > %s 2> NUL' % (drive_letter, lsdvdout)
    exitcode = subprocess.call(lsdvdcmd, shell=True, text=True)
    
    premis_list.append(premis_dict(timestamp, 'metadata extraction', exitcode, lsdvdcmd, 'Extracted content information from DVD, including titles, chapters, audio streams and video.', lsdvd_ver))     
    pickleDump('premis_list', premis_list, folders)
    
    #now verify how many titles are on the disk.  Set a default value of 0
    titlecount = 0
    
    #check file to see how many titles are on DVD using lsdvd XML output. 
    parser = etree.XMLParser(recover=True)

    try:
        doc = etree.parse(lsdvdout, parser=parser)
        titlecount = int(doc.xpath("count(//lsdvd//track)"))
        
        #check for PAL content, just in case...
        formats = doc.xpath("//format")
        if [f for f in formats if f.text == 'PAL']:
            titlecount = 'PAL'
            return titlecount
            
    #if lsdvd fails or information not in report, get the title count by parsing directory...
    except (OSError, lxml.etree.XMLSyntaxError):
        titlelist = glob.glob(os.path.join(drive_letter, '**/VIDEO_TS', '*_*_*.VOB'), recursive=True)
        count = []
        for t in titlelist:
            #parse VOB filenames to get # of titles
            count.append(int(t.rsplit('_', 2)[1]))
        if len(count) > 0:
            titlecount = max(set(count))
    
    #if we haven't identified titles (i.e., we do not have a DVD), delete lsdvd output
    if titlecount == 0:
        os.remove(lsdvdout)
        
    return titlecount

def normalize_dvd_content(folders, item_barcode, titlecount, drive_letter):
    temp_dir = folders['temp_dir']
    reports_dir = folders['reports_dir']
    files_dir = folders['files_dir']
    log_dir = folders['log_dir']
    
    #set up PREMIS list
    premis_list = pickleLoad('premis_list', folders, item_barcode)
        
    #check current directory; change to a temp directory to store files
    bdpl_cwd = 'C:\\BDPL\\scripts'
    
    ffmpeg_temp = os.path.join(temp_dir, 'ffmpeg')
    if not os.path.exists(ffmpeg_temp):
        os.makedirs(ffmpeg_temp)
    
    os.chdir(ffmpeg_temp)
    
    #get ffmpeg version
    ffmpeg_ver =  '; '.join(subprocess.check_output('"C:\\Program Files\\ffmpeg\\bin\\ffmpeg" -version', shell=True, text=True).splitlines()[0:2])
    
    print('\n\nMOVING IMAGE FILE NORMALIZATION: FFMPEG')
    
    #loop through titles and rip each one to mpeg using native streams
    for title in range(1, (titlecount+1)):
        titlelist = glob.glob(os.path.join(drive_letter, "**/VIDEO_TS", "VTS_%s_*.VOB" % str(title).zfill(2)), recursive=True)
        #be sure list is sorted
        sorted(titlelist)
        
        if len(titlelist) > 0:
            
            #check if title track is missing audio--could make trouble for other tracks...
            audio_test = {}
            print('\n\tChecking audio streams...')
            for t in titlelist:
                cmd = "ffprobe -i %s -hide_banner -show_streams -select_streams a -loglevel error" % t
                try:
                    audio_check = subprocess.check_output(cmd, shell=True, text=True)
                    audio_test[t] = audio_check
                except subprocess.CalledProcessError:
                    pass
            
            if len(audio_test) == 0:
                print('\nWARNING: unable to access information on DVD. Moving image normalization has failed...')
                return
            
            #if there's no audio in any track, it's OK
            if all(value == '' for value in audio_test.values()):
                pass
            #if our first track lacks audio, add a dummy track
            elif audio_test[titlelist[0]] == '':
                dummy_audio = os.path.join(temp_dir, 'added_silence.mpg')
                
                cmd = "ffmpeg -y -nostdin -loglevel warning -i %s -f lavfi -i anullsrc -c:v copy -c:a aac -shortest -target ntsc-dvd %s" % (titlelist[0], dummy_audio)
                
                print('\n\tCorrecting missing audio on first track...')
                
                subprocess.call(cmd, text=True)
                
                #replace original item from list
                del titlelist[0]
                titlelist.insert(0, dummy_audio)
            
            timestamp = str(datetime.datetime.now())
            
            ffmpegout = os.path.join(files_dir, '%s-%s.mpg' % (item_barcode, str(title).zfill(2)))
            ffmpeg_cmd = 'ffmpeg -y -nostdin -loglevel warning -report -stats -i "concat:%s" -c copy -target ntsc-dvd %s' % ('|'.join(titlelist), ffmpegout)
            
            print('\n\tGenerating title %s of %s: %s\n' % (str(title), str(titlecount), ffmpegout))
            
            exitcode = subprocess.call(ffmpeg_cmd, shell=True, text=True)
                
            premis_list.append(premis_dict(timestamp, 'normalization', exitcode, ffmpeg_cmd, 'Transformed object to an institutionally supported preservation format (.MPG) with a direct copy of all streams.', ffmpeg_ver))
            
            #move and rename ffmpeg log file
            ffmpeglog = glob.glob(os.path.join(ffmpeg_temp, 'ffmpeg-*.log'))[0]
            shutil.move(ffmpeglog, os.path.join(log_dir, '%s-%s-ffmpeg.log' % (item_barcode, str(title).zfill(2))))
    
    #save PREMIS to file       
    pickleDump('premis_list', premis_list, folders)
    
    #move back to original directory
    os.chdir(bdpl_cwd)
    
    print('\n\tMoving image normalization completed; proceed to content analysis.')

def cdda_image_creation(folders, item_barcode, sourceDevice):
    temp_dir = folders['temp_dir']
    reports_dir = folders['reports_dir']
    files_dir = folders['files_dir']
    log_dir = folders['log_dir']
    image_dir = folders['image_dir']
    
    #set up PREMIS list
    premis_list = pickleLoad('premis_list', folders, item_barcode)

    print('\n\nDISK IMAGE CREATION: CDRDAO\n\n\tSOURCE: %s \n\tDESTINATION: %s' % (sourceDevice, image_dir))
    
    #determine appropriate drive ID for cdrdao; save output of command to temp file
    cdr_scan = os.path.join(temp_dir, 'cdr_scan.txt')
    scan_cmd = 'cdrdao scanbus > %s 2>&1' % cdr_scan
    subprocess.check_output(scan_cmd, shell=True, text=True)

    #pull drive ID and cdrdao version from file
    with open(cdr_scan, 'r') as f:
        info = f.read().splitlines()
    cdrdao_ver = info[0].split(' - ')[0]
    drive_id = info[8].split(':')[0]
        
    #get info about CD using cdrdao; record this as a premis event, too.
    disk_info_report = os.path.join(reports_dir, '%s-cdrdao-diskinfo.txt' % item_barcode)
    cdrdao_cmd = 'cdrdao disk-info --device %s --driver generic-mmc-raw > %s 2>&1' % (drive_id, disk_info_report)
    timestamp = str(datetime.datetime.now())
    exitcode = subprocess.call(cdrdao_cmd, shell=True, text=True)
    
    premis_list.append(premis_dict(timestamp, 'metadata extraction', exitcode, cdrdao_cmd, 'Extracted information about the CD-R, including medium, TOC type, number of sessions, etc.', cdrdao_ver))

    #read log file to determine # of sessions on disk.
    with open(disk_info_report, 'r') as f:
        for line in f:
            if 'Sessions             :' in line:
                sessions = int(line.split(':')[1].strip())
    
    t2c_ver = subprocess.check_output('toc2cue -V', shell=True, text=True).strip()
    
    #for each session, create a bin/toc file
    for x in range(1, (sessions+1)):
        cdr_bin = os.path.join(image_dir, "%s-%s.bin") % (item_barcode, str(x).zfill(2))
        cdr_toc = os.path.join(image_dir, "%s-%s.toc") % (item_barcode, str(x).zfill(2))
        
        print('\n\tGenerating session %s of %s: %s\n\n' % (str(x), str(sessions), cdr_bin))
        
        #create separate bin/cue for each session
        cdr_cmd = 'cdrdao read-cd --read-raw --session %s --datafile %s --device %s --driver generic-mmc-raw -v 1 %s' % (str(x), cdr_bin, drive_id, cdr_toc)
        
        timestamp = str(datetime.datetime.now())
        
        exitcode = subprocess.call(cdr_cmd, shell=True, text=True)
        
        premis_list.append(premis_dict(timestamp, 'disk image creation', exitcode, cdr_cmd, 'Extracted a disk image from the physical information carrier.', cdrdao_ver))
                    
        #convert TOC to CUE
        cue = os.path.join(image_dir, "%s-%s.cue") % (item_barcode, str(sessions).zfill(2))
        cue_log = os.path.join(log_dir, "%s-%s_toc2cue.log") % (item_barcode, str(sessions).zfill(2))
        t2c_cmd = 'toc2cue %s %s > %s 2>&1' % (cdr_toc, cue, cue_log)
        timestamp = str(datetime.datetime.now())
        exitcode2 = subprocess.call(t2c_cmd, shell=True, text=True)
        
        #toc2cue may try to encode path information as binary data--let's fix that
        with open(cue, 'rb') as infile:
            cue_info = infile.readlines()[1:]
        
        with open(cue, 'w') as outfile:
            outfile.write('FILE "%s" BINARY\n' % os.path.basename(cdr_bin))
        
        with open(cue, 'ab') as outfile:
            for line in cue_info:
                outfile.write(line)           
        
        premis_list.append(premis_dict(timestamp, 'metadata modification', exitcode2, t2c_cmd, "Converted the CD's table of contents (TOC) file to the CUE format.", t2c_ver))
        
        #place a copy of the .cue file for the first session in files_dir for the forthcoming WAV; this session will have audio data
        if x == 1:
            new_cue = os.path.join(files_dir, '%s.cue' % item_barcode)
            
            #now write the new cue file
            with open(new_cue, 'w') as outfile:
                outfile.write('FILE "%s.wav" WAVE\n' % item_barcode)
                
            with open(new_cue, 'ab') as outfile:
                for line in cue_info:
                    outfile.write(line)
                
    #save PREMIS to file
    pickleDump('premis_list', premis_list, folders)
    
    print('\n\tCDDA disk image created; moving on to next step...')

def cdda_wav_creation(folders, item_barcode, sourceDevice):
    temp_dir = folders['temp_dir']
    reports_dir = folders['reports_dir']
    files_dir = folders['files_dir']
    log_dir = folders['log_dir']
    
    #set up PREMIS list
    premis_list = pickleLoad('premis_list', folders, item_barcode)
    
    #get cdparanoia version
    paranoia_temp = os.path.join(temp_dir, 'paranoia.txt')
    ver_cmd = 'cd-paranoia -V > %s 2>&1' % paranoia_temp
    
    exitcode = subprocess.call(ver_cmd, shell=True, text=True)
    with open(paranoia_temp, 'r') as f:
        paranoia_ver = f.read().splitlines()[0]
    
    paranoia_log = os.path.join(log_dir, '%s-cdparanoia.log' % item_barcode)
    paranoia_out = os.path.join(files_dir, '%s.wav' % item_barcode)
    
    print('\n\nAUDIO CONTENT NORMALIZATION: CDPARANOIA\n\n\tSOURCE: %s \n\tDESTINATION: %s\n' % (sourceDevice, paranoia_out))
    
    paranoia_cmd = 'cd-paranoia -l %s -w [00:00:00.00]- %s' % (paranoia_log, paranoia_out)
    
    timestamp = str(datetime.datetime.now())
    exitcode = subprocess.call(paranoia_cmd, shell=True, text=True)
    
    premis_list.append(premis_dict(timestamp, 'normalization', exitcode, paranoia_cmd, 'Transformed object to an institutionally supported preservation format (.WAV).', paranoia_ver))
    
    #save PREMIS to file
    pickleDump('premis_list', premis_list, folders)
    
    print('\n\tAudio normalization complete; proceed to content analysis.')

def transferContent(unit_name, shipmentDate, item_barcode, transfer_vars):    
    #set variables
    folders = bdpl_folders(unit_name, shipmentDate, item_barcode)
    
    files_dir = folders['files_dir']
    log_dir = folders['log_dir']
    imagefile = folders['imagefile']
    temp_dir = folders['temp_dir']
    reports_dir = folders['reports_dir']
    files_dir = folders['files_dir']
    image_dir = folders['image_dir']
    dfxml_output = folders['dfxml_output']
    
    jobType = transfer_vars['jobType'].get()
    source = transfer_vars['source'].get()
    sourceDevice = transfer_vars['sourceDevice'].get()
    disk525 = transfer_vars['disk525'].get()    
    mediaStatus = transfer_vars['mediaStatus'].get()
    other_device = transfer_vars['other_device'].get()
    platform = transfer_vars['platform']
    
    newscreen()
    
    #check that information is added to GUI
    if not verify_data(unit_name, shipmentDate, item_barcode):
        return
    
    if not os.path.exists(folders['destination']):
        print('\nWARNING: load record before proceeding')
        return

    if not mediaStatus and jobType != 'Copy_only':
        print('\n\nMake sure that media has been inserted/attached; check the "Attached?" box and continue.')
        return

    print('\n\nSTEP 1. TRANSFER CONTENT')
        
    #check to see if content will include disk image; if nothing entered, exit and prompt user to do so        
    if jobType == 'Copy_only':
        
        teracopy_source = source.replace('/', '\\')
        
        if 'bdpl_transfer_lists' in teracopy_source:
            teracopy_source = glob.glob(os.path.join('Z:\\bdpl_transfer_lists', '%s.txt' % item_barcode))
            if len(teracopy_source) != 1:
                print('\n\nOperation failed: could not find transfer list for this barcode. Please verify the list and try again.')
                return
            else: 
                teracopy_source = teracopy_source[0]
            
        secureCopy(teracopy_source, folders, item_barcode)
                
    elif jobType == 'Disk_image':     
        
        #special process for 5.25" floppies: use FC5025
        if sourceDevice == '5.25':
            if disk525 == 'N/A':
                print('\n\nError; be sure to select the appropriate 5.25" floppy disk type from the drop down menu.')
                return
            else:
                fc5025_image(folders, item_barcode, disk525)
        
        #all other disk imaging will use ddrescue
        else:    

            ddrescue_image(folders, item_barcode, sourceDevice, other_device)
        
        #exit if disk image doesn't exist
        if not os.path.isfile(imagefile):
            print('\nNOTE: Disk image not created. Exiting transfer process; correct issues and try again.')
            return
        
        #get info on the disk image (fsstat, ils, mmls, and disktype)
        disk_image_info(folders, item_barcode)
        
        #create a logical copy of content on disk image
        disk_image_replication(folders, item_barcode)
             
    elif jobType == 'DVD':
            
        #create disk image of DVD
        ddrescue_image(folders, item_barcode, sourceDevice, other_device)
        
        #check DVD for title information
        drive_letter = "%s\\" % optical_drive_letter()
        titlecount = lsdvd_check(folders, item_barcode, drive_letter)
        
        #if we titles found on DVD, go ahead and normalize to .MPG; if none found, exit
        if titlecount > 0:       
            drive_letter = "%s\\" % optical_drive_letter()
            normalize_dvd_content(folders, item_barcode, titlecount, drive_letter)
        elif titlecount == 'PAL':
            print('\n\nWARNING: DVD is PAL formatted! Need to figure out correct ffmpeg command; stop file normalization...')
            return
        else:
            print('\nWARNING: DVD does not appear to have any titles; job type should likely be Disk_image.  Manually review disc and re-transfer content if necessary.')
            return
    
    elif jobType == 'CDDA':
        
        #create a copy or raw pulse code modulated (PCM) audio data 
        cdda_image_creation(folders, item_barcode, sourceDevice)
        
        #now rip CDDA to WAV using cd-paranoia (Cygwin build; note hyphen)
        cdda_wav_creation(folders, item_barcode, sourceDevice)
        
    else: 
        print('\n\nError; please indicate the appropriate job type')
        return
    
    print('\n\n--------------------------------------------------------------------------------------------------\n\n')
    
def premis_dict(timestamp, event_type, event_outcome, event_detail, event_detail_note, agent_id):
    temp_dict = {}
    temp_dict['eventType'] = event_type
    temp_dict['eventOutcomeDetail'] = event_outcome
    temp_dict['timestamp'] = timestamp
    temp_dict['eventDetailInfo'] = event_detail
    temp_dict['eventDetailInfo_additional'] = event_detail_note
    temp_dict['linkingAgentIDvalue'] = agent_id
    return temp_dict


def run_antivirus(folders, item_barcode):
    files_dir = folders['files_dir']
    log_dir = folders['log_dir']
    metadata = folders['metadata']
    
    #get version
    cmd = 'clamscan -V'
    av_ver = subprocess.check_output(cmd, text=True).rstrip()
    
    virus_log = os.path.join(log_dir, 'viruscheck-log.txt')
    av_command = 'clamscan -i -l %s --recursive %s' % (virus_log, files_dir)  
    
    timestamp = str(datetime.datetime.now())
    exitcode = subprocess.call(av_command, shell=True, text=True)
    
    #write info to metadata_dict
    metadata_dict = pickleLoad('metadata_dict', folders, item_barcode)
        
    with open(virus_log, 'r') as f:
        if "Infected files: 0" not in f.read():
            metadata_dict['virus_scan_results'] = 'WARNING! Virus or malware found; see %s.' % virus_log
        
        else:
            metadata_dict['virus_scan_results'] = '-'

        
    pickleDump('metadata_dict', metadata_dict, folders)
    
    #save preservation to PREMIS
    premis_list = pickleLoad('premis_list', folders, item_barcode)
    premis_list.append(premis_dict(timestamp, 'virus check', exitcode, av_command, 'Scanned files for malicious programs.', av_ver))
    pickleDump('premis_list', premis_list, folders)
    
    print('\n\tVirus scan completed; moving on to next step...')
   

def run_bulkext(folders, item_barcode):

    bulkext_dir = folders['bulkext_dir']
    bulkext_log = folders['bulkext_log']
    files_dir = folders['files_dir']
    reports_dir = folders['reports_dir']
    
    #get bulk extractor version for premis
    try:
        be_ver = subprocess.check_output(['bulk_extractor', '-V'], shell=True, text=True).rstrip()
    except subprocess.CalledProcessError as e:
        be_ver = e.output.rstrip()
    
    print('\n\tScan underway...be patient!\n')
    
    #use default command with buklk_extractor
    bulkext_command = 'bulk_extractor -x aes -x base64 -x elf -x exif -x gps -x hiberfile -x httplogs -x json -x kml -x net -x pdf -x sqlite -x winlnk -x winpe -x winprefetch -S ssn_mode=2 -q -1 -o "%s" -R "%s" > "%s"' % (bulkext_dir, files_dir, bulkext_log)
    
    if os.path.exists(bulkext_dir):
        shutil.rmtree(bulkext_dir)
    
    try:
        os.makedirs(bulkext_dir)
    except OSError as exception:
        if exception.errno != errno.EEXIST:
            raise

    #create timestamp
    timestamp = str(datetime.datetime.now())        

    exitcode = subprocess.call(bulkext_command, shell=True, text=True)
       
    premis_list = pickleLoad('premis_list', folders, item_barcode)       
    premis_list.append(premis_dict(timestamp, 'sensitive data scan', exitcode, bulkext_command, 'Scanned files for potentially sensitive information, including Social Security and credit card numbers.', be_ver))
    pickleDump('premis_list', premis_list, folders)
    
    #create a cumulative BE report
    cumulative_report = os.path.join(bulkext_dir, 'cumulative.txt')
    if os.path.exists(cumulative_report):
        os.remove(cumulative_report)
        
    for myfile in ('pii.txt', 'ccn.txt', 'email.txt', 'telephone.txt', 'find.txt'):
        myfile = os.path.join(bulkext_dir, myfile)
        if os.path.exists(myfile) and os.stat(myfile).st_size > 0:
            with open(myfile, 'rb') as filein:
                data = filein.read().splitlines()    
            with open(cumulative_report, 'a', encoding='utf8') as outfile:
                outfile.write('%s: %s\n' % (os.path.basename(myfile), len(data[5:])))
    if not os.path.exists(cumulative_report):         
        open(cumulative_report, 'a').close()

    #move any b_e histogram files, if needed
    for myfile in ('email_domain_histogram.txt', 'find_histogram.txt', 'telephone_histogram.txt'):
        current_file = os.path.join(bulkext_dir, myfile)
        try:    
            if os.stat(current_file).st_size > 0:
                shutil.copy(current_file, reports_dir)
        except OSError:
            continue
    
    print('\n\tSensitive data scan completed; moving on to next step...')

def import_csv(folders):
    temp_dir = folders['temp_dir']
    reports_dir = folders['reports_dir']

    siegfried_db = os.path.join(temp_dir, 'siegfried.sqlite')
    conn = sqlite3.connect(siegfried_db)
    conn.text_factory = str  # allows utf-8 data to be stored
    cursor = conn.cursor()

    print('\n\tImporting siegried file to sqlite3 database...')
    """Import csv file into sqlite db"""
    sf_file = os.path.join(reports_dir, 'siegfried.csv')
    
    f = open(sf_file, 'r', encoding='utf8')
    
    try:
        reader = csv.reader(x.replace('\0', '') for x in f) # replace null bytes with empty strings on read
    except UnicodeDecodeError:
        f = (x.strip() for x in f) # skip non-utf8 encodable characters
        reader = csv.reader(x.replace('\0', '') for x in f) # replace null bytes with empty strings on read
    header = True
    for row in reader:
        if header:
            header = False # gather column names from first row of csv
            sql = "DROP TABLE IF EXISTS siegfried"
            cursor.execute(sql)
            sql = "CREATE TABLE siegfried (filename text, filesize text, modified text, errors text, namespace text, id text, format text, version text, mime text, basis text, warning text)"
            cursor.execute(sql)
            insertsql = "INSERT INTO siegfried VALUES (%s)" % (", ".join([ "?" for column in row ]))
            rowlen = len(row)
        else:
            # skip lines that don't have right number of columns
            if len(row) == rowlen:
                cursor.execute(insertsql, row)
    conn.commit()
    f.close()
    
    sqlite_done = os.path.join(temp_dir, 'sqlite_done.txt')
    open(sqlite_done, 'a').close()
    
    cursor.close()
    conn.close()

def generate_reports(cursor, html, folders, re_analyze, item_barcode):
    temp_dir = folders['temp_dir']
    reports_dir = folders['reports_dir']
    bulkext_dir = folders['bulkext_dir']
    
    print('\n\tGenerating format reports and writing html...')
    
    """Run sql queries on db to generate reports, write to csv and html"""
    full_header = ['Filename', 'Filesize', 'Date modified', 'Errors', 
                'Namespace', 'ID', 'Format', 'Format version', 'MIME type', 
                'Basis for ID', 'Warning']
    
    # sorted format list report
    path = os.path.join(reports_dir, 'formats.csv')
    if not os.path.exists(path) or re_analyze:
        sql = "SELECT format, id, COUNT(*) as 'num' FROM siegfried GROUP BY format ORDER BY num DESC"
        format_header = ['Format', 'ID', 'Count']
        sqlite_to_csv(sql, path, format_header, cursor)
    write_html('File formats', path, ',', html, folders, item_barcode)

    # sorted format and version list report
    path = os.path.join(reports_dir, 'formatVersions.csv')
    if not os.path.exists(path) or re_analyze:
        sql = "SELECT format, id, version, COUNT(*) as 'num' FROM siegfried GROUP BY format, version ORDER BY num DESC"
        version_header = ['Format', 'ID', 'Version', 'Count']
        sqlite_to_csv(sql, path, version_header, cursor)
    write_html('File format versions', path, ',', html, folders, item_barcode)

    # sorted mimetype list report
    path = os.path.join(reports_dir, 'mimetypes.csv')
    if not os.path.exists(path) or re_analyze:
        sql = "SELECT mime, COUNT(*) as 'num' FROM siegfried GROUP BY mime ORDER BY num DESC"
        mime_header = ['MIME type', 'Count']
        sqlite_to_csv(sql, path, mime_header, cursor)
    write_html('MIME types', path, ',', html, folders, item_barcode)

    # dates report
    path = os.path.join(reports_dir, 'years.csv')
    write_html('Last modified dates by year', path, ',', html, folders, item_barcode)
    
    # unidentified files report
    path = os.path.join(reports_dir, 'unidentified.csv')
    if not os.path.exists(path) or re_analyze:
        sql = "SELECT * FROM siegfried WHERE id='UNKNOWN';"
        sqlite_to_csv(sql, path, full_header, cursor)
    write_html('Unidentified', path, ',', html, folders, item_barcode)
    
    # errors report
    path = os.path.join(reports_dir, 'errors.csv')
    if not os.path.exists(path) or re_analyze:
        sql = "SELECT * FROM siegfried WHERE errors <> '';"
        sqlite_to_csv(sql, path, full_header, cursor)
    write_html('Errors', path, ',', html, folders, item_barcode)
    
    # duplicates report: retrieve our 'duplicates' file instead of CSV
    dup_list = pickleLoad('duplicates', folders, item_barcode)
    write_html('Duplicates', dup_list, ',', html, folders, item_barcode)
    
    #PII report: 
    cumulative_report = os.path.join(bulkext_dir, 'cumulative.txt')
    if os.path.exists(cumulative_report):
        write_html('Personally Identifiable Information (PII)', cumulative_report, '\n', html, folders, item_barcode)
    
    
def sqlite_to_csv(sql, path, header, cursor):
    """Write sql query result to csv"""
    # in python3, specify newline to prevent extra csv lines in windows
    # in python2, write csv in byte mode
    if (sys.version_info > (3, 0)):
        report = open(path, 'w', newline='', encoding='utf8')
    else:
        report = open(path, 'w')
    w = csv.writer(report, lineterminator='\n')
    w.writerow(header)
    for row in cursor.execute(sql):
        w.writerow(row)
    report.close()

def write_pronom_links(old_file, new_file):
    """Use regex to replace fmt/# and x-fmt/# PUIDs with link to appropriate PRONOM page"""

    in_file = open(old_file, 'r', encoding='utf8')
    out_file = open(new_file, 'w', encoding='utf8')


    for line in in_file:
        regex = r"fmt\/[0-9]+|x\-fmt\/[0-9]+" #regex to match fmt/# or x-fmt/#
        pronom_links_to_replace = re.findall(regex, line)
        new_line = line
        for match in pronom_links_to_replace:
            new_line = line.replace(match, "<a href=\"http://nationalarchives.gov.uk/PRONOM/" + 
                    match + "\" target=\"_blank\">" + match + "</a>")
            line = new_line # allow for more than one match per line
        out_file.write(new_line)

    in_file.close()
    out_file.close()

def write_html(header, path, file_delimiter, html, folders, item_barcode):
    """Write csv file to html table"""
    reports_dir = folders['reports_dir']

    # write header
    html.write('\n<a name="%s" style="padding-top: 40px;"></a>' % header)
    html.write('\n<h4>%s</h4>' % header)
    
    if header == 'Duplicates':
        html.write('\n<p><em>Duplicates are grouped by hash value.</em></p>')
        
        dup_list = path
        
        numline = len(dup_list)
        
        if numline > 1: #aka more rows than just header
            # read md5s from csv and write to list
            hash_list = []
            for row in dup_list:
                hash_list.append(row[3])
            # deduplicate md5_list
            hash_list = list(OrderedDict.fromkeys(hash_list))
            # for each hash in md5_list, print header, file info, and list of matching files
            for hash_value in hash_list:
                html.write('\n<p>Files matching checksum <strong>%s</strong>:</p>' % hash_value)
                html.write('\n<table class="table table-sm table-responsive table-bordered table-hover">')
                html.write('\n<thead>')
                html.write('\n<tr>')
                html.write('\n<th>Filename</th><th>Filesize</th>')
                html.write('<th>Date modified</th>')
                html.write('<th>Checksum</th>')
                html.write('\n</tr>')
                html.write('\n</thead>')
                html.write('\n<tbody>')
                for row in dup_list:
                    if row[3] == '%s' % hash_value:
                        # write data
                        html.write('\n<tr>')
                        for column in row:
                            html.write('\n<td>' + str(column) + '</td>')
                        html.write('\n</tr>')
                html.write('\n</tbody>')
                html.write('\n</table>')
        
            #save a copy of the duplicates for the reports
            dup_report = os.path.join(reports_dir, 'duplicates.csv')
            with open(dup_report, "w", newline="", encoding='utf-8') as f:
                writer = csv.writer(f)
                dup_header = ['Filename', 'Filesize', 'Date modified', 'Checksum']
                writer.writerow(dup_header)
                for item in dup_list:
                    writer.writerow(item)
        else:
            html.write('\nNone found.\n<br><br>')
        
    else:
        in_file = open(path, 'r', encoding="utf-8")
        # count lines and then return to start of file
        numline = len(in_file.readlines())
        in_file.seek(0)

        #open csv reader
        r = csv.reader(in_file, delimiter="%s" % file_delimiter)
        
        # if writing PII, handle separately
        if header == 'Personally Identifiable Information (PII)':
            html.write('\n<p><em>Potential PII in source, as identified by bulk_extractor.</em></p>')
            
            pii_list = []
        
            metadata_dict = pickleLoad('metadata_dict', folders, item_barcode)
            
            #check that there are any PII results
            if os.stat(path).st_size > 0:
                html.write('\n<table class="table table-sm table-responsive table-hover">')
                html.write('\n<thead>')
                html.write('\n<tr>')
                html.write('\n<th>PII type</th>')
                html.write('\n<th># of matches (may be false)</th>')
                html.write('\n<th>More information (if available)</th>')
                html.write('\n</tr>')
                html.write('\n</thead>')
                html.write('\n<tbody>')
                with open(path, 'r') as pii_info:
                    for line in pii_info:
                        html.write('\n<tr>')
                        if 'pii.txt' in line:
                            # write data
                            html.write('\n<td>SSNs, Account Nos., Birth Dates, etc.</td>')
                            html.write('\n<td>' + line.split()[1] + '</td>')
                            html.write('\n<td>Use BE_Viewer to verify results; report.xml file located at: %s.</td>' % folders['bulkext_dir'])
                            pii_list.append('ACCOUNT NOs')
                        if 'ccn.txt' in line:
                            html.write('\n<td>Credit Card Nos.</td>')
                            html.write('\n<td>' + line.split()[1] + '</td>')
                            html.write('\n<td>Use BE_Viewer to verify results; report.xml file located at: %s.</td>' % folders['bulkext_dir'])
                            pii_list.append('CCNs')
                        if 'email.txt' in line:
                            html.write('\n<td>Email address domains (may include 3rd party information)</td>')
                            html.write('\n<td>' + line.split()[1] + '</td>')
                            html.write('\n<td>See: <a href="./email_domain_histogram.txt">Email domain histogram</a></td>')
                            pii_list.append('EMAIL')
                        if 'telephone.txt' in line:
                            html.write('\n<td>Telephone numbers (may include 3rd party information)</td>')
                            html.write('\n<td>' + line.split()[1] + '</td>')
                            html.write('\n<td>See: <a href="./telephone_histogram.txt">Telephone # histogram</a></td>')
                            pii_list.append('TELEPHONE NOs')
                        if 'find.txt' in line:
                            html.write('\n<td>Sensitive terms and phrases</td>')
                            html.write('\n<td>' + line.split()[1] + '</td>')
                            html.write('\n<td>See: <a href="./find_histogram.txt">Keyword histogram</a></td>')
                            pii_list.append('TERMS')
                        html.write('\n</tr>')   
                html.write('\n</tbody>')
                html.write('\n</table>')
                
                if len(pii_list) > 0:
                    metadata_dict['pii_scan_results'] = '%s.' % ', '.join(pii_list)
                else: 
                    metadata_dict['pii_scan_results'] = '-'
        
            else:
                html.write('\nNone found.')
                metadata_dict['pii_scan_results'] = '-'
            
            pickleDump('metadata_dict', metadata_dict, folders)

        # otherwise write as normal
        else:
            if numline > 1: #aka more rows than just header
                # add borders to table for full-width tables only
                full_width_table_headers = ['Unidentified', 'Errors']
                if header in full_width_table_headers:
                    html.write('\n<table class="table table-sm table-responsive table-bordered table-hover">')
                else:
                    html.write('\n<table class="table table-sm table-responsive table-hover">')
                # write header row
                html.write('\n<thead>')
                html.write('\n<tr>')
                row1 = next(r)
                for column in row1:
                    html.write('\n<th>' + str(column) + '</th>')
                html.write('\n</tr>')
                html.write('\n</thead>')
                # write data rows
                html.write('\n<tbody>')
                for row in r:
                    # write data
                    html.write('\n<tr>')
                    for column in row:
                        html.write('\n<td>' + str(column) + '</td>')
                    html.write('\n</tr>')
                html.write('\n</tbody>')
                html.write('\n</table>')
            else:
                html.write('\nNone found.\n<br><br>')
    
        in_file.close()
    
def convert_size(size):
    # convert size to human-readable form
    if (size == 0):
        return '0 bytes'
    size_name = ("bytes", "KB", "MB", "GB", "TB", "PB", "EB", "ZB", "YB")
    i = int(math.floor(math.log(size,1024)))
    p = math.pow(1024,i)
    s = round(size/p)
    s = str(s)
    s = s.replace('.0', '')
    return '%s %s' % (s,size_name[i])
    
def close_html(html):
    """Add JavaScript and write html closing tags"""
    html.write('\n</div>')
    html.write('\n</div>')
    html.write('\n</div>')
    html.write('\n</div>')
    html.write('\n<script src="./assets//js/jquery-3.3.1.slim.min.js"></script>')
    html.write('\n<script src="./assets//js/popper.min.js"></script>')
    html.write('\n<script src="./assets//js/bootstrap.min.js"></script>')
    html.write('\n<script>$(".navbar-nav .nav-link").on("click", function(){ $(".navbar-nav").find(".active").removeClass("active"); $(this).addClass("active"); });</script>')
    html.write('\n<script>$(".navbar-brand").on("click", function(){ $(".navbar-nav").find(".active").removeClass("active"); });</script>')
    html.write('\n</body>')
    html.write('\n</html>')

def close_files_conns_on_exit(html, conn, cursor):
    cursor.close()
    conn.close()
    html.close()

def get_stats(folders, cursor, html, item_barcode, re_analyze, jobType):
    """Get aggregate statistics and write to html report"""
    temp_dir = folders['temp_dir']
    files_dir = folders['files_dir']
    reports_dir = folders['reports_dir']
    log_dir = folders['log_dir']
    
    print('\n\tGetting statistics about content...')
    
    # get stats from sqlite db
    cursor.execute("SELECT COUNT(*) from siegfried;") # total files
    num_files = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) from siegfried where filesize='0';") # empty files
    empty_files = cursor.fetchone()[0]
    
    #for DVDs, we will use stats from normalized files; however, we will also need disk image stats
    if jobType == 'DVD':
        file_stats = []
        for f in os.listdir(files_dir):
            file = os.path.join(files_dir, f)
            file_dict = {}
            size = os.path.getsize(file)
            mtime = datetime.datetime.fromtimestamp(os.path.getmtime(file)).isoformat()
            ctime = datetime.datetime.fromtimestamp(os.path.getctime(file)).isoformat()
            atime = datetime.datetime.fromtimestamp(os.path.getatime(file)).isoformat()[:-7]
            checksum = md5(file)
            
            file_dict = { 'name' : file, 'size' : size, 'mtime' : mtime, 'ctime' : ctime, 'atime' : atime, 'checksum' : checksum}
            file_stats.append(file_dict)
        
        try:
            with open(os.path.join(temp_dir, 'checksums_di.txt'), 'rb') as f:
                file_stats_di = pickle.load(f)
        except FileNotFoundError:
            pass
    else:
        file_stats = []
        try:
            with open(os.path.join(temp_dir, 'checksums.txt'), 'rb') as f:
                file_stats = pickle.load(f)
        except FileNotFoundError:
            pass
        
    #Get stats on duplicates. Just in case the bdpl ingest tool crashes after compiling a duplicates list, we'll check to see if it already exists
    dup_list = []
    if os.path.exists(os.path.join(temp_dir, 'duplicates.txt')) and not re_analyze:
        dup_list = pickleLoad('duplicates', folders, item_barcode)
    else:
        #next, create a new dictionary that IDs checksums that correspond to 1 or more files. NOTE: the 'file_stats' list will be empty for DVDs, so we'll skip this step in that case
        if len(file_stats) > 1:
            stat_dict = {}
            for dctnry in file_stats:
                if int(dctnry['size']) > 0:
                    if dctnry['checksum'] in stat_dict:
                        stat_dict[dctnry['checksum']].append(dctnry['name'])
                    else:
                        stat_dict[dctnry['checksum']] = [dctnry['name']]
           
            #go through new dict and find checksums with duplicates
            for chksm in [key for key, values in stat_dict.items() if len(values) > 1]:
                for fname in stat_dict[chksm]:
                    temp = [item for item in file_stats if item['checksum'] == chksm and item['name'] == fname][0]
                    dup_list.append([temp['name'], temp['size'], temp['mtime'], temp['checksum']])
            
        #save this duplicate file for later when we need to write to html
        pickleDump('duplicates', dup_list, folders)
    
    #total duplicates = total length of duplicate list
    all_dupes = len(dup_list)

    #distinct duplicates = # of unique checksums in the duplicates list
    distinct_dupes = len(set([c[3] for c in dup_list]))

    #duplicate copies = # of unique files that may have one or more copies
    duplicate_copies = int(all_dupes) - int(distinct_dupes) # number of duplicate copies of unique files
    duplicate_copies = str(duplicate_copies)
    
    distinct_files = int(num_files) - int(duplicate_copies)
    distinct_files = str(distinct_files)
        
    cursor.execute("SELECT COUNT(*) FROM siegfried WHERE id='UNKNOWN';") # unidentified files
    unidentified_files = cursor.fetchone()[0]

    #next get date information using info pulled from dfxml
    date_info = []
    
    #for dvd jobs, we need to use disk image metadata for dates...
    if jobType == 'DVD':
        file_stats = file_stats_di
    
    #let's not accept file mtimes that were set when content was replicated.  Compare file time against timestamp for replication...
    premis_list = pickleLoad('premis_list', folders, item_barcode)
    try:
        bdpl_time = [p for p in premis_list if p['eventType'] == 'replication'][0]['timestamp'].split('.')[0].replace('T', ' ')
    except IndexError:
        bdpl_time = datetime.datetime.fromtimestamp(os.path.getmtime(os.path.join(temp_dir, 'folders_created.txt'))).isoformat().replace('T', ' ').split('.')[0]
    
    bdpl_time = datetime.datetime.strptime(bdpl_time, "%Y-%m-%d %H:%M:%S")
    
    if len(file_stats) > 0:
        for dctnry in file_stats:
            dt_time = dctnry['mtime'].replace('T', ' ').split('.')[0]
            dt_time = datetime.datetime.strptime(dt_time, "%Y-%m-%d %H:%M:%S")
            if dt_time < bdpl_time:
                date_info.append(dctnry['mtime'])
            else:
                date_info.append('undated')
        
        #remove all occurences of 'undated', to get better info
        date_info_check = [x for x in date_info if x != 'undated']
        
        if len(date_info_check) > 0:
            begin_date = min(date_info_check)[:4]
            end_date = max(date_info_check)[:4]
            earliest_date = min(date_info_check)
            latest_date = max(date_info_check)   
        
        else:
            begin_date = "undated"
            end_date = "undated"
            earliest_date = "undated"
            latest_date = "undated"
    
    else:
        begin_date = "undated"
        end_date = "undated"
        earliest_date = "undated"
        latest_date = "undated"
        
    #generate a year count
    year_info = [x[:4] for x in date_info]
    year_info = [x if x != 'unda' else 'undated' for x in year_info]
    
    year_count = dict(Counter(year_info))
    
    path = os.path.join(reports_dir, 'years.csv')    
    with open(path, 'w', newline='') as f:
        writer = csv.writer(f)
        year_header = ['Year Last Modified', 'Count']
        writer.writerow(year_header)
        if len(year_count) > 0:
            for key, value in year_count.items():
                writer.writerow([key, value])

    cursor.execute("SELECT COUNT(DISTINCT format) as formats from siegfried WHERE format <> '';") # number of identfied file formats
    num_formats = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM siegfried WHERE errors <> '';") # number of siegfried errors
    num_errors = cursor.fetchone()[0]

    # calculate size from recursive dirwalk and format
    size_bytes = 0
    for root, dirs, files in os.walk(files_dir):
        for f in files:
            file_path = os.path.join(root, f)
            file_info = os.stat(file_path)
            size_bytes += file_info.st_size

    size = convert_size(size_bytes)
    
    # write html
    html.write('<!DOCTYPE html>')
    html.write('\n<html lang="en">')
    html.write('\n<head>')
    html.write('\n<title>IUL Born Digital Preservation Lab report: %s</title>' % item_barcode)
    html.write('\n<meta http-equiv="Content-Type" content="text/html; charset=utf-8">')
    html.write('\n<meta name="description" content="HTML report based upon a template developed by Tim Walsh and distributed as part of Brunnhilde v. 1.7.2">')
    html.write('\n<link rel="stylesheet" href="./assets//css/bootstrap.min.css">')
    html.write('\n</head>')
    html.write('\n<body style="padding-top: 80px">')
    # navbar
    html.write('\n<nav class="navbar navbar-expand-lg navbar-dark bg-dark fixed-top">')
    html.write('\n<a class="navbar-brand" href="#">Brunnhilde</a>')
    html.write('\n<button class="navbar-toggler" type="button" data-toggle="collapse" data-target="#navbarNavAltMarkup" aria-controls="navbarNavAltMarkup" aria-expanded="false" aria-label="Toggle navigation">')
    html.write('\n<span class="navbar-toggler-icon"></span>')
    html.write('\n</button>')
    html.write('\n<div class="collapse navbar-collapse" id="navbarNavAltMarkup">')
    html.write('\n<div class="navbar-nav">')
    html.write('\n<a class="nav-item nav-link" href="#Provenance">Provenance</a>')
    html.write('\n<a class="nav-item nav-link" href="#Stats">Statistics</a>')
    html.write('\n<a class="nav-item nav-link" href="#File formats">File formats</a>')
    html.write('\n<a class="nav-item nav-link" href="#File format versions">Versions</a>')
    html.write('\n<a class="nav-item nav-link" href="#MIME types">MIME types</a>')
    html.write('\n<a class="nav-item nav-link" href="#Last modified dates by year">Dates</a>')
    html.write('\n<a class="nav-item nav-link" href="#Unidentified">Unidentified</a>')
    html.write('\n<a class="nav-item nav-link" href="#Errors">Errors</a>')
    html.write('\n<a class="nav-item nav-link" href="#Duplicates">Duplicates</a>')
    html.write('\n<a class="nav-item nav-link" href="#Personally Identifiable Information (PII)">PII</a>')
    html.write('\n</div>')
    html.write('\n</div>')
    html.write('\n</nav>')
    # content
    html.write('\n<div class="container-fluid">')
    html.write('\n<h1 style="text-align: center; margin-bottom: 40px;">Brunnhilde HTML report</h1>')
    # provenance
    html.write('\n<a name="Provenance" style="padding-top: 40px;"></a>')
    html.write('\n<div class="container-fluid" style="margin-bottom: 40px;">')
    html.write('\n<div class="card">')
    html.write('\n<h2 class="card-header">Provenance</h2>')
    html.write('\n<div class="card-body">')
    '''need to check if disk image or not'''
    if jobType == 'Copy_only':
        html.write('\n<p><strong>Input source: File directory</strong></p>')
    else:
        html.write('\n<p><strong>Input source: Physical media</strong></p>')
    html.write('\n<p><strong>Accession/identifier:</strong> %s</p>' % item_barcode)
    html.write('\n</div>')
    html.write('\n</div>')
    html.write('\n</div>')
    # statistics
    html.write('\n<a name="Stats" style="padding-top: 40px;"></a>')
    html.write('\n<div class="container-fluid" style="margin-bottom: 40px;">')
    html.write('\n<div class="card">')
    html.write('\n<h2 class="card-header">Statistics</h2>')
    html.write('\n<div class="card-body">')
    html.write('\n<h4>Overview</h4>')
    html.write('\n<p><strong>Total files:</strong> %s (includes contents of archive files)</p>' % num_files)
    html.write('\n<p><strong>Total size:</strong> %s</p>' % size)
    html.write('\n<p><strong>Years (last modified):</strong> %s - %s</p>' % (begin_date, end_date))
    html.write('\n<p><strong>Earliest date:</strong> %s</p>' % earliest_date)
    html.write('\n<p><strong>Latest date:</strong> %s</p>' % latest_date)
    html.write('\n<h4>File counts and contents</h4>')
    html.write('\n<p><em>Calculated by hash value. Empty files are not counted in first three categories. Total files = distinct + duplicate + empty files.</em></p>')
    html.write('\n<p><strong>Distinct files:</strong> %s</p>' % distinct_files)
    html.write('\n<p><strong>Distinct files with duplicates:</strong> %s</p>' % distinct_dupes)
    html.write('\n<p><strong>Duplicate files:</strong> %s</p>' % duplicate_copies)
    html.write('\n<p><strong>Empty files:</strong> %s</p>' % empty_files)
    html.write('\n<h4>Format identification</h4>')
    html.write('\n<p><strong>Identified file formats:</strong> %s</p>' % num_formats)
    html.write('\n<p><strong>Unidentified files:</strong> %s</p>' % unidentified_files)
    html.write('\n<h4>Errors</h4>')
    html.write('\n<p><strong>Siegfried errors:</strong> %s</p>' % num_errors)
    html.write('\n<h2>Virus scan report</h2>')
    with open(os.path.join(log_dir, 'viruscheck-log.txt'), 'r', encoding='utf-8') as f:
        virus_report = f.read()
    html.write('\n<p>%s</p>' % virus_report)
    html.write('\n</div>')
    html.write('\n</div>')
    html.write('\n</div>')
    # detailed reports
    html.write('\n<div class="container-fluid" style="margin-bottom: 40px;">')
    html.write('\n<div class="card">')
    html.write('\n<h2 class="card-header">Detailed reports</h2>')
    html.write('\n<div class="card-body">')
    
    #save information to metadata_dict
    metadata_dict = pickleLoad('metadata_dict', folders, item_barcode)
    
    metadata_dict.update({'Source': item_barcode, 'begin_date': begin_date, 'end_date' : end_date, 'extent_normal': size, 'extent_raw': size_bytes, 'item_file_count': num_files, 'item_duplicate_count': distinct_dupes, 'FormatCount': num_formats, 'item_unidentified_count': unidentified_files})  
    
    pickleDump('metadata_dict', metadata_dict, folders)
    
def print_premis(premis_path, folders, item_barcode):   
    
    premis_list = pickleLoad('premis_list', folders, item_barcode)
    
    attr_qname = etree.QName("http://www.w3.org/2001/XMLSchema-instance", "schemaLocation")

    PREMIS_NAMESPACE = "http://www.loc.gov/premis/v3"

    PREMIS = "{%s}" % PREMIS_NAMESPACE

    NSMAP = {'premis' : PREMIS_NAMESPACE,
            "xsi": "http://www.w3.org/2001/XMLSchema-instance"}

    events = []
    
    #if our premis file already exists, we'll just delete it and write a new one
    if os.path.exists(premis_path):
        os.remove(premis_path)
        
    root = etree.Element(PREMIS + 'premis', {attr_qname: "http://www.loc.gov/premis/v3 https://www.loc.gov/standards/premis/premis.xsd"}, version="3.0", nsmap=NSMAP)
    
    object = etree.SubElement(root, PREMIS + 'object', attrib={etree.QName(NSMAP['xsi'], 'type'): 'premis:file'})
    objectIdentifier = etree.SubElement(object, PREMIS + 'objectIdentifier')
    objectIdentifierType = etree.SubElement(objectIdentifier, PREMIS + 'objectIdentifierType')
    objectIdentifierType.text = 'local'
    objectIdentifierValue = etree.SubElement(objectIdentifier, PREMIS + 'objectIdentifierValue')
    objectIdentifierValue.text = item_barcode
    objectCharacteristics = etree.SubElement(object, PREMIS + 'objectCharacteristics')
    compositionLevel = etree.SubElement(objectCharacteristics, PREMIS + 'compositionLevel')
    compositionLevel.text = '0'
    format = etree.SubElement(objectCharacteristics, PREMIS + 'format')
    formatDesignation = etree.SubElement(format, PREMIS + 'formatDesignation')
    formatName = etree.SubElement(formatDesignation, PREMIS + 'formatName')
    formatName.text = 'Tape Archive Format'
    formatRegistry = etree.SubElement(format, PREMIS + 'formatRegistry')
    formatRegistryName = etree.SubElement(formatRegistry, PREMIS + 'formatRegistryName')
    formatRegistryName.text = 'PRONOM'
    formatRegistryKey = etree.SubElement(formatRegistry, PREMIS + 'formatRegistryKey')
    formatRegistryKey.text = 'x-fmt/265' 

    for entry in premis_list:
        event = etree.SubElement(root, PREMIS + 'event')
        eventID = etree.SubElement(event, PREMIS + 'eventIdentifier')
        eventIDtype = etree.SubElement(eventID, PREMIS + 'eventIdentifierType')
        eventIDtype.text = 'UUID'
        eventIDval = etree.SubElement(eventID, PREMIS + 'eventIdentifierValue')
        eventIDval.text = str(uuid.uuid4())

        eventType = etree.SubElement(event, PREMIS + 'eventType')
        eventType.text = entry['eventType']

        eventDateTime = etree.SubElement(event, PREMIS + 'eventDateTime')
        eventDateTime.text = entry['timestamp']

        eventDetailInfo = etree.SubElement(event, PREMIS + 'eventDetailInformation')
        eventDetail = etree.SubElement(eventDetailInfo, PREMIS + 'eventDetail')
        eventDetail.text = entry['eventDetailInfo']
        
        #include additional eventDetailInfo to clarify action; older transfers may not include this element, so skip if KeyError
        try:
            eventDetailInfo = etree.SubElement(event, PREMIS + 'eventDetailInformation')
            eventDetail = etree.SubElement(eventDetailInfo, PREMIS + 'eventDetail')
            eventDetail.text = entry['eventDetailInfo_additional']
        except KeyError:
            pass
            
        eventOutcomeInfo = etree.SubElement(event, PREMIS + 'eventOutcomeInformation')
        eventOutcome = etree.SubElement(eventOutcomeInfo, PREMIS + 'eventOutcome')
        eventOutcome.text = str(entry['eventOutcomeDetail'])
        eventOutDetail = etree.SubElement(eventOutcomeInfo, PREMIS + 'eventOutcomeDetail')
        eventOutDetailNote = etree.SubElement(eventOutDetail, PREMIS + 'eventOutcomeDetailNote')
        if entry['eventOutcomeDetail'] == '0':
            eventOutDetailNote.text = 'Successful completion'
        elif entry['eventOutcomeDetail'] == 0:
            eventOutDetailNote.text = 'Successful completion'
        else:
            eventOutDetailNote.text = 'Unsuccessful completion; refer to logs.'

        linkingAgentID = etree.SubElement(event, PREMIS + 'linkingAgentIdentifier')
        linkingAgentIDtype = etree.SubElement(linkingAgentID, PREMIS + 'linkingAgentIdentifierType')
        linkingAgentIDtype.text = 'local'
        linkingAgentIDvalue = etree.SubElement(linkingAgentID, PREMIS + 'linkingAgentIdentifierValue')
        linkingAgentIDvalue.text = 'IUL BDPL'
        linkingAgentRole = etree.SubElement(linkingAgentID, PREMIS + 'linkingAgentRole')
        linkingAgentRole.text = 'implementer'
        linkingAgentID = etree.SubElement(event, PREMIS + 'linkingAgentIdentifier')
        linkingAgentIDtype = etree.SubElement(linkingAgentID, PREMIS + 'linkingAgentIdentifierType')
        linkingAgentIDtype.text = 'local'
        linkingAgentIDvalue = etree.SubElement(linkingAgentID, PREMIS + 'linkingAgentIdentifierValue')
        linkingAgentIDvalue.text = entry['linkingAgentIDvalue']
        linkingAgentRole = etree.SubElement(linkingAgentID, PREMIS + 'linkingAgentRole')
        linkingAgentRole.text = 'executing software'
        linkingObjectID = etree.SubElement(event, PREMIS + 'linkingObjectIdentifier')
        linkingObjectIDtype = etree.SubElement(linkingObjectID, PREMIS + 'linkingObjectIdentifierType')
        linkingObjectIDtype.text = 'local'
        linkingObjectIDvalue = etree.SubElement(linkingObjectID, PREMIS + 'linkingObjectIdentifierValue')
        linkingObjectIDvalue.text = item_barcode
    
    premis_tree = etree.ElementTree(root)
    
    premis_tree.write(premis_path, pretty_print=True, xml_declaration=True, encoding="utf-8")

def checkFiles(some_dir):
    #check to see if it exists
    if not os.path.exists(some_dir):
        print('\n\nError; folder "%s" does not exist.' % some_dir)
        return False
    
    #make sure there are files in the 'files' directory
    for dirpath, dirnames, contents in os.walk(some_dir):
        for file in contents:
            if os.path.isfile(os.path.join(dirpath, file)):
                return True
            else: 
                continue
            
    print('\n\nError; no files located at %s. Check settings and run again; you may need to manually copy or extract files.' % some_dir)
    return False

def md5(fname):
    hash_md5 = hashlib.md5()
    with open(fname, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_md5.update(chunk)
    return hash_md5.hexdigest()

def produce_dfxml(target, folders, item_barcode, jobType=None):
    dfxml_output = folders['dfxml_output']
    temp_dir = folders['temp_dir']
    imagefile = folders['imagefile']
    files_dir = folders['files_dir']
    
    timestamp = str(datetime.datetime.now())
    
    file_stats = []
    
    #use fiwalk if we have an image file
    if os.path.isfile(target):
        print('\n\nDIGITAL FORENSICS XML CREATION: FIWALK')
        dfxml_ver_cmd = 'fiwalk-0.6.3 -V'
        dfxml_ver = subprocess.check_output(dfxml_ver_cmd, shell=True, text=True).splitlines()[0]
        
        dfxml_cmd = 'fiwalk-0.6.3 -x %s > %s' % (target, dfxml_output)
        
        exitcode = subprocess.call(dfxml_cmd, shell=True, text=True)
                
        #Note that for DVD jobs we will also get stats on the files themselves later on; parse dfxml to get info for later; because large DFXML files pose a challenge; use iterparse to avoid crashing 
        print('\n\tCollecting file statistics...\n')
        counter = 0
        for event, element in etree.iterparse(dfxml_output, events = ("end",), tag="fileobject"):
            
            #refresh dict for each fileobject
            file_dict = {}
            
            #default values; will make sure that we don't record info about non-allocated files and that we have a default timestamp value
            good = True
            mt = False
            mtime = 'undated'
            target = ''
            size = ''
            checksum = ''
            
            for child in element:
                
                if child.tag == "filename":
                    target = child.text
                if child.tag == "name_type":
                    if child.text != "r":
                        element.clear()
                        good = False
                        break
                if child.tag == "alloc":
                    if child.text != "1":
                        good = False
                        element.clear()
                        break
                if child.tag == "unalloc":
                    if child.text == "1":
                        good = False
                        element.clear()
                        break
                if child.tag == "filesize":
                    size = child.text
                if child.tag == "hashdigest":
                    if child.attrib['type'] == 'md5':
                        checksum = child.text
                if child.tag == "mtime":
                    mtime = datetime.datetime.utcfromtimestamp(int(child.text)).isoformat()
                    mt = True
                if child.tag == "crtime" and mt == False:
                    mtime = datetime.datetime.utcfromtimestamp(int(child.text)).isoformat()
            
            if good and not '' in file_dict.values():
                file_dict = { 'name' : target, 'size' : size, 'mtime' : mtime, 'checksum' : checksum}
                file_stats.append(file_dict)
                
                counter+=1            
                print('\r\tWorking on file #: %s' % counter, end='')

            element.clear()
 
    #use custom operation for other cases    
    elif os.path.isdir(target):
        print('\n\nDIGITAL FORENSICS XML CREATION: bdpl_ingest')
        
        dfxml_ver = 'https://github.com/IUBLibTech/bdpl_ingest'
        dfxml_cmd = 'bdpl_ingest.py'
        
        timestamp = str(datetime.datetime.now().isoformat())
        
        done_list = []
        
        temp_dfxml = os.path.join(temp_dir, 'temp_dfxml.txt')
        if os.path.exists(temp_dfxml):
            with open(temp_dfxml, 'r', encoding='utf-8') as f:
                done_so_far = f.read().splitlines()
                for d in done_so_far:
                    line = d.split(' | ')
                    done_list.append(line[0])
                    file_dict = { 'name' : line[0], 'size' : line[1], 'mtime' : line[2], 'ctime' : line[3], 'atime' : line[4], 'checksum' : line[5], 'counter' : line[6] }
                    file_stats.append(file_dict)
        
        if len(file_stats) > 0:
            counter = int(file_stats[-1]['counter'])
        else:
            counter = 0
        
        print('\n')
        
        #get total number of files
        total = sum([len(files) for r, d, files in os.walk(target)])
        
        #now loop through, keeping count
        for root, dirnames, filenames in os.walk(target):
            for file in filenames:
                
                #check to make sure that we haven't already added info for this file
                file_target = os.path.join(root, file)
                if file_target in done_list:
                    continue
                
                counter += 1
                
                size = os.path.getsize(file_target)
                mtime = datetime.datetime.fromtimestamp(os.path.getmtime(file_target)).isoformat()
                ctime = datetime.datetime.fromtimestamp(os.path.getctime(file_target)).isoformat()
                atime = datetime.datetime.fromtimestamp(os.path.getatime(file_target)).isoformat()[:-7]
                checksum = md5(file_target)
                
                file_dict = { 'name' : file_target, 'size' : size, 'mtime' : mtime, 'ctime' : ctime, 'atime' : atime, 'checksum' : checksum, 'counter' : counter }
                
                print('\r\tCalculating checksum for file %d out of %d' % (counter, total), end='')
                
                file_stats.append(file_dict)
                
                #add this file to our 'done list'
                done_list.append(file_target)
                
                #save this list to file just in case we crash...
                raw_stats = "%s | %s | %s | %s | %s | %s | %s\n" % (file_target, size, mtime, ctime, atime, checksum, counter)
                with open(temp_dfxml, 'a', encoding='utf8') as f:
                    f.write(raw_stats)
        print('\n')
        
        dc_namespace = 'http://purl.org/dc/elements/1.1/'
        dc = "{%s}" % dc_namespace
        NSMAP = {None : 'http://www.forensicswiki.org/wiki/Category:Digital_Forensics_XML',
                'xsi': "http://www.w3.org/2001/XMLSchema-instance",
                'dc' : dc_namespace}

        dfxml = etree.Element("dfxml", nsmap=NSMAP, version="1.0")
        
        metadata = etree.SubElement(dfxml, "metadata")
        
        dctype = etree.SubElement(metadata, dc + "type")
        dctype.text = "Hash List"
        
        creator = etree.SubElement(dfxml, 'creator')
        
        program = etree.SubElement(creator, 'program')
        program.text = 'bdpl_ingest'
        
        execution_environment = etree.SubElement(creator, 'execution_environment')
        
        start_time = etree.SubElement(execution_environment, 'start_time')
        start_time.text = timestamp
        
        for f in file_stats:
            fileobject = etree.SubElement(dfxml, 'fileobject')
            
            filename = etree.SubElement(fileobject, 'filename')
            filename.text = f['name']
            
            filesize = etree.SubElement(fileobject, 'filesize')
            filesize.text = str(f['size'])

            modifiedtime = etree.SubElement(fileobject, 'mtime')
            modifiedtime.text = f['mtime']
        
            createdtime = etree.SubElement(fileobject, 'ctime')
            createdtime.text = f['ctime']
            
            accesstime = etree.SubElement(fileobject, 'atime')
            accesstime.text = f['atime']
                
            hashdigest = etree.SubElement(fileobject, 'hashdigest', type='md5')
            hashdigest.text = f['checksum']

        tree = etree.ElementTree(dfxml)
        
        tree.write(dfxml_output, pretty_print=True, xml_declaration=True, encoding="utf-8")      
    
    else:
        print('\n\tERROR: %s does not appear to exist...' % target)
        return
    
    #save stats for reporting...
    if jobType in ['DVD', 'CDDA']:
        checksums = os.path.join(temp_dir, 'checksums_di.txt')
    else:
        checksums = os.path.join(temp_dir, 'checksums.txt')
    with open (checksums, 'wb') as f:
        pickle.dump(file_stats, f)
    
    #save PREMIS
    premis_list = pickleLoad('premis_list', folders, item_barcode)        
    premis_list.append(premis_dict(timestamp, 'message digest calculation', 0, dfxml_cmd, 'Extracted information about the structure and characteristics of content, including file checksums.', dfxml_ver))
    pickleDump('premis_list', premis_list, folders)
    
    print('\n\n\tDFXML creation completed; moving on to next step...')

def optical_drive_letter():
    drive_cmd = 'wmic logicaldisk get caption, drivetype | FINDSTR /C:"5"'
    drive_ltr = subprocess.check_output(drive_cmd, shell=True, text=True).split()[0]
    return drive_ltr

def disk_image_info(folders, item_barcode):
    imagefile = folders['imagefile']
    reports_dir = folders['reports_dir']
    
    premis_list = pickleLoad('premis_list', folders, item_barcode) 
    
    print('\n\nDISK IMAGE METADATA EXTRACTION: FSSTAT, ILS, MMLS')
    
    #run disktype to get information on file systems on disk
    disktype_output = os.path.join(reports_dir, 'disktype.txt')
    disktype_command = 'disktype %s > %s' % (imagefile, disktype_output)
        
    timestamp = str(datetime.datetime.now())
    exitcode = subprocess.call(disktype_command, shell=True, text=True)
    premis_list.append(premis_dict(timestamp, 'forensic feature analysis', exitcode, disktype_command, 'Determined disk image file system information.', 'disktype v9'))
    
    #take disktype output; print to screen and get a list  of all partition information
    charenc = get_encoding(disktype_output)
    
    with open(disktype_output, 'r', encoding=charenc) as f:
        print(f.read(), end="")
    
    with open(disktype_output, 'r', encoding=charenc) as f:
        dt_info = f.read().split('Partition ')

    fs_list = []
    for dt in dt_info:
        if 'file system' in dt:
            fs_list.append([d for d in dt.split('\n') if ' file system' in d][0].split(' file system')[0].lstrip().lower())
    
    #save file system list for later...
    pickleDump('fs_list', fs_list, folders)
    
    #run fsstat: get range of meta-data values (inode numbers) and content units (blocks or clusters)
    fsstat_output = os.path.join(reports_dir, 'fsstat.txt')
    fsstat_ver = 'fsstat: %s' % subprocess.check_output('fsstat -V', shell=True, text=True).strip()
    fsstat_command = 'fsstat %s > %s 2>&1' % (imagefile, fsstat_output)
    
    timestamp = str(datetime.datetime.now())
    try:
        exitcode = subprocess.call(fsstat_command, shell=True, text=True, timeout=60)   
    except subprocess.TimeoutExpired:
        #if there was output before timeout, then rerun command
        if os.path.getsize(fsstat_output) > 0:
            exitcode = subprocess.call(fsstat_command, shell=True, text=True)
        #if the command did nothing, kill the process and report as a failure
        else:
            for proc in psutil.process_iter():
                if proc.name() == 'fsstat.exe':
                    psutil.Process(proc.pid).terminate()
            exitcode = 1
        
    premis_list.append(premis_dict(timestamp, 'forensic feature analysis', exitcode, fsstat_command, 'Determined range of meta-data values (inode numbers) and content units (blocks or clusters)', fsstat_ver))

    #run ils to document inode information
    ils_output = os.path.join(reports_dir, 'ils.txt')
    ils_ver = 'ils: %s' % subprocess.check_output('ils -V', shell=True, text=True).strip()
    ils_command = 'ils -e %s > %s 2>&1' % (imagefile, ils_output)
    
    timestamp = str(datetime.datetime.now())
    try:
        exitcode = subprocess.call(ils_command, shell=True, text=True, timeout=60)
    except subprocess.TimeoutExpired:
        #if there was output before timeout, then rerun command
        if os.path.getsize(ils_output) > 0:
            exitcode = subprocess.call(ils_command, shell=True, text=True)
        #if the command did nothing, kill the process and report as a failure
        else:
            for proc in psutil.process_iter():
                if proc.name() == 'ils.exe':
                    psutil.Process(proc.pid).terminate()
            exitcode = 1
    
    premis_list.append(premis_dict(timestamp, 'forensic feature analysis', exitcode, ils_command, 'Documented all inodes found on disk image.', ils_ver))
    
    #run mmls to document the layout of partitions in a volume system
    mmls_output = os.path.join(reports_dir, 'mmls.txt')
    mmls_ver = 'mmls: %s' % subprocess.check_output('mmls -V', shell=True, text=True).strip()
    mmls_command = 'mmls %s > %s 2>NUL' % (imagefile, mmls_output)
    
    timestamp = str(datetime.datetime.now())
    exitcode = subprocess.call(mmls_command, shell=True, text=True) 
    premis_list.append(premis_dict(timestamp, 'forensic feature analysis', exitcode, mmls_command, 'Determined the layout of partitions in a volume system.', mmls_ver))
    
    #check mmls output for partition information
    partition_info_list = []
    if os.stat(mmls_output).st_size > 0:
        
        with open(mmls_output, 'r', encoding='utf8') as f:
            mmls_info = [m.split('\n') for m in f.read().splitlines()[5:]] 
        
        for mm in mmls_info:
            temp = {}
            for dt in dt_info:
                if 'file system' in dt and ', {} sectors from {})'.format(mm[0].split()[4].lstrip('0'), mm[0].split()[2].lstrip('0')) in dt:
                    fsname = [d for d in dt.split('\n') if ' file system' in d][0].split(' file system')[0].lstrip().lower()
                    temp['start'] = mm[0].split()[2]
                    temp['desc'] = fsname
                    temp['slot'] = mm[0].split()[1]
                    #now save this dictionary to our list of partition info
                    if not temp in partition_info_list:
                        partition_info_list.append(temp)
            
        pickleDump('partition_info_list', partition_info_list, folders)
                                
    pickleDump('premis_list', premis_list, folders)

def dir_tree(folders, item_barcode):
        
    #make a directory tree to document original structure
    target = folders['files_dir']
    reports_dir = folders['reports_dir']
    tree_dest = os.path.join(reports_dir, 'tree.txt')
    
    tree_ver = subprocess.check_output('tree --version', shell=True, text=True).split(' (')[0]
    tree_command = 'tree.exe -tDhR "%s" > "%s"' % (target, tree_dest)
    
    timestamp = str(datetime.datetime.now())
    exitcode = subprocess.call(tree_command, shell=True, text=True)
    
    premis_list = pickleLoad('premis_list', folders, item_barcode)
    premis_list.append(premis_dict(timestamp, 'metadata extraction', exitcode, tree_command, 'Documented the organization and structure of content within a directory tree.', tree_ver))
    pickleDump('premis_list', premis_list, folders)
    
    print('\n\tDirectory structure documented; moving on to next step...')

def droid_to_siegfried(infile, outfile):

    counter = 0

    with open(outfile, 'w', newline='') as f1:
        csvWriter = csv.writer(f1)
        header = ['filename', 'filesize', 'modified', 'errors', 'namespace', 'id', 'format', 'version', 'mime', 'basis', 'warning']
        csvWriter.writerow(header)
        with open(infile, 'r', encoding='utf8') as f2:
            csvReader = csv.reader(f2)
            next(csvReader)
            for row in csvReader:
                counter+=1
                print('\rWorking on row %d' % counter, end='')
                
                if 'zip:file:' in row[2]:
                    filename = row[2].split('zip:file:/', 1)[1].replace('.zip!', '.zip#').replace('/', '\\')
                else:
                    filename = row[2].split('file:/', 1)[1]
                filename = unquote(filename)
                
                filesize = row[7]
                modified = row[10]
                errors = ''
                namespace = 'pronom'
                if row[14] == "":
                    id = 'UNKNOWN'
                else:
                    id = row[14]
                format = row[16]
                version = row[17]
                mime = row[15]
                basis = ''
                if row[11].lower() == 'true':
                    warning = 'extension mismatch'
                else:
                    warning = ''
                
                data = [filename, filesize, modified, errors, namespace, id, format, version, mime, basis, warning]
                
                csvWriter.writerow(data)

def format_analysis(folders, item_barcode):
    
    files_dir = folders['files_dir']
    reports_dir = folders['reports_dir']
    log_dir = folders['log_dir']
    temp_dir = folders['temp_dir']
    
    print('\n\tFile format identification with siegfried...') 

    sfcmd = 'sf -version'
    format_version = subprocess.check_output(sfcmd, shell=True, text=True).replace('\n', ' ')
    
    sf_file = os.path.join(reports_dir, 'siegfried.csv')
    format_command = 'sf -z -csv "%s" > "%s"' % (files_dir, sf_file)
    
    #create timestamp
    timestamp = str(datetime.datetime.now())
    
    if os.path.exists(sf_file):
        os.remove(sf_file)                                                                 
    
    exitcode = subprocess.call(format_command, shell=True, text=True)
    
    #if siegfried fails, then we'll run DROID
    if exitcode != 0 and os.path.getsize(sf_file) == 0:
        print('\n\tFile format identification with siegfried failed; now attempting with DROID...\n') 
        
        droid_profile = os.path.join(temp_dir, 'droid.droid')
        droid_out = os.path.join(temp_dir, 'droid.csv')
        
        droid_ver = "DROID v%s" % subprocess.check_output('droid -v', shell=True, text=True).strip()
        
        cmd = 'droid -RAq -a "%s" -p "%s"' % (files_dir, droid_profile)
        
        exitcode = subprocess.call(cmd, shell=True)
        
        cmd2 = 'droid -p "%s" -e "%s"' % (droid_profile, droid_out)
        
        subprocess.call(cmd2, shell=True)
        
        #consolidate commands for premis
        format_command = "%s && %s" % (cmd, cmd2)
        
        #now reformat droid output to be like sf output
        droid_to_siegfried(droid_out, sf_file)
        
    premis_list = pickleLoad('premis_list', folders, item_barcode)
    premis_list.append(premis_dict(timestamp, 'format identification', exitcode, format_command, 'Determined file format and version numbers for content recorded in the PRONOM format registry.', format_version))
    pickleDump('premis_list', premis_list, folders)

def stats_and_report_creation(folders, item_barcode, re_analyze, jobType):
    
    temp_dir = folders['temp_dir']
    files_dir = folders['files_dir']
    reports_dir = folders['reports_dir']
    log_dir = folders['log_dir']
    
    #set up html for report
    temp_html = os.path.join(temp_dir, 'temp.html')
    html = open(temp_html, 'w', encoding='utf8')
    
    #prepare sqlite database and variables
    siegfried_db = os.path.join(temp_dir, 'siegfried.sqlite')
    conn = sqlite3.connect(siegfried_db)
    conn.text_factory = str  # allows utf-8 data to be stored
    cursor = conn.cursor() 
    
    get_stats(folders, cursor, html, item_barcode, re_analyze, jobType) # get aggregate stats and write to html file
    generate_reports(cursor, html, folders, re_analyze, item_barcode) # run sql queries, print to html and csv
    close_html(html) # close HTML file tags
    
    # close database connections
    cursor.close()
    conn.close()
    
    print('\n\tFormat analysis completed!')
    
    # close HTML file
    html.close()

    # write new html file, with hrefs for PRONOM IDs   
    new_html = os.path.join(reports_dir, 'report.html')
    
    if os.path.exists(new_html):
        os.remove(new_html)

    write_pronom_links(temp_html, new_html)

    # get format list and add to metadata dictionary
    metadata_dict = pickleLoad('metadata_dict', folders, item_barcode)
    
    fileformats = []
    formatcount = 0
    formatlist = ''
    formatcsv = os.path.join(reports_dir, 'formats.csv')
    try:
        with open(formatcsv, 'r') as csvfile:
            formatreader = csv.reader(csvfile)
            next(formatreader)
            for row in formatreader:
                formatcount += 1
                fileformats.append(row[0])
            fileformats = [element or 'Unidentified' for element in fileformats] # replace empty elements with 'Unidentified'
            if formatcount > 0:
                metadata_dict['format_overview'] = "Top file formats (out of %s total) are: %s" % (formatcount, ' | '.join(fileformats[:10]))
            else:
                metadata_dict['format_overview'] = "-"
            
    except IOError:
        metadata_dict['format_overview'] = "ERROR! No formats.csv file to pull formats from."
            
    pickleDump('metadata_dict', metadata_dict, folders)

def analyzeContent(unit_name, shipmentDate, item_barcode, analysis_vars):
    
    folders = bdpl_folders(unit_name, shipmentDate, item_barcode)
    
    files_dir = folders['files_dir']
    log_dir = folders['log_dir']
    metadata = folders['metadata']
    reports_dir = folders['reports_dir']
    imagefile = folders['imagefile']
    files_dir = folders['files_dir']
    bulkext_dir = folders['bulkext_dir']
    bulkext_log = folders['bulkext_log']
    temp_dir = folders['temp_dir']
    image_dir = folders['image_dir']
    dfxml_output = folders['dfxml_output']

    platform = analysis_vars['platform']    
    if platform == 'bdpl_ingest':
        jobType = analysis_vars['jobType'].get()
        re_analyze = analysis_vars['re_analyze'].get()
        gui_vars = analysis_vars['gui_vars']
    else:
        jobType = analysis_vars['jobType']
        re_analyze = analysis_vars['re_analyze']
        gui_vars = analysis_vars['gui_vars']
    
    if platform == 'bdpl_ingest':
        newscreen()
    
        #if information not 'verified' then go into 'first run'; exit if anything is wrong
        if not verify_data(unit_name, shipmentDate, item_barcode):
            return

    print('\n\n--------------------------------------------------------------------------------------------------\n\nSTEP 2: CONTENT ANALYSIS') 

    #return if no job type is selected
    if jobType not in ['Disk_image', 'Copy_only', 'DVD', 'CDDA']:
        print('\n\nError; please indicate the appropriate job type')
        return
        
    '''copy .css and .jc files to assets directory'''
    assets_dir = os.path.join('C:\\BDPL\\resources\\assets')
    assets_target = os.path.join(reports_dir, 'assets')
    if os.path.exists(assets_target):
        pass
    else:
        shutil.copytree(assets_dir, assets_target)
                                                                                                                              
    '''run antivirus scan using clamscan; skip if virus scan already run'''
    print('\nVIRUS SCAN: clamscan.exe')
    if check_premis('virus check', folders, item_barcode) and not re_analyze:
        print('\n\tVirus scan already completed; moving on to next step...')
    else:
        run_antivirus(folders, item_barcode)
    
    '''create DFXML (if not already done so)'''
    if check_premis('message digest calculation', folders, item_barcode) and not re_analyze:
        print('\n\nDIGITAL FORENSICS XML CREATION:')
        print('\n\tDFXML already created; moving on to next step...')
    else:
        if jobType == 'Disk_image':
            #DFXML creation for disk images will depend on the image's file system; check fs_list
            fs_list = pickleLoad('fs_list', folders, item_barcode)
            
            #if it's an HFS+ file system, we can use fiwalk on the disk image; otherwise, use bdpl_ingest on the file directory
            if 'hfs+' in [fs.lower() for fs in fs_list]:
                produce_dfxml(imagefile, folders, item_barcode)
            else:
                produce_dfxml(files_dir, folders, item_barcode)
        
        elif jobType == 'Copy_only':
            produce_dfxml(files_dir, folders, item_barcode)
        
        elif jobType == 'DVD':
            produce_dfxml(imagefile, folders, item_barcode, jobType)
        
        elif jobType == 'CDDA':
            produce_dfxml(image_dir, folders, item_barcode, jobType)
        
    '''document directory structure'''
    print('\n\nDOCUMENTING FOLDER/FILE STRUCTURE: TREE')
    if check_premis('metadata extraction', folders, item_barcode) and not re_analyze:
        print('\n\tDirectory structure already documented with tree command; moving on to next step...')
    else:
        dir_tree(folders, item_barcode) 
    
    '''run bulk_extractor to identify potential sensitive information (only if disk image or copy job type). Skip if b_e was run before'''
    print('\n\nSENSITIVE DATA SCAN: BULK_EXTRACTOR')
    if check_premis('sensitive data scan', folders, item_barcode) and not re_analyze:
        print('\n\tSensitive data scan already completed; moving on to next step...')
    else:
        if jobType in ['Copy_only', 'Disk_image']:
            run_bulkext(folders, item_barcode)
        else:
            print('\n\tSensitive data scan not required for DVD-Video or CDDA content; moving on to next step...')
            
    '''run siegfried to characterize file formats'''
    print('\n\nFILE FORMAT ANALYSIS')
    if check_premis('format identification', folders, item_barcode) and not re_analyze:
        print('\n\tFile format analysis already completed; moving on to next operation...')
    else:
        format_analysis(folders, item_barcode)
    
    #load siegfried.csv into sqlite database; skip if it's already completed
    if not os.path.exists(os.path.join(temp_dir, 'sqlite_done.txt')) or re_analyze:
        import_csv(folders) # load csv into sqlite db
    
    '''create HTML and CSV reports'''
    stats_and_report_creation(folders, item_barcode, re_analyze, jobType)
    
    #generate PREMIS preservation metadata file
    premis_path = os.path.join(metadata, '%s-premis.xml' % item_barcode)
    print_premis(premis_path, folders, item_barcode)
    
    #write info to spreadsheet for collecting unit to review
    if not writeSpreadsheet(folders, unit_name, shipmentDate, item_barcode, gui_vars, jobType, platform):
        return
       
    #create file to indicate that process was completed
    done_file = os.path.join(temp_dir, 'done.txt')
    if not os.path.exists(done_file):
        open(done_file, 'a').close()
        
    '''clean up; delete disk image folder if empty and remove temp_html'''
    try:
        os.rmdir(image_dir)
    except (WindowsError, PermissionError):
        pass

    # remove temp html file
    try:
        os.remove(os.path.join(temp_dir, 'temp.html'))
    except WindowsError:
        pass
    
    '''if using gui, print final details about item'''
    if analysis_vars['platform'] == 'bdpl_ingest':
        print('\n\n--------------------------------------------------------------------------------------------------\n\nINGEST PROCESS COMPLETED FOR ITEM %s\n\nResults:\n' % item_barcode)
        
        du_cmd = 'du64.exe -nobanner "%s" > %s' % (files_dir, os.path.join(temp_dir, 'final_stats.txt'))
        
        subprocess.call(du_cmd, shell=True, text=True)   
        
        if os.path.exists(image_dir):
            di_count = len(os.listdir(image_dir))
            if di_count > 0:
                print('Disk Img(s):   %s' % di_count)
        du_list = ['Files:', 'Directories:', 'Size:', 'Size on disk:']
        with open(os.path.join(temp_dir, 'final_stats.txt'), 'r') as f:
            for line, term in zip(f.readlines(), du_list):
                if "Directories:" in term:
                    print(term, ' ', str(int(line.split(':')[1]) - 1).rstrip())
                else: 
                    print(term, line.split(':')[1].rstrip())
        print('\n\n')      

def return_spreadsheet_row(ws, item_barcode):
 
    newrow = ws.max_row+1
    found = False
    
    #if barcode exists in spreadsheet, set variable to that row
    for cell in ws['A']:
        if (cell.value is not None):
            if item_barcode == str(cell.value).strip():
                newrow = cell.row
                found = True
                break
    
    return found, newrow
    
def writeNote(unit_name, shipmentDate, item_barcode, gui_vars):

    folders = bdpl_folders(unit_name, shipmentDate, item_barcode)
    ship_dir = folders['ship_dir']
    temp_dir = folders['temp_dir']
    
    if not verify_data(unit_name, shipmentDate, item_barcode):
        return
    
    metadata_dict = pickleLoad('metadata_dict', folders, item_barcode)
    metadata_dict['label_transcription'] = gui_vars['label_transcription'].get(1.0, END).replace('LABEL TRANSCRIPTION:\n\n', '')
    metadata_dict['technician_note'] = gui_vars['technician_note'].get(1.0, END)
    
    #additional steps if we are noting failed transfer of item...
    if gui_vars['noteFail'].get() == True:
        metadata_dict['migration_date'] = str(datetime.datetime.now())
        metadata_dict['migration_outcome'] = "Failure"
        
        done_file = os.path.join(temp_dir, 'done.txt')
        if not os.path.exists(done_file):
            open(done_file, 'a').close()
    
    #save our metadata, just in case...
    pickleDump('metadata_dict', metadata_dict, folders)
    
    #get spreadsheet
    spreadsheet = find_spreadsheet(folders, unit_name, shipmentDate)
    if not os.path.exists(spreadsheet):
        print('\nWARNING: unable to find spreadsheet.  Check shipment directory...')
        return
        
    wb = openpyxl.load_workbook(spreadsheet)
    app_ws = wb['Appraisal']

    #use function to see if barcode is already recorded; if not, use next available row
    status, current_row = return_spreadsheet_row(app_ws, item_barcode)
    
    #get all column #s
    ws_cols = get_spreadsheet_columns(app_ws)
    
    for key in ws_cols.keys():
        if key in metadata_dict:
            app_ws.cell(row=current_row, column=ws_cols[key], value=metadata_dict[key])

    #save and close spreadsheet
    wb.save(spreadsheet)
    
    print('\n\nInformation saved to Appraisal worksheet.') 
    
def writeSpreadsheet(folders, unit_name, shipmentDate, item_barcode, gui_vars, jobType, platform):

    premis_list = pickleLoad('premis_list', folders, item_barcode)
            
    metadata_dict = pickleLoad('metadata_dict', folders, item_barcode)
    
    #get additional metadata from PREMIS about transfer
    if jobType in ['Disk_image', 'DVD', 'CDDA']:
        try:
            temp_dict = [f for f in premis_list if f['eventType'] == 'disk image creation'][-1]
        except IndexError:
            try: 
                temp_dict = [f for f in premis_list if f['eventType'] == 'normalization'][-1]
            except IndexError:
                temp_dict = {'linkingAgentIDvalue' : '-', 'timestamp' : '-', 'eventOutcomeDetail' : 'Operation not completed.'}
    elif jobType == 'Copy_only':
        try:
            temp_dict = [f for f in premis_list if f['eventType'] == 'replication'][-1]
        except IndexError:
            temp_dict = {'linkingAgentIDvalue' : '-', 'timestamp' : '-', 'eventOutcomeDetail' : 'Operation not completed.'}
    
    metadata_dict['jobType'] = jobType
    metadata_dict['transfer_method'] = temp_dict['linkingAgentIDvalue']
    metadata_dict['migration_date'] = temp_dict['timestamp']
    
    if temp_dict['eventOutcomeDetail'] == '0' or temp_dict['eventOutcomeDetail'] == 0:
        metadata_dict['migration_outcome'] = 'Success'
    else:
        metadata_dict['migration_outcome'] = 'Failure'
    
    #if using the GUI ingest tool, update any notes provided by technician
    if platform == 'bdpl_ingest':
        metadata_dict['label_transcription'] = gui_vars['label_transcription'].get(1.0, END).replace('LABEL TRANSCRIPTION:\n\n', '')
        metadata_dict['technician_note'] = gui_vars['technician_note'].get(1.0, END)
    
    #add linked information
    metadata_dict['full_report'] = '=HYPERLINK("{}", "{}")'.format(".\\%s\\metadata\\reports\\report.html" % item_barcode, "View report")
    metadata_dict['transfer_link'] = '=HYPERLINK("{}", "{}")'.format(".\\%s" % item_barcode, "View transfer folder")
    
    try:
        if metadata_dict['initial_appraisal'] == "No appraisal needed":
            metadata_dict['initial_appraisal'] = "Transfer to SDA"
        elif metadata_dict['initial_appraisal'] == '-':
            del metadata_dict['initial_appraisal']
    except KeyError:
        pass
        
    #write back metadata, just in case...
    pickleDump('metadata_dict', metadata_dict, folders)
    
    #get spreadsheet
    spreadsheet = find_spreadsheet(folders, unit_name, shipmentDate)
    if not os.path.exists(spreadsheet):
        return False
        
    wb = openpyxl.load_workbook(spreadsheet)
    app_ws = wb['Appraisal']
        
    #use function to see if barcode is already recorded; if not, use next available row
    status, current_row = return_spreadsheet_row(app_ws, item_barcode)
    
    ws_cols = get_spreadsheet_columns(app_ws)
    
    for key in ws_cols.keys():
        if key in metadata_dict:
            app_ws.cell(row=current_row, column=ws_cols[key], value=metadata_dict[key])

    #save and close spreadsheet
    wb.save(spreadsheet)   

    return True
        
def cleanUp(cleanUp_vars):
    
    newscreen()
    
    #deselect all radio buttons
    try:
        for name, button in cleanUp_vars['radio_buttons'].items():
            button.deselect()
    except NameError as e:
        pass
        
    #clear string variables
    for name, str in cleanUp_vars['str_vars'].items():
        if name == 'disk525':
            str.set('N/A')
        elif name == 'jobType':
            str.set(None)
        elif name == 'sourceDevice':
            str.set(None)
        else:
            str.set('')
    
    #reset checkboxes
    for checkbox in cleanUp_vars['checkboxes'].values():
        checkbox.set(False)
    
    #reset text widgets
    for name, widget in cleanUp_vars['text_widgets'].items():
        widget.configure(state='normal')
        widget.delete('1.0', END)
        if name == 'bdpl_instructions':
            widget.insert(INSERT, "TECHNICIAN NOTES:\n")
        elif name == 'appraisal_notes':
            widget.insert(INSERT, "APPRAISAL NOTES:\n")
        elif name == 'label_transcription':
            widget.insert(INSERT, "LABEL TRANSCRIPTION:\n")
    
    #reset entry widgets
    for widget in cleanUp_vars['entry_widgets'].values():
        widget.delete(0, END)        

def closeUp(window):    
    
    try:
        close_files_conns_on_exit(html, conn, cursor)
    except (NameError, sqlite3.ProgrammingError) as e:
        pass
    
    window.destroy()

def verify_data(unit_name, shipmentDate, item_barcode, check=None):
    
    if unit_name == '':
        print('\n\nError; please make sure you have entered a unit ID abbreviation.')
        return False 
    
    if item_barcode == '':
        if check in ['check_progress', 'check_unfinished']:
            pass
        else:
            print('\n\nError; please make sure you have entered a barcode.')
            return False 
    
    if shipmentDate == '':
        print('\n\nError; please make sure you have entered a shipment date.')
        return False
    else:
        ship_dir = os.path.join('Z:\\', unit_name, 'ingest', shipmentDate)
        if not os.path.exists(ship_dir):
            print('\n\nError; the shipment directory %s does not exist.\n\tPlease check the unit folder to identify the correct shipment folder.')
            return False
    
    #if we get through all the above, then we are good to go!
    return True

def get_spreadsheet_columns(ws):

    spreadsheet_columns = {}
    
    for cell in ws[1]:
        if not cell.value is None:
        
            if 'identifier' in str(cell.value).lower():
                spreadsheet_columns['item_barcode'] = cell.column
                
            elif 'accession' in cell.value.lower():
                spreadsheet_columns['current_accession'] = cell.column
                
            elif 'collection title' in cell.value.lower():
                spreadsheet_columns['collection_title'] = cell.column
                
            elif 'collection id' in cell.value.lower():
                spreadsheet_columns['current_coll_id'] = cell.column
                
            elif 'creator' in cell.value.lower():
                spreadsheet_columns['collection_creator'] = cell.column
                
            elif 'physical location' in cell.value.lower():
                spreadsheet_columns['phys_loc'] = cell.column
                
            elif 'source type' in cell.value.lower():
                spreadsheet_columns['content_source_type'] = cell.column
                
            elif cell.value.strip().lower() == 'title':
                spreadsheet_columns['item_title'] = cell.column
                
            elif 'label transcription' in cell.value.lower():
                spreadsheet_columns['label_transcription'] = cell.column
                
            elif cell.value.strip().lower() == 'description':
                spreadsheet_columns['item_description'] = cell.column
                
            elif 'initial appraisal notes' in cell.value.lower():
                spreadsheet_columns['appraisal_notes'] = cell.column
                
            elif 'content date range' in cell.value.lower():
                spreadsheet_columns['assigned_dates'] = cell.column
                
            elif 'instructions' in cell.value.lower():
                spreadsheet_columns['bdpl_instructions'] = cell.column
                
            elif 'restriction statement' in cell.value.lower():
                spreadsheet_columns['restriction_statement'] = cell.column
                
            elif 'restriction end date' in cell.value.lower():
                spreadsheet_columns['restriction_end_date'] = cell.column
                
            elif 'sda' in cell.value.lower():
                spreadsheet_columns['initial_appraisal'] = cell.column
                
            elif 'transfer method' in cell.value.lower():
                spreadsheet_columns['transfer_method'] = cell.column
                
            elif 'migration date' in cell.value.lower():
                spreadsheet_columns['migration_date'] = cell.column
                
            elif 'migration notes' in cell.value.lower():
                spreadsheet_columns['technician_note'] = cell.column
                
            elif 'migration outcome' in cell.value.lower():
                spreadsheet_columns['migration_outcome'] = cell.column
                
            elif 'extent (normalized)' in cell.value.lower():
                spreadsheet_columns['extent_normal'] = cell.column
                
            elif 'extent (raw)' in cell.value.lower():
                spreadsheet_columns['extent_raw'] = cell.column
                
            elif 'no. of files' in cell.value.lower():
                spreadsheet_columns['item_file_count'] = cell.column
                
            elif 'no. of duplicate files' in cell.value.lower():
                spreadsheet_columns['item_duplicate_count'] = cell.column
                
            elif 'no. of unidentified files' in cell.value.lower():
                spreadsheet_columns['item_unidentified_count'] = cell.column
                
            elif 'file formats' in cell.value.lower():
                spreadsheet_columns['format_overview'] = cell.column
                
            elif 'begin date' in cell.value.lower():
                spreadsheet_columns['begin_date'] = cell.column
                
            elif 'end date' in cell.value.lower():
                spreadsheet_columns['end_date'] = cell.column
                
            elif 'virus status' in cell.value.lower():
                spreadsheet_columns['virus_scan_results'] = cell.column
                
            elif 'pii status' in cell.value.lower():
                spreadsheet_columns['pii_scan_results'] = cell.column
                
            elif 'full report' in cell.value.lower():
                spreadsheet_columns['full_report'] = cell.column
                
            elif 'link to transfer' in cell.value.lower():
                spreadsheet_columns['transfer_link'] = cell.column
                
            elif 'appraisal results' in cell.value.lower():
                spreadsheet_columns['initial_appraisal'] = cell.column
                
            elif 'job type' in cell.value.lower():
                spreadsheet_columns['jobType'] = cell.column
    
    return spreadsheet_columns    

def load_metadata(folders, item_barcode, spreadsheet):

    ship_dir = folders['ship_dir']
    metadata = folders['metadata']
    temp_dir = folders['temp_dir']
    
    wb = openpyxl.load_workbook(spreadsheet)
    inv_ws = wb['Inventory']
    app_ws = wb['Appraisal']

    #Find the barcode in the inventory sheet; save information to a dictionary so that it can be written to the Appraisal sheet later.
    metadata_dict = pickleLoad('metadata_dict', folders, item_barcode)
    
    #if dictionary is empty, read info from spreadsheet; otherwise, retain dictionary
    if len(metadata_dict) == 0:
        
        status, current_row = return_spreadsheet_row(inv_ws, item_barcode)
        
        #if search status was false, then barcode isn't listed in spreadsheet.  Report error!
        if not status:
            print('\n\nError; barcode not found in spreadsheet.\n\nPlease review spreadsheet and correct barcode or add item to spreadsheet at %s.' % spreadsheet)
            return False
        
        ws_cols = get_spreadsheet_columns(inv_ws)
    
        for key in ws_cols.keys():
            if key == 'item_barcode':
                metadata_dict['item_barcode'] = item_barcode
            else:
                metadata_dict[key] = inv_ws.cell(row=current_row, column=ws_cols[key]).value
                
        #clean up any None values
        for val in metadata_dict:
            if metadata_dict[val] is None:
                metadata_dict[val] = '-'
       
    #Next, check if barcode has already been written to appraisal sheet
    status, current_row = return_spreadsheet_row(app_ws, item_barcode)    
    
    #if barcode has been written to the appraisal sheet...
    if status:
        
        #get our list of columns; if the 'note field' cell has content, add to the dictionary
        ws_cols = get_spreadsheet_columns(app_ws)
        
        for key in ws_cols.keys():
            if key == 'item_barcode':
                pass
            else:
                if not app_ws.cell(row=current_row, column=ws_cols[key]).value is None:
                    metadata_dict[key] = app_ws.cell(row=current_row, column=ws_cols[key]).value
           
    pickleDump('metadata_dict', metadata_dict, folders)

    return True

def metadata_to_gui(gui_vars, folders, item_barcode):
        
    #Find the barcode in the inventory sheet; save information to a dictionary so that it can be written to the Appraisal sheet later.
    metadata_dict = pickleLoad('metadata_dict', folders, item_barcode)
        
    gui_vars['collection_title'].set(metadata_dict['collection_title'])
    gui_vars['collection_creator'].set(metadata_dict['collection_creator'])
    gui_vars['content_source_type'].set(metadata_dict['content_source_type'])
    
    gui_vars['label_transcription'].configure(state='normal')
    gui_vars['label_transcription'].delete('1.0', END)
    gui_vars['label_transcription'].insert(INSERT, 'LABEL TRANSCRIPTION:\n\n%s' % metadata_dict['label_transcription'])
    
    gui_vars['bdpl_instructions'].configure(state='normal')
    gui_vars['bdpl_instructions'].delete('1.0', END)
    gui_vars['bdpl_instructions'].insert(INSERT, "TECHNICIAN NOTES:\n\n%s" % metadata_dict['bdpl_instructions'])
    gui_vars['bdpl_instructions'].configure(state='disabled')
    
    gui_vars['appraisal_notes'].configure(state='normal')
    gui_vars['appraisal_notes'].delete('1.0', END)
    gui_vars['appraisal_notes'].insert(INSERT, "APPRAISAL NOTES:\n\n%s" % metadata_dict['appraisal_notes'])
    gui_vars['appraisal_notes'].configure(state='disabled')
    
    try:
        if metadata_dict['technician_note'] == '-':
            notevalue = ''
        else:
            notevalue = metadata_dict['technician_note']
    except KeyError:
        notevalue = ''
        
    gui_vars['technician_note'].configure(state='normal')
    gui_vars['technician_note'].delete('1.0', END)
    gui_vars['technician_note'].insert(INSERT, notevalue)
    
def check_unfinished(unit_name, shipmentDate):
    if not verify_data(unit_name, shipmentDate, '', 'check_unfinished'):
        return   
    folders = bdpl_folders(unit_name, shipmentDate)
    ship_dir = folders['ship_dir']
    
    #loop through shipmnt directory/  Check for 'done.txt' file.  If not present, check premis.xml file and report what steps are completed.
    print('list:', [x for x in os.listdir(ship_dir) if os.path.isdir(x)])
    for item in [x for x in os.listdir(ship_dir) if os.path.isdir(os.path.join(ship_dir, x))]:
        if not os.path.exists(os.path.join(ship_dir, item, 'temp', 'done.txt')):
            
            premis_file = os.path.join(ship_dir, item, 'temp', 'premis_list.txt')
            if os.path.exists(premis_file):
                with open(premis_file, 'rb') as f:
                    premis_list = pickle.load(f)
            else:
                premis_list = []
                
            if len(premis_list) == 0:
                print('\nBarcode: %s' % item)
                print('\tItem folder structure has been created, but no ingest procedures have been completed.')
            else:
                print('\nBarcode: %s' % item)
                print('The following procedures have been completed:')
                print('\t%s' % '\n\t'.join(list(set((i['%s' % 'eventType'] for i in premis_list)))))

def check_progress(unit_name, shipmentDate):
    
    if not verify_data(unit_name, shipmentDate, '', 'check_progress'):
        return
    
    folders = bdpl_folders(unit_name, shipmentDate)
    ship_dir = folders['ship_dir']
    
    spreadsheet = find_spreadsheet(folders, unit_name, shipmentDate)    
    if os.path.exists(spreadsheet):
        wb = openpyxl.load_workbook(spreadsheet)
    else:
        return
    
    try:
        app_ws = wb['Appraisal']
    except KeyError:
        print('\n\nConsult with Digital Preservation Librarian; "Appraisal" worksheet does not exist.')
        return
    
    try:
        inv_ws = wb['Inventory']
    except KeyError:
        print('\n\nConsult with Digital Preservation Librarian; "Inventory" worksheet does not exist.')
        return    
    
    #get list of all barcodes on appraisal spreadsheet
    app_barcodes = []
    for col in app_ws['A'][1:]:
        if not col.value is None:
            app_barcodes.append(str(col.value))
    
    #get list of all barcodes on inventory spreadsheet
    inv_barcodes = {}
    for col in inv_ws['A'][1:]:
        if not col.value is None:
            inv_barcodes[str(col.value)] = col.row
    
    inv_list = list(inv_barcodes.keys())        
    
    #check to see if there are any duplicate barcodes in the inventory; print warning if so
    duplicate_barcodes = [item for item, count in Counter(inv_list).items() if count > 1]
    
    if duplicate_barcodes:
        print('\n\nWARNING! Inventory contains at least one duplicate barcode:')
        for dup in duplicate_barcodes:
            print('\t%s\tRow: %s' % (dup, inv_barcodes[dup]))
    
    current_total = len(inv_list) - len(app_barcodes)
    
    items_not_done = list(set(inv_list) - set(app_barcodes))
    
    print('\n\nCurrent status: %s out of %s items have been ingested. \n\n%s remain.' % (len(app_barcodes), len(inv_list), current_total))
    
    if len(items_not_done) > 0:
        print('\n\nThe following barcodes require ingest:\n%s' % '\n'.join(items_not_done))
    
    #reprint total if list is long...
    if len(items_not_done) > 15:    
        print('\n\nCurrent status: %s out of %s items have been ingested. \n\n%s remain.' % (len(app_barcodes), len(inv_list), current_total))


def move_media_images(unit_name, shipmentDate):
    
    folders = bdpl_folders(unit_name, shipmentDate)
    media_image_dir = folders['media_image_dir']
    unit_home = folders['unit_home']
    
    if unit_name == '':
        '\n\nError; please make sure you have entered a unit ID abbreviation.'
        return 
    
    if len(os.listdir(media_image_dir)) == 0:
        print('\n\nNo images of media at %s' % media_image_dir)
        return
    
    # get a list of barcodes in each shipment
    shipList = list(filter(lambda f: os.path.isdir(f), glob.glob('%s\\*\\*' % unit_home)))

    #list of files with no parent
    bad_file_list = []
    
    #loop through a list of all images in this folder; try to find match in list of barcodes; if not, add to 'bad file list'
    for f in os.listdir(media_image_dir):
        pic = f.split('-')[0]
        match = [s for s in shipList if pic in s]
        if len(match) > 0:
            media_pics = os.path.join(match[0], 'metadata', 'media-image')
            if not os.path.exists(media_pics):
                os.makedirs(media_pics)
            try:
                shutil.move(os.path.join(media_image_dir, f), media_pics)
            except shutil.Error as e:
                print('NOTE: ', e)
                print('\n\nCheck the media image folder to determine if a file already exists or a filename is being duplicated.')
                
        else:
            bad_file_list.append(f)
        
    if len(bad_file_list) > 0:
        print('\n\nFilenames for the following images do not match current barcodes:\n%s' % '\n'.join(bad_file_list))
        print('\nPlease correct filenames and try again.')
    else:
        print('\n\nMedia images successfully copied!')

def find_spreadsheet(folders, unit_name, shipmentDate):
    ship_dir = folders['ship_dir']
    
    #check to see if spreadsheet is present
    spreadsheet = [x for x in os.listdir(ship_dir) if x == '%s_%s.xlsx' % (unit_name, shipmentDate)]
    
    if len(spreadsheet) == 1: 
        workbook = os.path.join(ship_dir, spreadsheet[0])
        return workbook
    
    else:
        print('\n\nWARNING: spreadsheet not found; check shipment directory and make sure it is named according to the convention [unit]_[shipment].xlsx (e.g., "UAC_20190805.xlsx").')
        return 'ERROR'

def updateCombobox(unit_name, unit_shipment_date):
    
    if unit_name == '':
        comboList = []
    else:
        unit_home = os.path.join('Z:\\', unit_name, 'ingest')
        comboList = glob.glob1(unit_home, '*')
    
    unit_shipment_date['values'] = comboList

def update_software():
    #make sure PRONOM and antivirus signatures are up to date
    sfup = 'sf -update'
    fresh_up = 'freshclam'
    droid_up = 'droid -d'
    
    clam_sig1 = "C:/BDPL/resources/clamav/database/daily.cvd" 
    clam_sig2 = "C:/BDPL/resources/clamav/database/daily.cld"
    
    if os.path.exists(clam_sig1):
        clam_sig = clam_sig1
    elif os.path.exists(clam_sig2):
        clam_sig = clam_sig2
    else:
        subprocess.check_output(fresh_up, shell=True, text=True)
        clam_sig = "C:/BDPL/resources/clamav/database/daily.cvd"
        
    file_mod_time = datetime.datetime.fromtimestamp(os.stat(clam_sig).st_mtime).strftime('%Y%m%d')
    now = datetime.datetime.today().strftime('%Y%m%d')
    
    #if signature is older than today, run updates
    if now > file_mod_time:  
        print('\n\nUpdating PRONOM and antivirus signatures...')
        
        subprocess.check_output(sfup, shell=True, text=True)
        subprocess.check_output(fresh_up, shell=True, text=True)
        subprocess.check_output(droid_up, shell=True, text=True)
        
        print('\nUpdate complete!  Time to ingest some date...')

def main():
    
    #global window, source, jobType, unit, barcode, mediaStatus, source1, source2, source3, source4, source5, disk525, jobType1, jobType2, jobType3, jobType4, sourceDevice, barcodeEntry, sourceEntry, unitEntry, spreadsheet, coll_creator, coll_title, xfer_source, appraisal_notes, bdpl_instructions, noteField, label_transcription, shipmentDateList, noteFail, re_analyze, other_device
    
    update_software()
      
    window = Tk()
    window.title("Indiana University Library Born-Digital Preservation Lab")
    window.geometry('650x750')

    #if user tries to use 'X' button, make sure program closes correctly
    window.protocol('WM_DELETE_WINDOW', lambda: closeUp(window))

    '''
    
    GUI section for barcode, shipment, and unit info
    
    '''
    
    topFrame = Frame(window, width=650, height=50)
    topFrame.pack(fill=BOTH)    

    #Get unit name and barcode; provide dynamically updated drop-down to select shipment              
    barcode = StringVar()
    barcode.set('')           
    barcodeTxt = Label(topFrame, text="Barcode:")
    barcodeTxt.pack(in_=topFrame, side=LEFT, padx=(10,0), pady=10)
    barcodeEntry = Entry(topFrame, width=20, textvariable=barcode)
    barcodeEntry.pack(in_=topFrame, side=LEFT, padx=(0,10), pady=10)
    
    unit = StringVar()
    unit.set('')
    unitTxt = Label(topFrame, text="Unit:")
    unitTxt.pack(in_=topFrame, side=LEFT, padx=(10,0), pady=10)
    unitEntry = Entry(topFrame, width=5, textvariable=unit)
    unitEntry.pack(in_=topFrame, side=LEFT, padx=(0,10), pady=10)

    shipLabel = Label(topFrame, text="Shipment ID: ")
    shipLabel.pack(in_=topFrame, side=LEFT, padx=(10,0), pady=10)
    
    #User can either select an existng shipment date or add new one
    global unit_shipment_date
    unit_shipment_date = ttk.Combobox(topFrame, width=20, postcommand= lambda: updateCombobox(unit.get(), unit_shipment_date))
    unit_shipment_date.pack(in_=topFrame, side=LEFT, padx=(0,10), pady=10)
    
    #alternative approach: text entry
    # unit_shipment_date= StringVar()
    # unit_shipment_date.set('')
    # shipmentDateEntry = Entry(topFrame, width=10, textvariable=unit_shipment_date)
    # shipmentDateEntry.pack(in_=topFrame, side=LEFT, padx=5, pady=5)

    '''
    
    GUI section for job info
    
    '''

    middleFrame = Frame(window, width=650, height=150)
    middleFrame.pack(fill=BOTH)
    middleFrame.pack_propagate(False)
    
    '''
                UPPER MIDDLE
    '''
    
    upperMiddle = Frame(middleFrame, width=650, height=50)
    upperMiddle.pack(fill=BOTH)
    
    #job types: these determine which operations run on content
    jobTypeLabel = Label(upperMiddle, text="Job type:")
    jobTypeLabel.grid(column=0, row=1, padx=5, pady=5)

    jobType = StringVar()
    jobType.set(None)

    jobType1 = Radiobutton(upperMiddle, text='Copy only', value='Copy_only', variable=jobType)                     
    jobType1.grid(column=1, row=1, padx=15, pady=5)

    jobType2 = Radiobutton(upperMiddle, text='Disk image', value='Disk_image', variable=jobType)
    jobType2.grid(column=2, row=1, padx=15, pady=5)

    jobType3 = Radiobutton(upperMiddle, text='DVD', value='DVD', variable=jobType)
    jobType3.grid(column=3, row=1, padx=15, pady=5)
    
    jobType4 = Radiobutton(upperMiddle, text='CDDA', value='CDDA', variable=jobType)
    jobType4.grid(column=4, row=1, padx=15, pady=5)
    
    re_analyze = BooleanVar()
    re_analyze.set(False)
    re_analyzeChk = Checkbutton(upperMiddle, text='Re-analyze files', variable=re_analyze)
    re_analyzeChk.grid(column=5, row=1, padx=15, pady=5)
    
    '''
                MID MIDDLE
    '''
    midMiddle = Frame(middleFrame, width=650, height=25)
    midMiddle.pack(fill=BOTH)
    
    #Get path to source, if needed
    source = StringVar()
    source.set('')
    sourceTxt = Label(midMiddle, text='Source / file list\n("COPY" only): ')
    sourceTxt.pack(in_=midMiddle, side=LEFT, padx=5, pady=5)
    sourceEntry = Entry(midMiddle, width=55, textvariable=source)
    sourceEntry.pack(in_=midMiddle, side=LEFT, padx=5, pady=5)
    sourceBtn = Button(midMiddle, text="Browse", command= lambda: source_browse(window, source))
    sourceBtn.pack(in_=midMiddle, side=LEFT, padx=5, pady=5)
    
    '''
            LOWER MIDDLE
    '''
    lowerMiddle = Frame(middleFrame, width=650, height=100)
    lowerMiddle.pack(fill=BOTH)
    
    lowerMiddle1 = Frame(lowerMiddle, width=650, height=25)
    lowerMiddle1.pack(fill=BOTH)
    lowerMiddle2 = Frame(lowerMiddle, width=650, height=75)
    lowerMiddle2.pack(fill=BOTH)
     
    #Get source device, if needed
    sourceDevice = StringVar()
    sourceDevice.set(None)
    
    disk_type_options = ['N/A', 'Apple DOS 3.3 (16-sector)', 'Apple DOS 3.2 (13-sector)', 'Apple ProDOS', 'Commodore 1541', 'TI-99/4A 90k', 'TI-99/4A 180k', 'TI-99/4A 360k', 'Atari 810', 'MS-DOS 1200k', 'MS-DOS 360k', 'North Star MDS-A-D 175k', 'North Star MDS-A-D 350k', 'Kaypro 2 CP/M 2.2', 'Kaypro 4 CP/M 2.2', 'CalComp Vistagraphics 4500', 'PMC MicroMate', 'Tandy Color Computer Disk BASIC', 'Motorola VersaDOS']
    
    disk525 = StringVar()
    disk525.set('N/A')
            
    sourceDeviceLabel = Label(lowerMiddle1, text='Media:')
    sourceDeviceLabel.grid(column=0, row=0)
        
    source1 = Radiobutton(lowerMiddle1, text='CD/DVD', value='/dev/sr0', variable=sourceDevice)
    source2 = Radiobutton(lowerMiddle1, text='3.5" fd', value='/dev/fd0', variable=sourceDevice)
    source3 = Radiobutton(lowerMiddle1, text='5.25" fd', value='5.25', variable=sourceDevice)
    disk_menu = OptionMenu(lowerMiddle1, disk525, *disk_type_options)    
    source4 = Radiobutton(lowerMiddle1, text='Zip', value='Zip', variable=sourceDevice)
    source5 = Radiobutton(lowerMiddle1, text='Other', value='Other', variable=sourceDevice)
    
    other_device = StringVar()
    other_device.set('')
    other_deviceTxt = Label(lowerMiddle1, text="(& name)")
    other_deviceEntry = Entry(lowerMiddle1, width=5, textvariable=other_device)

    source1.grid(column=1, row=0, padx=5, pady=5)
    source2.grid(column=2, row=0, padx=5, pady=5)
    source3.grid(column=3, row=0, padx=5, pady=5)
    disk_menu.grid(column=4, row=0, padx=5, pady=5)
    source4.grid(column=5, row=0, padx=5, pady=5)
    source5.grid(column=6, row=0, padx=(5,0), pady=5)
    other_deviceTxt.grid(column=7, row=0, padx=(0,5), pady=5)
    other_deviceEntry.grid(column=8, row=0, padx=5, pady=5)

    
    #buttons: kick off various functions    
    newBtn = Button(lowerMiddle2, text="New", bg='light slate gray', width = 8, command= lambda: cleanUp(cleanUp_vars))
    newBtn.grid(column=0, row=2, padx=(30,20), pady=5)

    createBtn = Button(lowerMiddle2, text="Load", bg='light slate gray', width = 8, command= lambda: first_run(unit.get(), unit_shipment_date.get(), barcode.get(), gui_vars))
    createBtn.grid(column=1, row=2, padx=20, pady=5)

    transferBtn = Button(lowerMiddle2, text="Transfer", bg='light slate gray', width = 8, command= lambda: transferContent(unit.get(), unit_shipment_date.get(), barcode.get(), transfer_vars))
    transferBtn.grid(column=2, row=2, padx=20, pady=5)

    analyzeBtn = Button(lowerMiddle2, text="Analyze", bg='light slate gray', width = 8, command= lambda: analyzeContent(unit.get(), unit_shipment_date.get(), barcode.get(), analysis_vars))
    analyzeBtn.grid(column=3, row=2, padx=20, pady=5)
        
    closeBtn = Button(lowerMiddle2, text="Quit", bg='light slate gray', width = 8, command= lambda: closeUp(window))
    closeBtn.grid(column=4, row=2, padx=20, pady=5)

    mediaStatus = BooleanVar()
    mediaStatus.set(False)
    mediaStatusChk = Checkbutton(lowerMiddle2, text="Attached?", variable=mediaStatus)
    mediaStatusChk.grid(column=5, row=2, padx=(10, 20), pady=5)
    
    '''
    
    GUI section for BDPL technician note
    
    '''
    noteFrame = Frame(window, width=650, height=40)
    noteFrame.pack(fill=BOTH)
    
    noteLabel = Label(noteFrame, text="BDPL\nnote:", anchor='w')
    noteLabel.grid(row=1, column=0, pady=10)
    
    noteScroll = Scrollbar(noteFrame)
    noteField = Text(noteFrame, height=3)
    noteScroll.config(command=noteField.yview)
    noteField.config(yscrollcommand=noteScroll.set)
    
    noteField.grid(row=1, column=1, sticky="nsew", padx=(10, 0), pady=10)
    noteFrame.grid_rowconfigure(1, weight=1)
    noteFrame.grid_columnconfigure(1, weight=1)
    
    noteScroll.grid(row=1, column=2, padx=(0, 10), pady=(10, 0), sticky=NS)
    
    noteSave = Button(noteFrame, text="Save\nnote", width=5, command= lambda: writeNote(unit.get(), unit_shipment_date.get(), barcode.get(), gui_vars))
    noteSave.grid(row=1, column=3, padx=10)
    
    noteFail = BooleanVar()
    noteFail.set(False)
    noteFailChk = Checkbutton(noteFrame, text="Record failed transfer with note", variable=noteFail)
    noteFailChk.grid(row=2, column=1, pady=(0, 10))
    
    '''
    GUI section for additional actions/features
    '''
    bottomFrame = Frame(window, width=650, height=50)
    bottomFrame.pack(fill=BOTH)
    bottomFrame.pack_propagate(False)
    
    check_spreadsheet = Button(bottomFrame, text="Check spreadsheet", width = 20, command= lambda: check_progress(unit.get(), unit_shipment_date.get()))
    check_spreadsheet.grid(row=0, column=0, padx=30)
    
    move_pics = Button(bottomFrame, text="Move media images", width = 20, command= lambda: move_media_images(unit.get(), unit_shipment_date.get()))
    move_pics.grid(row=0, column=1, padx=30)
    
    unfinished_check = Button(bottomFrame, text="Check unfinished", width = 20, command= lambda: check_unfinished(unit.get(), unit_shipment_date.get()))
    unfinished_check.grid(row=0, column=2, padx=30)
    
    '''
    GUI section with metadata      
    '''
    
    borderFrame = Frame(window, width=650, height=5, bg='black')
    borderFrame.pack(fill=BOTH, padx=10, pady=10)
    borderLabel = Label(borderFrame, text="Information about transfer:")
    borderLabel.pack()
    borderLabel.config(fg='white', bg='black')
    
    inventoryFrame = Frame(window, width=650, height=300)
    inventoryFrame.pack(fill=BOTH)
    
    inventoryTop = Frame(inventoryFrame, width=650, height=50)
    inventoryTop.pack(fill=BOTH)
    #inventoryTop.pack_propagate(0)
    inventoryTop.grid_columnconfigure(1, weight=1)
    
    inventoryBottom = Frame(inventoryFrame, width=650, height=250)
    inventoryBottom.pack(fill=BOTH)
    #inventoryBottom.pack_propagate(0)
    #inventoryBottom.grid_columnconfigure(0, weight=1)
    
    #pull in information from spreadsheet so tech can see what's going on
    coll_title = StringVar()
    coll_title_Label = Label(inventoryTop, text="Coll.\ntitle:")
    coll_title_Display = Label(inventoryTop, wraplength=250, justify=LEFT, textvariable=coll_title)
    coll_title_Label.grid(row=0, column=0, padx=5)
    coll_title_Display.grid(row=0, column=1, padx=5, sticky='w')
    
    coll_creator = StringVar()
    coll_creator_Label = Label(inventoryTop, text="Creator:")
    coll_creator_Display = Label(inventoryTop, wraplength=250, justify=LEFT, textvariable=coll_creator)
    coll_creator_Label.grid(row=1, column=0, padx=5)
    coll_creator_Display.grid(row=1, column=1, padx=5, sticky='w')

    xfer_source = StringVar()
    xfer_source_Label = Label(inventoryTop, text="Source:")
    xfer_source_Display = Label(inventoryTop, textvariable=xfer_source)
    xfer_source_Label.grid(row=2, column=0, padx=5)
    xfer_source_Display.grid(row=2, column=1, padx=5, sticky='w')   
    
    #some larger fields with potential for more text   
    appraisal_notes = Text(inventoryBottom, height=4, width=70)
    appraisal_scroll = Scrollbar(inventoryBottom)
    appraisal_scroll.config(command=appraisal_notes.yview)
    appraisal_notes.config(yscrollcommand=appraisal_scroll.set)
    appraisal_notes.insert(INSERT, "APPRAISAL NOTES:\n")
    appraisal_notes.grid(row=0, column=0, pady=5, padx=(5,0))
    appraisal_scroll.grid(row=0, column=1, pady=5, sticky='ns')
    appraisal_notes.configure(state='disabled')
    
    label_transcription = Text(inventoryBottom, height=4, width=70)
    label_scroll = Scrollbar(inventoryBottom)
    label_scroll.config(command=label_transcription.yview)
    label_transcription.config(yscrollcommand=label_scroll.set)
    label_transcription.insert(INSERT, "LABEL TRANSCRIPTION:\n")
    label_transcription.grid(row=1, column=0, pady=5, padx=(5,0))
    label_scroll.grid(row=1, column=1, pady=5, sticky='ns')
    #label_transcription.configure(state='disabled')
    
    bdpl_instructions = Text(inventoryBottom, height=4, width=70)
    bdpl_scroll = Scrollbar(inventoryBottom)
    bdpl_scroll.config(command=bdpl_instructions.yview)
    bdpl_instructions.config(yscrollcommand=bdpl_scroll.set)
    bdpl_instructions.insert(INSERT, "TECHNICIAN NOTES:\n")
    bdpl_instructions.grid(row=2, column=0, pady=5, padx=(5,0))
    bdpl_scroll.grid(row=2, column=1, pady=5, sticky='ns')
    bdpl_instructions.configure(state='disabled')
    
    '''Variables for main functions'''
     
    gui_vars = {'collection_creator' : coll_creator, 'collection_title' : coll_title, 'content_source_type' : xfer_source, 'bdpl_instructions' : bdpl_instructions, 'appraisal_notes' : appraisal_notes, 'label_transcription': label_transcription, 'technician_note': noteField, 'noteFail' : noteFail, 'platform' : 'bdpl_ingest'}
    
    transfer_vars = {'platform' : 'bdpl_ingest', 'jobType' : jobType, 'sourceDevice' : sourceDevice, 'source' : source, 'other_device' : other_device, 'disk525' : disk525, 'mediaStatus' : mediaStatus}
    
    analysis_vars = {'platform' : 'bdpl_ingest', 'jobType' : jobType, 're_analyze' : re_analyze, 'gui_vars' : {'collection_creator' : coll_creator, 'collection_title' : coll_title, 'content_source_type' : xfer_source, 'bdpl_instructions' : bdpl_instructions, 'appraisal_notes' : appraisal_notes, 'label_transcription': label_transcription, 'technician_note': noteField}}
    
    cleanUp_vars = {'radio_buttons' : {'jobType1' : jobType1, 'jobType2': jobType2, 'jobType3': jobType3, 'jobType4' : jobType4, 'source1' : source1, 'source2' : source2, 'source3' : source3, 'source4' : source4, 'source5' : source5}, 'str_vars' : {'jobType' : jobType, 'sourceDevice' : sourceDevice, 'barcode' : barcode, 'source' : source, 'other_device' : other_device, 'disk525' : disk525, 'collection_creator' : coll_creator, 'collection_title' : coll_title, 'content_source_type' : xfer_source}, 'checkboxes' : {'mediaStatus' : mediaStatus, 'noteFail' : noteFail, 're_analyze' : re_analyze}, 'entry_widgets' : {'barcodeEntry' : barcodeEntry, 'sourceEntry' : sourceEntry}, 'text_widgets' : {'bdpl_instructions' : bdpl_instructions, 'appraisal_notes' : appraisal_notes, 'label_transcription': label_transcription, 'technician_note': noteField}}
    
    window.mainloop()

def newscreen():
    os.system('cls')
    
    #print BDPL screen
    fname = "C:/BDPL/scripts/bdpl.txt"
    if os.path.exists(fname):
        with open(fname, 'r') as fin:
            print(fin.read())
            print('\n')
    else:
        print('Missing ASCII art header file; download to: %s' % fname)
        
def source_browse(window, source):
    currdir = "Z:\\"
    selected_dir = tkinter.filedialog.askdirectory(parent=window, initialdir=currdir, title='Please select the source directory')
    if len(selected_dir) > 0:
        source.set(selected_dir)
        

if __name__ == '__main__':
    main()
