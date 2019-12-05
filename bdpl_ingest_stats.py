'''
script to pull stats from BDPL spreadsheets currently in process
'''

import openpyxl
import os
import glob
import collections
import math
import subprocess
import shutil
import datetime

from bdpl_ingest import get_spreadsheet_columns

def convert_size(size):
    # convert size to human-readable form
    if (size == 0):
        return '0 bytes'
    size_name = ("bytes", "KB", "MB", "GB", "TB", "PB", "EB", "ZB", "YB")
    i = int(math.floor(math.log(size,1024)))
    p = math.pow(1024,i)
    s = round(size/p)
    s = str(s)
    s = s.replace('.0', '')
    return '%s %s' % (s,size_name[i])

def main():
    
    '''GET STATS FROM CURRENT SHIPMENTS--CONTENT THAT IS READY FOR INGEST BUT NOT DEPOSITED TO SCANDIUM'''
    print('\nGetting stats for current work...')
    
    #add any other temp/work folders, if they are present
    working_dirs = ['TEST', 'media-images', 'bdpl_transfer_lists', 'Ripstation']
    
    #assumes that bdpl\workspace is mapped to Z:\ drive; change if needed. Get a list of projects from that location. 
    workspace = 'Z:\\'
    projects = os.listdir('Z:/')
    
    #remove any working_dirs from list
    projects = [dir for dir in projects if dir not in working_dirs]
    
    #specify a location for output file
    output = "C:/BDPL/bdpl_stats.txt"
    if os.path.exists(output):
        os.remove(output)

    stats = {'total_size' : 0, 'total_files' : 0, 'total_items' : 0}
    
    with open(output, 'w') as f:
        f.write('BDPL STATISTICS\nPrepared on: %s\n\nCURRENT SHIPMENTS:\n' % datetime.date.today().strftime("%B %d, %Y"))
    
    for project in projects:
        project_added = False
        
        home_dir = os.path.join(workspace, project, 'ingest')
        
        if not os.path.exists(home_dir):
            continue
        
        stats[project] = []
        
        for shipment in [s for s in os.listdir(home_dir) if os.path.isdir(os.path.join(home_dir, s))]:

            temp = {}
                       
            if glob.glob(os.path.join(home_dir, shipment, '*.xlsx')):
                spreadsheet = glob.glob(os.path.join(home_dir, shipment, '*.xlsx'))[0]
            else:
                continue
            
            spreadsheet_name = os.path.basename(spreadsheet)
            
            if os.path.exists(os.path.join(workspace, project, 'completed_shipments', spreadsheet_name)):
                continue
            
            wb = openpyxl.load_workbook(spreadsheet)
            
            ws = wb['Appraisal']
            
            #get columns and adjust to work with iterrows
            ws_columns = get_spreadsheet_columns(ws)
            
            for key in ws_columns.keys():
                ws_columns[key] = ws_columns[key] - 1
            
            temp['shipment'] = shipment
            temp['raw_formats'] = []
            temp['size'] = []
            temp['item_count'] = 0
            temp['file_count'] = 0
            
            iterrows = ws.iter_rows()
            next(iterrows)
            
            for row in iterrows:
                if row[ws_columns['migration_outcome']].value == 'Success':
                    temp['item_count'] += 1
                    temp['file_count'] += int(row[ws_columns['item_file_count']].value)
                    temp['raw_formats'].append(row[ws_columns['content_source_type']].value.split(' (')[0].lower().replace(' ', '').replace('?', ''))
                    temp['size'].append(row[ws_columns['extent_normal']].value)

            #tally our sizes
            final_size = 0
            for unit in ["bytes", "KB", "MB", "GB", "TB"]:
                subgroup = [x for x in temp['size'] if unit in x]
                if len(subgroup) > 0:
                    for item in subgroup:
                        if 'bytes' in item:
                            volume = int(item.split(' ')[0])
                        elif 'KB' in item:
                            volume = int(item.split(' ')[0]) * 1000
                        elif 'MB' in item:
                            volume = int(item.split(' ')[0]) * 1000000
                        elif 'GB' in item:
                            volume = int(item.split(' ')[0]) * 1000000000
                        elif 'GB' in item:
                            volume = int(item.split(' ')[0]) * 1000000000000
                        final_size += volume
            
            temp['size'] = final_size
            
            format_count = collections.Counter(temp['raw_formats'])
            temp['final_formats'] = dict(format_count)
            
            if temp['item_count'] > 0:
                if not project_added:
                    with open(output, 'a') as f:
                        f.write('%s\n' % project)
                    project_added = True
                    
                with open(output, 'a') as f:
                    f.write('\n\tShipment: %s\n' % temp['shipment'])
                    f.write('\t\tItem count: %s\n' % temp['item_count'])
                    f.write('\t\tFile count: %s\n' % temp['file_count'])
                    f.write('\t\tSize (in bytes): %s (%s)\n' % (temp['size'], convert_size(temp['size'])))
                    f.write('\t\tSource formats:\n')
                    for key, value in temp['final_formats'].items():
                        f.write('\t\t\t%s: %s\n' % (key, value))
                  
            stats[project].append(temp)
            
        total_unit_items = 0
        total_unit_count = 0
        total_size = 0
        total_formats = []
        
        for totals in stats[project]:
            total_unit_count += totals['file_count']
            total_unit_items += totals['item_count']
            total_formats = total_formats + totals['raw_formats']
            total_size += totals['size']
        
        format_count = collections.Counter(total_formats)
        
        if total_unit_items > 0:
            with open(output, 'a') as f:
                f.write('\n\t%s TOTALS:\n' % project)
                f.write('\t\tTotal items: %s\n' % total_unit_items)
                f.write('\t\tTotal files: %s\n' % total_unit_count)
                f.write('\t\tTotal size: %s (%s)\n' % (total_size, convert_size(total_size)))
                f.write('\t\tTotal format tallies:\n')
                for key, value in dict(format_count).items():
                    f.write('\t\t\t%s: %s\n' % (key, value))
        
        stats['total_files'] += total_unit_count
        stats['total_items'] += total_unit_items
        stats['total_size'] += total_size
        
    with open(output, 'a') as f:
        f.write('\n\nGRAND TOTALS FOR CURRENT WORK:\n')
        f.write('\tItems: %s\n' % stats['total_items'])
        f.write('\tFiles: %s\n' % stats['total_files'])
        f.write('\tSize: %s (%s)\n' % (stats['total_size'], convert_size(stats['total_size'])))
    
    '''NEXT, GET STATS ON CONTENT DEPOSITED TO SDA'''
    print('\nCollecting statistics for content deposited to SDA...')

    if not os.path.exists('W:/spreadsheets/bdpl_master_spreadsheet.xlsx'):
        book = input('\nPath to master spreadsheet: ')
    else:
        book = 'W:/spreadsheets/bdpl_master_spreadsheet.xlsx'
    
    spreadsheet_copy = os.path.join('C:/BDPL/', '%s_COPY.xlsx' % os.path.basename(book))
    
    shutil.copy(book, spreadsheet_copy)

    wb = openpyxl.load_workbook(spreadsheet_copy)

    ws_master_all = wb['Cumulative']

    iterrows = ws_master_all.iter_rows()

    next(iterrows)

    master_stats = {}

    with open(output, 'a') as f:
        f.write('\n\nCONTENT DEPOSITED TO SDA:\n')
        for row in iterrows:
            unit = row[0].value.split()[0]
            if not unit in master_stats.keys():
                master_stats[unit] = {'count' : 1, 'items' : int(row[2].value), 'size' : int(row[5].value)}
            else:
                master_stats[unit]['count'] += 1
                master_stats[unit]['items'] += int(row[2].value)
                master_stats[unit]['size'] += int(row[5].value)
        
        unit_totals = {}
        for key, value in master_stats.items():
            sized = convert_size(value['size'])
            unit_totals[key] = {'items' : value['items'], 'size' : sized}
            
            
        ws_master_item = wb['Item']
        
        iterrows2 = ws_master_item.iter_rows()
        next(iterrows2)
        
        stats_items = {}
        by_year = {}
        for row in iterrows2:
            unit = row[1].value
            year = str(row[14].value)[:4]
            
            if year not in by_year.keys():
                by_year[year] = [[int(row[17].value)], [1]]
            else:
                by_year[year][0].append(int(row[17].value))
                by_year[year][1].append(1)
            
            if not unit in stats_items.keys():
                stats_items[unit] = {year : {'items' : 1, 'size' : int(row[17].value)}}
            else:
                if not year in stats_items[unit].keys():
                    stats_items[unit][year] = {'items' : 1, 'size' : int(row[17].value)}
                else:
                    
                    stats_items[unit][year]['items'] += 1
                    stats_items[unit][year]['size'] += int(row[17].value)
        
        for unit, data in stats_items.items():
            f.write('%s\n' % unit)
            for year, info in sorted(data.items()):
                sized = convert_size(info['size'])
                
                f.write('\t%s\n\t\tNumber of items: %s\n\t\tOverall size: %s\n' % (year, info['items'], sized))
            
            f.write('\tTOTAL:\n\t\tNumber of items: %s\n\t\tOverall size: %s\n' % (unit_totals[unit]['items'], unit_totals[unit]['size']))

        f.write('\n\n')
        
        for key, values in sorted(by_year.items()):
            f.write('%s : %s (%s items)\n' % (key, convert_size(sum(values[0])), sum(values[1])))
        
    print('\nText file with these statistics located at: %s' % output)
    
    os.remove(spreadsheet_copy)
    
    cmd = 'notepad %s' % output
    subprocess.call(cmd)

if __name__ == '__main__':
    main()       
                       
        